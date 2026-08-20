"""Construcción de los reportes exportables: HU-19 (PDF/Excel) y HU-20 (base
consolidada en Excel/CSV).

Funciones puras: reciben los datos ya armados —el diccionario de
`Encuesta.resumen_para_reporte()`, o la lista de diccionarios de
`Integrante.base_consolidada()`— y devuelven el archivo construido. No
conocen `request` ni `HttpResponse` — la vista es quien decide el
`Content-Type` y el nombre del archivo —, así que se prueban directamente,
sin simular HTTP. Es el mismo criterio que ya aplican `usuarios/seguridad.py`
y `usuarios/auditoria.py`: la lógica que no depende de la petición no vive en
la vista.

Los bloques del reporte de resultados —resumen general, por estado, por
sector, por censista— se arman con la misma descripción (`BLOQUES`) en los
dos formatos, para que agregar un bloque el día de mañana sea una entrada más
en esa lista y no un bloque de código copiado entre el Excel y el PDF. La
base consolidada usa el mismo recurso con `COLUMNAS_BASE_CONSOLIDADA`.
"""

import csv

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

TITULO_REPORTE = "Reporte de resultados del operativo"
FORMATO_FECHA = "%d-%m-%Y %H:%M"

# Cada bloque describe sus columnas y cómo sacarle una fila a un elemento del
# resumen. `ancho_pdf` está en centímetros y solo lo usa el PDF; el Excel
# fija el ancho de columna una sola vez, al final, por letra.
BLOQUES = (
    {
        "titulo": "Por estado",
        "clave": "por_estado",
        "columnas": ("Estado", "Total", "%"),
        "fila": lambda f: (f["etiqueta"], f["total"], f"{f['porcentaje']}%"),
        "ancho_pdf": (8 * cm, 3 * cm, 2 * cm),
    },
    {
        "titulo": "Por sector",
        "clave": "por_sector",
        "columnas": ("Sector", "Comuna", "Total", "%"),
        "fila": lambda f: (f["sector"], f["comuna"], f["total"], f"{f['porcentaje']}%"),
        "ancho_pdf": (5 * cm, 5 * cm, 2 * cm, 1.5 * cm),
    },
    {
        "titulo": "Por censista",
        "clave": "por_censista",
        "columnas": ("Censista", "Total", "Validadas", "Observadas"),
        "fila": lambda f: (f["censista"], f["total"], f["validadas"], f["observadas"]),
        "ancho_pdf": (6 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm),
    },
)


def _filas_resumen_general(resumen):
    """El bloque de arriba: el total y las tres alertas de calidad (HU-16).

    Reutiliza `resumen_alertas_calidad()` tal cual: el reporte no inventa un
    segundo criterio de "ficha con problemas" distinto del que ya usa el panel
    del supervisor.
    """
    alertas = resumen["alertas"]
    return [
        ("Total de encuestas", resumen["total"]),
        ("Con datos incompletos", alertas["datos_incompletos"]),
        ("Esperando revisión hace demasiado", alertas["espera_prolongada"]),
        ("Visitas vencidas", alertas["visitas_vencidas"]),
    ]


def construir_reporte_excel(resumen):
    """Arma el libro .xlsx con el resumen general y los bloques de `BLOQUES`."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Resultados"

    hoja.append([TITULO_REPORTE])
    hoja.cell(row=1, column=1).font = Font(bold=True, size=14)
    hoja.append([f"Generado el {resumen['generado_en']:{FORMATO_FECHA}}"])
    hoja.append([])

    hoja.append(["Resumen general"])
    hoja.cell(row=hoja.max_row, column=1).font = Font(bold=True)
    hoja.append(["Indicador", "Total"])
    for etiqueta, total in _filas_resumen_general(resumen):
        hoja.append([etiqueta, total])
    hoja.append([])

    for bloque in BLOQUES:
        hoja.append([bloque["titulo"]])
        hoja.cell(row=hoja.max_row, column=1).font = Font(bold=True)
        hoja.append(list(bloque["columnas"]))

        filas = resumen[bloque["clave"]]
        if filas:
            for elemento in filas:
                hoja.append(list(bloque["fila"](elemento)))
        else:
            hoja.append(["Sin datos para este filtro."])
        hoja.append([])

    for columna, ancho in zip("ABCD", (34, 20, 14, 14)):
        hoja.column_dimensions[columna].width = ancho

    return libro


def construir_reporte_pdf(buffer, resumen):
    """Escribe en `buffer` el PDF con el resumen general y los bloques de `BLOQUES`.

    `buffer` es cualquier objeto con `.write()` — un `HttpResponse` calza sin
    envoltorio adicional, y así lo usa `ExportarReportePDFView`.
    """
    documento = SimpleDocTemplate(buffer, pagesize=letter)
    estilos = getSampleStyleSheet()

    elementos = [
        Paragraph(TITULO_REPORTE, estilos["Title"]),
        Paragraph(
            f"Generado el {resumen['generado_en']:{FORMATO_FECHA}}", estilos["Normal"]
        ),
        Spacer(1, 14),
        Paragraph("Resumen general", estilos["Heading2"]),
        _tabla_pdf(
            ("Indicador", "Total"),
            _filas_resumen_general(resumen),
            (10 * cm, 3 * cm),
        ),
        Spacer(1, 14),
    ]

    for bloque in BLOQUES:
        elementos.append(Paragraph(bloque["titulo"], estilos["Heading2"]))
        filas = [bloque["fila"](elemento) for elemento in resumen[bloque["clave"]]]
        elementos.append(_tabla_pdf(bloque["columnas"], filas, bloque["ancho_pdf"]))
        elementos.append(Spacer(1, 14))

    documento.build(elementos)


def _tabla_pdf(columnas, filas, anchos):
    """Una tabla con encabezado, o un único aviso si el bloque no tiene filas.

    Un bloque vacío es un resultado válido —el filtro pudo no dejar ninguna
    encuesta de un sector, por ejemplo— y el PDF lo dice en vez de mostrar una
    tabla con solo el encabezado, que se leería como un error de armado.
    """
    if not filas:
        return Paragraph("Sin datos para este filtro.", getSampleStyleSheet()["Italic"])

    datos = [list(columnas), *[list(fila) for fila in filas]]
    tabla = Table(datos, colWidths=list(anchos))
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d3b66")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    return tabla


# ==========================================================================
# HU-20 — BASE CONSOLIDADA (una fila por persona)
# ==========================================================================

# (clave del diccionario de `Integrante.base_consolidada()`, encabezado de columna).
# La misma lista arma el Excel y el CSV, para que agregar una columna el día de
# mañana sea una entrada más aquí y no un cambio en dos formatos por separado.
COLUMNAS_BASE_CONSOLIDADA = (
    ("operativo", "Operativo"),
    ("region", "Región"),
    ("comuna", "Comuna"),
    ("sector", "Sector"),
    ("zona", "Zona"),
    ("direccion", "Dirección"),
    ("referencia", "Referencia"),
    ("tipo_vivienda", "Tipo de vivienda"),
    ("tenencia", "Tenencia"),
    ("materialidad_muros", "Materialidad de los muros"),
    ("origen_agua", "Origen del agua"),
    ("sistema_sanitario", "Sistema sanitario"),
    ("tiene_electricidad", "Tiene electricidad"),
    ("latitud", "Latitud"),
    ("longitud", "Longitud"),
    ("jefe_hogar_nombre", "Jefe de hogar"),
    ("jefe_hogar_rut", "RUT jefe de hogar"),
    ("telefono_contacto", "Teléfono de contacto"),
    ("integrantes_declarados", "Personas declaradas"),
    ("ingreso_mensual", "Ingreso mensual del hogar"),
    ("nombres", "Nombres"),
    ("apellidos", "Apellidos"),
    ("rut", "RUT"),
    ("parentesco", "Parentesco"),
    ("sexo", "Sexo"),
    ("fecha_nacimiento", "Fecha de nacimiento"),
    ("edad", "Edad"),
    ("nivel_educacional", "Nivel educacional"),
    ("situacion_ocupacional", "Situación ocupacional"),
    ("pueblo_originario", "Pueblo originario"),
    ("tiene_discapacidad", "Presenta discapacidad"),
    ("estado_encuesta", "Estado de la encuesta"),
    ("cerrada_en", "Fecha de cierre"),
)


def _tabla_base_consolidada(filas):
    """Los encabezados y el cuerpo, en el mismo orden, listos para escribir."""
    claves = [clave for clave, _ in COLUMNAS_BASE_CONSOLIDADA]
    encabezados = [etiqueta for _, etiqueta in COLUMNAS_BASE_CONSOLIDADA]
    cuerpo = [[fila[clave] for clave in claves] for fila in filas]
    return encabezados, cuerpo


def construir_base_excel(filas):
    """Arma el libro .xlsx de la base consolidada: una fila por persona.

    A diferencia de `construir_reporte_excel()`, no hay bloques ni resumen:
    es una tabla ancha y plana, el formato que un análisis externo espera.
    """
    encabezados, cuerpo = _tabla_base_consolidada(filas)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Base consolidada"

    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True)
    hoja.freeze_panes = "A2"  # el encabezado queda fijo al desplazarse

    for fila in cuerpo:
        hoja.append(fila)

    return libro


def construir_base_csv(buffer, filas):
    """Escribe en `buffer` la base consolidada en CSV: una fila por persona.

    `buffer` es cualquier objeto con `.write()` — un `HttpResponse` calza sin
    envoltorio adicional, y así lo usa `ExportarBaseCSVView`.
    """
    encabezados, cuerpo = _tabla_base_consolidada(filas)

    escritor = csv.writer(buffer)
    escritor.writerow(encabezados)
    escritor.writerows(cuerpo)
