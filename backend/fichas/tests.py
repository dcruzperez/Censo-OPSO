"""Pruebas automáticas de la HU-07 «Visualizar las encuestas asignadas y su estado».

    python manage.py test fichas

Cada prueba sigue el patrón PREPARAR -> ACTUAR -> VERIFICAR y su nombre describe la
regla que comprueba, para que la salida del comando se lea como una lista de
requisitos cumplidos.

Las cuatro cosas que esta batería vigila con más insistencia, porque son las que
romperían la historia sin dar ningún error visible:

  1. Que nadie vea las encuestas de otra persona.
  2. Que el estado y sus dos fechas no puedan contradecirse.
  3. Que la pantalla ordene por urgencia y no por casualidad.
  4. Que el coste en consultas no crezca con el número de encuestas.
"""

import shutil
import tempfile
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError, connection, transaction
from django.db.models import ProtectedError
from django.test import TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from operativos.models import (
    AsignacionSector,
    Comuna,
    EstadoOperativo,
    Operativo,
    Region,
    Sector,
    Zona,
)
from usuarios.models import Permiso, Rol, RolCodigo, Usuario

from .forms import (
    AnularEncuestaForm,
    BorradorForm,
    CerrarSinDatosForm,
    DevolverEncuestaForm,
    FiltroMisEncuestasForm,
    FiltroRevisionForm,
    FotografiaForm,
    GrupoFamiliarForm,
    IntegranteForm,
    UbicacionForm,
    ValidarEncuestaForm,
    ViviendaForm,
    zonas_disponibles,
)
from .models import (
    ESTADOS_ABIERTOS,
    ESTADOS_CERRADOS,
    ESTADOS_RESUELTOS,
    ESTADOS_SIN_LEVANTAR,
    Encuesta,
    EstadoEncuesta,
    Fotografia,
    GrupoFamiliar,
    Integrante,
    MaterialidadMuros,
    NivelEducacional,
    OrigenAgua,
    Parentesco,
    PuebloOriginario,
    Sexo,
    SistemaSanitario,
    SituacionOcupacional,
    TenenciaVivienda,
    TipoFotografia,
    TipoVivienda,
    Vivienda,
)
from .reportes import construir_reporte_excel, construir_reporte_pdf

CLAVE_VALIDA = "Censo2026#Opso"


class BaseEncuestaTest(TestCase):
    """Escenario común: un operativo en curso con un sector de dos zonas.

    Marta es la encuestadora de la historia; Juan existe para comprobar que no ve
    lo de Marta ni Marta lo suyo.
    """

    @classmethod
    def setUpTestData(cls):
        cls.rol_admin = Rol.objects.get(codigo=RolCodigo.ADMINISTRADOR)
        cls.rol_supervisor = Rol.objects.get(codigo=RolCodigo.SUPERVISOR)
        cls.rol_censista = Rol.objects.get(codigo=RolCodigo.CENSISTA)

        cls.admin = Usuario.objects.create_user(
            email="admin@opso.cl",
            password=CLAVE_VALIDA,
            first_name="Ana",
            last_name="Rojas",
            rol=cls.rol_admin,
        )
        cls.supervisor = Usuario.objects.create_user(
            email="supervisor@opso.cl",
            password=CLAVE_VALIDA,
            first_name="Luis",
            last_name="Pérez",
            rol=cls.rol_supervisor,
        )
        cls.marta = Usuario.objects.create_user(
            email="marta@opso.cl",
            password=CLAVE_VALIDA,
            first_name="Marta",
            last_name="Soto",
            rol=cls.rol_censista,
        )
        cls.juan = Usuario.objects.create_user(
            email="juan@opso.cl",
            password=CLAVE_VALIDA,
            first_name="Juan",
            last_name="Vera",
            rol=cls.rol_censista,
        )
        cls.sin_rol = Usuario.objects.create_user(
            email="sinrol@opso.cl",
            password=CLAVE_VALIDA,
            first_name="Sin",
            last_name="Rol",
        )
        cls.biobio = Region.objects.get(codigo="08")

    def setUp(self):
        self.concepcion = Comuna.objects.create(region=self.biobio, nombre="Concepción")
        self.operativo = Operativo.objects.create(
            nombre="Censo Social 2026",
            fecha_inicio=date(2026, 3, 1),
            fecha_termino=date(2026, 3, 31),
            estado=EstadoOperativo.EN_CURSO,
        )
        self.boldos = Sector.objects.create(
            operativo=self.operativo, comuna=self.concepcion, nombre="Los Boldos"
        )
        self.norte = Sector.objects.create(
            operativo=self.operativo, comuna=self.concepcion, nombre="Barrio Norte"
        )
        self.zona1 = Zona.objects.create(
            sector=self.boldos, nombre="Zona 1", viviendas_estimadas=20
        )
        self.zona2 = Zona.objects.create(
            sector=self.boldos, nombre="Zona 2", viviendas_estimadas=10
        )
        self.zona_norte = Zona.objects.create(sector=self.norte, nombre="Zona única")

        self.url_lista = reverse("fichas:mis_encuestas")

    # -- ayudantes ---------------------------------------------------------

    def crear_vivienda(self, direccion="Av. Central 100", zona=None, **extra):
        """Una vivienda, descrita por defecto.

        Se describe por defecto porque lo contrario —una vivienda a medias— es el
        caso excepcional que solo interesa a las pruebas que lo comprueban, y
        dejarlo como norma haría que cualquier prueba de otra cosa arrastrara un
        dato incompleto sin quererlo.
        """
        datos = {
            "tipo": TipoVivienda.CASA,
            "tenencia": TenenciaVivienda.ARRENDADA,
            "materialidad_muros": MaterialidadMuros.ALBANILERIA,
            "origen_agua": OrigenAgua.RED_PUBLICA,
            "sistema_sanitario": SistemaSanitario.ALCANTARILLADO,
            "tiene_electricidad": True,
        }
        datos.update(extra)
        return Vivienda.objects.create(
            zona=zona or self.zona1, direccion=direccion, **datos
        )

    def crear(
        self,
        direccion="Av. Central 100",
        zona=None,
        censista=None,
        estado=EstadoEncuesta.PENDIENTE,
        vivienda=None,
        referencia="",
        **extra,
    ):
        """Crea una encuesta coherente con el estado pedido, y su vivienda.

        Pasa por cambiar_estado() y no escribe las fechas a mano, porque escribirlas
        a mano en 30 pruebas sería repetir 30 veces la regla que el modelo ya sabe
        —y equivocarse en alguna—.

        Desde la HU-08 crea también la vivienda, salvo que se le pase una: así las
        pruebas de la HU-07 siguieron leyéndose igual después de que la dirección
        cambiara de tabla.
        """
        if vivienda is None:
            vivienda = self.crear_vivienda(
                direccion=direccion, zona=zona, referencia=referencia
            )

        encuesta = Encuesta(
            vivienda=vivienda,
            censista=censista or self.marta,
            **extra,
        )

        # Desde la HU-10, cerrar sin levantar exige un motivo escrito
        # (`encuesta_cierre_con_motivo`). El ayudante lo pone cuando hace falta, por
        # lo mismo que pasa por cambiar_estado() para las fechas: producir filas que
        # la base de datos acepta es su trabajo, no el de cada prueba.
        if estado in ESTADOS_SIN_LEVANTAR and not encuesta.motivo_cierre:
            encuesta.motivo_cierre = "Motivo de prueba: no se pudo levantar."

        # Y desde la HU-14, resolver en contra exige un comentario del supervisor
        # (`encuesta_resolucion_con_motivo`, que la HU-15 extendió a OBSERVADA).
        if estado == EstadoEncuesta.ANULADA and not encuesta.comentario_revision:
            encuesta.comentario_revision = "Duplicada: motivo de prueba."

        if estado == EstadoEncuesta.OBSERVADA and not encuesta.comentario_revision:
            encuesta.comentario_revision = (
                "Corregir: Integrantes del hogar.\n\nMotivo de prueba: faltan personas."
            )

        encuesta.cambiar_estado(estado, guardar=False)
        encuesta.save()
        return encuesta

    def devolver(self, encuesta, observaciones="Faltan dos personas por registrar."):
        """Deja una encuesta OBSERVADA por el camino real (HU-15).

        Antes de la HU-15 las pruebas hacían `cambiar_estado(OBSERVADA)` a secas, y
        funcionaba porque nada exigía un motivo. Al aparecer
        `encuesta_resolucion_con_motivo` esas filas dejaron de ser válidas, y la
        respuesta correcta no era añadir el comentario a mano en cada prueba: era
        pasar por `devolver()`, que es lo único que produce una encuesta observada en
        la aplicación de verdad.

        Una prueba que construye a mano un estado que el sistema no puede alcanzar
        comprueba algo que no existe.
        """
        return encuesta.devolver(usuario=self.supervisor, observaciones=observaciones)

    def url_detalle(self, encuesta):
        return reverse("fichas:encuesta_detalle", kwargs={"pk": encuesta.pk})

    def operativo_cerrado(self):
        """Un segundo operativo, ya cerrado, con su propia zona."""
        cerrado = Operativo.objects.create(
            nombre="Censo 2025",
            fecha_inicio=date(2025, 3, 1),
            fecha_termino=date(2025, 3, 31),
            estado=EstadoOperativo.CERRADO,
        )
        sector = Sector.objects.create(
            operativo=cerrado, comuna=self.concepcion, nombre="Sector antiguo"
        )
        return Zona.objects.create(sector=sector, nombre="Zona vieja")


# ==========================================================================
# 1. EL MODELO Encuesta
# ==========================================================================


class EncuestaModeloTest(BaseEncuestaTest):
    def test_una_encuesta_nace_pendiente_y_sin_fechas(self):
        encuesta = Encuesta.objects.create(
            vivienda=self.crear_vivienda(), censista=self.marta
        )

        self.assertEqual(encuesta.estado, EstadoEncuesta.PENDIENTE)
        self.assertIsNone(encuesta.iniciada_en)
        self.assertIsNone(encuesta.cerrada_en)

    def test_registra_cuando_se_creo(self):
        encuesta = self.crear()

        self.assertIsNotNone(encuesta.creada_en)
        self.assertIsNotNone(encuesta.actualizada_en)

    def test_el_texto_incluye_la_direccion_y_el_estado(self):
        encuesta = self.crear(direccion="Pasaje Los Robles 15")

        self.assertIn("Pasaje Los Robles 15", str(encuesta))
        self.assertIn("Pendiente", str(encuesta))

    def test_la_ubicacion_lleva_el_camino_completo(self):
        """Una dirección suelta no identifica nada fuera de su zona."""
        encuesta = self.crear(direccion="Av. Central 100")

        self.assertEqual(
            encuesta.ubicacion, "Av. Central 100 · Zona 1 · Los Boldos"
        )

    def test_atajos_a_la_jerarquia_territorial(self):
        encuesta = self.crear()

        self.assertEqual(encuesta.sector, self.boldos)
        self.assertEqual(encuesta.comuna, self.concepcion)
        self.assertEqual(encuesta.operativo, self.operativo)

    def test_get_absolute_url_apunta_a_su_ficha(self):
        encuesta = self.crear()

        self.assertEqual(encuesta.get_absolute_url(), self.url_detalle(encuesta))

    def test_el_orden_por_defecto_es_geografico(self):
        """El orden del modelo es el del recorrido, no el de urgencia."""
        self.crear(direccion="B", zona=self.zona2)
        self.crear(direccion="A", zona=self.zona2)
        self.crear(direccion="C", zona=self.zona1)

        direcciones = list(
            Encuesta.objects.values_list("vivienda__direccion", flat=True)
        )

        self.assertEqual(direcciones, ["C", "A", "B"])

    def test_la_zona_no_se_puede_borrar_si_tiene_viviendas(self):
        """PROTECT: borrar la zona dejaría fichas del censo sin ubicación."""
        self.crear()

        with self.assertRaises(ProtectedError):
            self.zona1.delete()

    def test_el_encuestador_no_se_puede_borrar_si_tiene_encuestas(self):
        """PROTECT: las cuentas se deshabilitan, no se borran (HU-03)."""
        self.crear(censista=self.juan)

        with self.assertRaises(ProtectedError):
            self.juan.delete()

    def test_la_encuesta_es_accesible_desde_la_vivienda_y_desde_la_persona(self):
        encuesta = self.crear()

        self.assertIn(encuesta, encuesta.vivienda.encuestas.all())
        self.assertIn(encuesta, self.marta.encuestas.all())

    def test_la_vivienda_es_accesible_desde_su_zona(self):
        encuesta = self.crear()

        self.assertIn(encuesta.vivienda, self.zona1.viviendas.all())


# ==========================================================================
# 2. LOS ESTADOS
# ==========================================================================


class EstadosTest(BaseEncuestaTest):
    """La partición abierto/cerrado es la que sostiene toda la historia."""

    def test_los_dos_grupos_cubren_todos_los_estados(self):
        """Exhaustiva: ningún estado se queda fuera de los dos grupos."""
        agrupados = set(ESTADOS_ABIERTOS) | set(ESTADOS_CERRADOS)

        self.assertEqual(agrupados, set(EstadoEncuesta.values))

    def test_los_dos_grupos_no_se_solapan(self):
        """Excluyente: ningún estado está en los dos a la vez."""
        self.assertEqual(set(ESTADOS_ABIERTOS) & set(ESTADOS_CERRADOS), set())

    def test_requiere_trabajo_solo_en_los_estados_abiertos(self):
        for estado in EstadoEncuesta.values:
            with self.subTest(estado=estado):
                encuesta = Encuesta(estado=estado)
                self.assertEqual(
                    encuesta.requiere_trabajo, estado in ESTADOS_ABIERTOS
                )

    def test_esta_cerrada_solo_en_los_estados_cerrados(self):
        for estado in EstadoEncuesta.values:
            with self.subTest(estado=estado):
                encuesta = Encuesta(estado=estado)
                self.assertEqual(encuesta.esta_cerrada, estado in ESTADOS_CERRADOS)

    def test_en_revision_es_solo_completada(self):
        self.assertTrue(Encuesta(estado=EstadoEncuesta.COMPLETADA).en_revision)
        self.assertFalse(Encuesta(estado=EstadoEncuesta.VALIDADA).en_revision)

    def test_necesita_correccion_es_solo_observada(self):
        self.assertTrue(Encuesta(estado=EstadoEncuesta.OBSERVADA).necesita_correccion)
        self.assertFalse(Encuesta(estado=EstadoEncuesta.BORRADOR).necesita_correccion)

    def test_una_observada_requiere_trabajo_aunque_ya_se_visito(self):
        """Es el caso que justifica el estado: vuelve a ser trabajo del encuestador."""
        encuesta = self.crear(estado=EstadoEncuesta.OBSERVADA)

        self.assertTrue(encuesta.requiere_trabajo)
        self.assertFalse(encuesta.esta_cerrada)

    def test_todos_los_estados_tienen_color(self):
        """Sin esto, un estado nuevo se pintaría gris sin que nadie lo notara."""
        for estado in EstadoEncuesta.values:
            with self.subTest(estado=estado):
                self.assertTrue(Encuesta(estado=estado).color_estado)

    def test_el_color_distingue_lo_urgente_de_lo_terminado(self):
        self.assertEqual(
            Encuesta(estado=EstadoEncuesta.OBSERVADA).color_estado, "danger"
        )
        self.assertEqual(
            Encuesta(estado=EstadoEncuesta.VALIDADA).color_estado, "success"
        )


# ==========================================================================
# 3. cambiar_estado(): EL ESTADO Y SUS DOS FECHAS SE MUEVEN JUNTOS
# ==========================================================================


class CambiarEstadoTest(BaseEncuestaTest):
    def test_empezarla_marca_la_fecha_de_inicio(self):
        encuesta = self.crear()

        encuesta.cambiar_estado(EstadoEncuesta.BORRADOR)

        self.assertEqual(encuesta.estado, EstadoEncuesta.BORRADOR)
        self.assertIsNotNone(encuesta.iniciada_en)
        self.assertIsNone(encuesta.cerrada_en)

    def test_completarla_marca_la_fecha_de_cierre(self):
        encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)

        encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)

        self.assertIsNotNone(encuesta.cerrada_en)

    def test_completarla_desde_pendiente_marca_las_dos_fechas(self):
        encuesta = self.crear()

        encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)

        self.assertIsNotNone(encuesta.iniciada_en)
        self.assertIsNotNone(encuesta.cerrada_en)

    def test_no_ubicada_desde_pendiente_tambien_marca_el_inicio(self):
        """Registrar que no se pudo ubicar implica haber ido a buscarla.

        El motivo se escribe antes de la transición porque desde la HU-10 la base de
        datos lo exige (`encuesta_cierre_con_motivo`). cambiar_estado() se ocupa de
        las FECHAS, que son lo que nadie más puede mantener coherente; el motivo es
        contenido y lo aporta quien pide el cambio.
        """
        encuesta = self.crear()
        encuesta.motivo_cierre = "La dirección no existe en ese pasaje."

        encuesta.cambiar_estado(EstadoEncuesta.NO_UBICADA)

        self.assertIsNotNone(encuesta.iniciada_en)
        self.assertIsNotNone(encuesta.cerrada_en)

    def test_observarla_borra_la_fecha_de_cierre(self):
        """El caso que justifica el método: vuelve a ser trabajo pendiente.

        Se llega por `devolver()` y no por `cambiar_estado()` a secas porque desde la
        HU-15 es el único camino que produce una encuesta observada válida: la base de
        datos exige el comentario del supervisor. La transición sigue siendo lo que se
        comprueba —devolver() no toca ninguna fecha—.
        """
        encuesta = self.crear(estado=EstadoEncuesta.COMPLETADA)
        self.assertIsNotNone(encuesta.cerrada_en)

        self.devolver(encuesta)

        self.assertIsNone(encuesta.cerrada_en)
        self.assertIsNotNone(encuesta.iniciada_en)

    def test_volver_a_pendiente_borra_la_fecha_de_inicio(self):
        encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)

        encuesta.cambiar_estado(EstadoEncuesta.PENDIENTE)

        self.assertIsNone(encuesta.iniciada_en)
        self.assertIsNone(encuesta.cerrada_en)

    def test_no_se_pisa_la_fecha_de_inicio_original(self):
        """Volver a tocar la encuesta no borra cuándo se visitó por primera vez."""
        encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        primera = encuesta.iniciada_en

        encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)

        self.assertEqual(encuesta.iniciada_en, primera)

    def test_no_se_pisa_la_fecha_de_cierre_al_validar(self):
        """Validar no cambia cuándo se terminó el trabajo de terreno."""
        encuesta = self.crear(estado=EstadoEncuesta.COMPLETADA)
        cierre = encuesta.cerrada_en

        encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        self.assertEqual(encuesta.cerrada_en, cierre)

    def test_guarda_en_la_base_de_datos_por_defecto(self):
        encuesta = self.crear()

        encuesta.cambiar_estado(EstadoEncuesta.BORRADOR)

        encuesta.refresh_from_db()
        self.assertEqual(encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_con_guardar_false_no_toca_la_base_de_datos(self):
        encuesta = self.crear()

        encuesta.cambiar_estado(EstadoEncuesta.BORRADOR, guardar=False)

        self.assertEqual(
            Encuesta.objects.get(pk=encuesta.pk).estado, EstadoEncuesta.PENDIENTE
        )

    def test_acepta_el_estado_como_texto(self):
        encuesta = self.crear()

        encuesta.cambiar_estado("COMPLETADA")

        self.assertEqual(encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_rechaza_un_estado_que_no_existe(self):
        encuesta = self.crear()

        with self.assertRaises(ValueError):
            encuesta.cambiar_estado("INVENTADO")

    def test_devuelve_la_propia_encuesta(self):
        encuesta = self.crear()

        self.assertIs(encuesta.cambiar_estado(EstadoEncuesta.BORRADOR), encuesta)


# ==========================================================================
# 4. LO QUE GARANTIZA LA BASE DE DATOS
# ==========================================================================


class RestriccionesBaseDatosTest(BaseEncuestaTest):
    """Las tres restricciones valen aunque nadie pase por un formulario.

    Cada prueba se envuelve en su propio atomic() porque un IntegrityError deja la
    transacción rota: sin eso, la primera prueba que falla arrastra a las demás.
    """

    def test_no_se_puede_guardar_un_estado_que_no_existe(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(),
                    censista=self.marta,
                    estado="INVENTADO",
                )

    def test_una_pendiente_no_puede_tener_fecha_de_inicio(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(),
                    censista=self.marta,
                    estado=EstadoEncuesta.PENDIENTE,
                    iniciada_en=timezone.now(),
                )

    def test_una_empezada_tiene_que_tener_fecha_de_inicio(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(),
                    censista=self.marta,
                    estado=EstadoEncuesta.BORRADOR,
                )

    def test_una_abierta_no_puede_tener_fecha_de_cierre(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(),
                    censista=self.marta,
                    estado=EstadoEncuesta.BORRADOR,
                    iniciada_en=timezone.now(),
                    cerrada_en=timezone.now(),
                )

    def test_una_cerrada_tiene_que_tener_fecha_de_cierre(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(),
                    censista=self.marta,
                    estado=EstadoEncuesta.COMPLETADA,
                    iniciada_en=timezone.now(),
                )

    def test_una_observada_no_puede_conservar_la_fecha_de_cierre(self):
        """Si la conservara, seguiría contando como terminada en cualquier recuento."""
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(),
                    censista=self.marta,
                    estado=EstadoEncuesta.OBSERVADA,
                    iniciada_en=timezone.now(),
                    cerrada_en=timezone.now(),
                )

    def test_los_ocho_estados_se_pueden_guardar_por_el_camino_correcto(self):
        """La contrapartida: cambiar_estado() produce filas que la base acepta.

        Eran siete hasta la HU-14, que agregó ANULADA. La prueba recorre
        EstadoEncuesta.values y no una lista escrita a mano, así que un estado nuevo
        entra aquí solo.
        """
        for numero, estado in enumerate(EstadoEncuesta.values):
            with self.subTest(estado=estado):
                encuesta = self.crear(direccion=f"Calle {numero}", estado=estado)
                encuesta.refresh_from_db()
                self.assertEqual(encuesta.estado, estado)

    def test_una_vivienda_puede_tener_dos_hogares(self):
        """Es el caso que la HU-08 vino a modelar bien: una casa, dos encuestas."""
        vivienda = self.crear_vivienda(direccion="Pasaje Los Robles 47")
        self.crear(vivienda=vivienda)
        self.crear(vivienda=vivienda)

        self.assertEqual(vivienda.encuestas.count(), 2)
        self.assertTrue(vivienda.tiene_varios_hogares)

    def test_dos_viviendas_pueden_compartir_direccion(self):
        """La casa del fondo y la de adelante: misma dirección, dos viviendas."""
        self.crear_vivienda(direccion="Pasaje Los Robles 47", referencia="Del fondo")
        self.crear_vivienda(direccion="Pasaje Los Robles 47", referencia="De adelante")

        self.assertEqual(
            Vivienda.objects.filter(direccion="Pasaje Los Robles 47").count(), 2
        )


# ==========================================================================
# 5. clean(): EL MISMO ERROR, PERO LEGIBLE
# ==========================================================================


class ValidacionModeloTest(BaseEncuestaTest):
    def encuesta(self, **campos):
        base = {"vivienda": self.crear_vivienda(), "censista": self.marta}
        return Encuesta(**{**base, **campos})

    def test_pendiente_con_fecha_de_inicio_avisa_en_el_campo_correcto(self):
        encuesta = self.encuesta(
            estado=EstadoEncuesta.PENDIENTE, iniciada_en=timezone.now()
        )

        with self.assertRaises(ValidationError) as error:
            encuesta.full_clean()

        self.assertIn("iniciada_en", error.exception.message_dict)

    def test_empezada_sin_fecha_de_inicio_avisa(self):
        encuesta = self.encuesta(estado=EstadoEncuesta.BORRADOR)

        with self.assertRaises(ValidationError) as error:
            encuesta.full_clean()

        self.assertIn("iniciada_en", error.exception.message_dict)

    def test_cerrada_sin_fecha_de_cierre_avisa(self):
        encuesta = self.encuesta(
            estado=EstadoEncuesta.COMPLETADA, iniciada_en=timezone.now()
        )

        with self.assertRaises(ValidationError) as error:
            encuesta.full_clean()

        self.assertIn("cerrada_en", error.exception.message_dict)

    def test_abierta_con_fecha_de_cierre_avisa(self):
        encuesta = self.encuesta(
            estado=EstadoEncuesta.BORRADOR,
            iniciada_en=timezone.now(),
            cerrada_en=timezone.now(),
        )

        with self.assertRaises(ValidationError) as error:
            encuesta.full_clean()

        self.assertIn("cerrada_en", error.exception.message_dict)

    def test_una_encuesta_coherente_pasa_la_validacion(self):
        encuesta = self.encuesta()

        encuesta.full_clean()  # no debe lanzar


# ==========================================================================
# 6. CONTROL DE ACCESO
# ==========================================================================


class AccesoTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.encuesta = self.crear()

    def test_un_visitante_anonimo_va_al_login(self):
        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(reverse("usuarios:login"), respuesta.url)

    def test_la_ficha_tambien_exige_sesion(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(reverse("usuarios:login"), respuesta.url)

    def test_el_censista_entra_con_fichas_ver_propias(self):
        """El permiso ya se lo dio la HU-04: esta historia no concede nada."""
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.status_code, 200)

    def test_el_supervisor_entra(self):
        self.client.force_login(self.supervisor)

        self.assertEqual(self.client.get(self.url_lista).status_code, 200)

    def test_el_administrador_entra(self):
        self.client.force_login(self.admin)

        self.assertEqual(self.client.get(self.url_lista).status_code, 200)

    def test_sin_el_permiso_no_se_entra(self):
        """Se le retira fichas.ver_propias al rol y la puerta se cierra."""
        permiso = Permiso.objects.get(codigo="fichas.ver_propias")
        self.rol_censista.permisos.remove(permiso)
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.status_code, 302)

    def test_sin_el_permiso_se_explica_por_que(self):
        permiso = Permiso.objects.get(codigo="fichas.ver_propias")
        self.rol_censista.permisos.remove(permiso)
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_lista, follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("permiso" in m.lower() for m in mensajes))

    def test_un_permiso_desactivado_deja_de_conceder(self):
        """Regla de la HU-04: activo=False retira el permiso sin tocar la matriz."""
        Permiso.objects.filter(codigo="fichas.ver_propias").update(activo=False)
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_lista).status_code, 302)

    def test_un_rol_desactivado_deja_de_conceder(self):
        self.rol_censista.activo = False
        self.rol_censista.save()
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_lista).status_code, 302)

    def test_una_cuenta_sin_rol_no_entra(self):
        self.client.force_login(self.sin_rol)

        self.assertEqual(self.client.get(self.url_lista).status_code, 302)


# ==========================================================================
# 7. EL LISTADO: SOLO LO MÍO
# ==========================================================================


class MisEncuestasTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def test_muestra_las_encuestas_propias(self):
        self.crear(direccion="Av. Central 100")

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "Av. Central 100")

    def test_no_muestra_las_de_otra_persona(self):
        """El filtro por request.user es lo que hace segura esta pantalla."""
        self.crear(direccion="Calle de Juan 1", censista=self.juan)

        respuesta = self.client.get(self.url_lista)

        self.assertNotContains(respuesta, "Calle de Juan 1")

    def test_el_supervisor_ve_las_suyas_y_no_las_de_todos(self):
        """Esta pantalla es «lo mío» aunque el rol pueda ver las fichas de todos."""
        self.crear(direccion="Calle de Marta 1")
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_lista)

        self.assertNotContains(respuesta, "Calle de Marta 1")

    def test_muestra_el_estado_de_cada_encuesta(self):
        self.crear(direccion="Av. Central 100", estado=EstadoEncuesta.BORRADOR)

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "Borrador")

    def test_sin_encuestas_explica_que_va_a_pasar(self):
        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "Todavía no tienes encuestas")

    def test_las_de_operativos_cerrados_no_salen_por_defecto(self):
        zona_vieja = self.operativo_cerrado()
        self.crear(direccion="Calle antigua 1", zona=zona_vieja)

        respuesta = self.client.get(self.url_lista)

        self.assertNotContains(respuesta, "Calle antigua 1")

    def test_las_de_operativos_cerrados_salen_si_se_piden(self):
        zona_vieja = self.operativo_cerrado()
        self.crear(direccion="Calle antigua 1", zona=zona_vieja)

        respuesta = self.client.get(self.url_lista, {"historicas": "on"})

        self.assertContains(respuesta, "Calle antigua 1")

    def test_pagina_de_veinte_en_veinte(self):
        for numero in range(25):
            self.crear(direccion=f"Calle {numero:02d}")

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(len(respuesta.context["encuestas"]), 20)
        self.assertTrue(respuesta.context["is_paginated"])

    def test_la_segunda_pagina_trae_el_resto(self):
        for numero in range(25):
            self.crear(direccion=f"Calle {numero:02d}")

        respuesta = self.client.get(self.url_lista, {"page": 2})

        self.assertEqual(len(respuesta.context["encuestas"]), 5)


# ==========================================================================
# 8. EL ORDEN DE LA JORNADA
# ==========================================================================


class OrdenPorUrgenciaTest(BaseEncuestaTest):
    """Lo primero de la lista tiene que ser lo primero que hay que hacer."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def estados_listados(self):
        respuesta = self.client.get(self.url_lista)
        return [e.estado for e in respuesta.context["encuestas"]]

    def test_lo_devuelto_por_el_supervisor_va_primero(self):
        self.crear(direccion="A", estado=EstadoEncuesta.PENDIENTE)
        self.crear(direccion="B", estado=EstadoEncuesta.OBSERVADA)

        self.assertEqual(self.estados_listados()[0], EstadoEncuesta.OBSERVADA)

    def test_el_orden_completo_es_observada_borrador_pendiente_y_despues_cerradas(self):
        self.crear(direccion="A", estado=EstadoEncuesta.VALIDADA)
        self.crear(direccion="B", estado=EstadoEncuesta.PENDIENTE)
        self.crear(direccion="C", estado=EstadoEncuesta.COMPLETADA)
        self.crear(direccion="D", estado=EstadoEncuesta.BORRADOR)
        self.crear(direccion="E", estado=EstadoEncuesta.OBSERVADA)

        self.assertEqual(
            self.estados_listados(),
            [
                EstadoEncuesta.OBSERVADA,
                EstadoEncuesta.BORRADOR,
                EstadoEncuesta.PENDIENTE,
                EstadoEncuesta.COMPLETADA,
                EstadoEncuesta.VALIDADA,
            ],
        )

    def test_dentro_del_mismo_estado_manda_el_orden_del_recorrido(self):
        self.crear(direccion="Calle C", zona=self.zona1)
        self.crear(direccion="Calle A", zona=self.zona1)
        self.crear(direccion="Calle B", zona=self.zona1)

        respuesta = self.client.get(self.url_lista)
        direcciones = [e.direccion for e in respuesta.context["encuestas"]]

        self.assertEqual(direcciones, ["Calle A", "Calle B", "Calle C"])

    def test_el_orden_se_aplica_antes_de_paginar(self):
        """Ordenar en Python solo ordenaría la página ya traída de la base."""
        for numero in range(24):
            self.crear(direccion=f"Calle {numero:02d}")
        self.crear(direccion="ZZZ última", estado=EstadoEncuesta.OBSERVADA)

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.context["encuestas"][0].direccion, "ZZZ última")


# ==========================================================================
# 9. LOS FILTROS
# ==========================================================================


class FiltrosTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.robles = self.crear(
            direccion="Pasaje Los Robles 15", referencia="Casa verde"
        )
        self.central = self.crear(
            direccion="Av. Central 100",
            zona=self.zona_norte,
            estado=EstadoEncuesta.COMPLETADA,
        )

    def direcciones(self, **filtros):
        respuesta = self.client.get(self.url_lista, filtros)
        return [e.direccion for e in respuesta.context["encuestas"]]

    def test_busca_por_direccion(self):
        self.assertEqual(self.direcciones(q="robles"), ["Pasaje Los Robles 15"])

    def test_busca_por_referencia(self):
        self.assertEqual(self.direcciones(q="casa verde"), ["Pasaje Los Robles 15"])

    def test_la_busqueda_no_distingue_mayusculas(self):
        self.assertEqual(self.direcciones(q="ROBLES"), ["Pasaje Los Robles 15"])

    def test_filtra_por_un_estado_concreto(self):
        self.assertEqual(self.direcciones(estado="COMPLETADA"), ["Av. Central 100"])

    def test_filtra_por_las_que_requieren_trabajo(self):
        self.assertEqual(self.direcciones(estado="ABIERTAS"), ["Pasaje Los Robles 15"])

    def test_filtra_por_las_ya_cerradas(self):
        self.assertEqual(self.direcciones(estado="CERRADAS"), ["Av. Central 100"])

    def test_filtra_por_sector(self):
        self.assertEqual(
            self.direcciones(sector=self.norte.pk), ["Av. Central 100"]
        )

    def test_los_filtros_se_combinan(self):
        self.crear(direccion="Pasaje Los Robles 23", estado=EstadoEncuesta.COMPLETADA)

        self.assertEqual(
            self.direcciones(q="robles", estado="CERRADAS"),
            ["Pasaje Los Robles 23"],
        )

    def test_un_estado_inventado_no_rompe_la_pantalla(self):
        """Una URL manipulada devuelve la lista, no un error."""
        respuesta = self.client.get(self.url_lista, {"estado": "BORRAD0R"})

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(len(respuesta.context["encuestas"]), 2)

    def test_un_sector_inventado_no_rompe_la_pantalla(self):
        respuesta = self.client.get(self.url_lista, {"sector": 99999})

        self.assertEqual(respuesta.status_code, 200)

    def test_sin_resultados_lo_dice_sin_sugerir_que_termino_el_trabajo(self):
        respuesta = self.client.get(self.url_lista, {"q": "no existe"})

        self.assertContains(respuesta, "Ninguna encuesta coincide")

    def test_el_desplegable_solo_ofrece_los_sectores_propios(self):
        """Una opción que nunca devuelve resultados es una opción que estorba."""
        self.crear(direccion="De Juan", zona=self.zona1, censista=self.juan)
        formulario = FiltroMisEncuestasForm(censista=self.juan)

        self.assertEqual(list(formulario.fields["sector"].queryset), [self.boldos])

    def test_el_desplegable_esta_vacio_para_quien_no_tiene_encuestas(self):
        formulario = FiltroMisEncuestasForm(censista=self.sin_rol)

        self.assertEqual(list(formulario.fields["sector"].queryset), [])

    def test_el_desplegable_de_estado_ofrece_los_dos_grupos_y_los_siete_estados(self):
        formulario = FiltroMisEncuestasForm(censista=self.marta)
        valores = [valor for valor, _ in formulario.fields["estado"].choices]

        self.assertIn("ABIERTAS", valores)
        self.assertIn("CERRADAS", valores)
        for estado in EstadoEncuesta.values:
            self.assertIn(estado, valores)

    def test_la_busqueda_recorta_los_espacios(self):
        formulario = FiltroMisEncuestasForm({"q": "  robles  "}, censista=self.marta)

        self.assertTrue(formulario.is_valid())
        self.assertEqual(formulario.cleaned_data["q"], "robles")

    def test_los_filtros_sobreviven_al_cambiar_de_pagina(self):
        for numero in range(25):
            self.crear(direccion=f"Robles {numero:02d}")

        respuesta = self.client.get(self.url_lista, {"q": "robles"})

        self.assertIn("q=robles", respuesta.context["parametros"])


# ==========================================================================
# 10. LOS CONTADORES
# ==========================================================================


class ResumenTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.crear(direccion="A", estado=EstadoEncuesta.PENDIENTE)
        self.crear(direccion="B", estado=EstadoEncuesta.PENDIENTE)
        self.crear(direccion="C", estado=EstadoEncuesta.BORRADOR)
        self.crear(direccion="D", estado=EstadoEncuesta.OBSERVADA)
        self.crear(direccion="E", estado=EstadoEncuesta.COMPLETADA)
        self.crear(direccion="F", estado=EstadoEncuesta.VALIDADA)
        self.crear(direccion="G", estado=EstadoEncuesta.NO_UBICADA)
        self.crear(direccion="H", estado=EstadoEncuesta.RECHAZADA)

    def resumen(self, **filtros):
        return self.client.get(self.url_lista, filtros).context["resumen"]

    def test_cuenta_el_total(self):
        self.assertEqual(self.resumen()["total"], 8)

    def test_cuenta_lo_que_queda_por_trabajar(self):
        """Pendientes, borradores y observadas: los tres estados abiertos."""
        self.assertEqual(self.resumen()["por_trabajar"], 4)

    def test_cuenta_cada_estado_por_separado(self):
        resumen = self.resumen()

        self.assertEqual(resumen["pendientes"], 2)
        self.assertEqual(resumen["borradores"], 1)
        self.assertEqual(resumen["observadas"], 1)
        self.assertEqual(resumen["completadas"], 1)
        self.assertEqual(resumen["validadas"], 1)

    def test_cuenta_las_cerradas(self):
        self.assertEqual(self.resumen()["cerradas"], 4)

    def test_calcula_el_avance_en_porcentaje(self):
        self.assertEqual(self.resumen()["avance"], 50)

    def test_el_avance_es_cero_sin_encuestas(self):
        self.client.force_login(self.juan)

        self.assertEqual(self.resumen()["avance"], 0)

    def test_los_contadores_no_cambian_al_filtrar(self):
        """Un contador que cambia al filtrar no responde «¿cuánto me queda?»."""
        self.assertEqual(self.resumen(estado="PENDIENTE")["total"], 8)

    def test_los_contadores_no_cambian_al_pasar_de_pagina(self):
        self.assertEqual(self.resumen(page=1)["total"], 8)

    def test_los_contadores_no_incluyen_las_de_otra_persona(self):
        self.crear(direccion="De Juan", censista=self.juan)

        self.assertEqual(self.resumen()["total"], 8)

    def test_los_contadores_no_incluyen_los_operativos_cerrados(self):
        zona_vieja = self.operativo_cerrado()
        self.crear(direccion="Antigua", zona=zona_vieja)

        self.assertEqual(self.resumen()["total"], 8)

    def test_avisa_de_las_fichas_devueltas(self):
        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "devolvió")

    def test_no_avisa_si_no_hay_ninguna_devuelta(self):
        Encuesta.objects.filter(estado=EstadoEncuesta.OBSERVADA).delete()

        respuesta = self.client.get(self.url_lista)

        self.assertNotContains(respuesta, "con observaciones")


class AvancePorZonaTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def avance(self):
        return self.client.get(self.url_lista).context["avance_por_zona"]

    def test_agrupa_por_zona(self):
        self.crear(direccion="A", zona=self.zona1)
        self.crear(direccion="B", zona=self.zona1)
        self.crear(direccion="C", zona=self.zona2)

        self.assertEqual(len(self.avance()), 2)

    def test_cuenta_lo_que_queda_en_cada_zona(self):
        self.crear(direccion="A", zona=self.zona1)
        self.crear(direccion="B", zona=self.zona1, estado=EstadoEncuesta.VALIDADA)

        fila = self.avance()[0]

        self.assertEqual(fila["total"], 2)
        self.assertEqual(fila["por_trabajar"], 1)

    def test_las_zonas_con_mas_trabajo_van_primero(self):
        """La primera pastilla es la zona a la que conviene ir hoy."""
        self.crear(direccion="A", zona=self.zona2)
        self.crear(direccion="B", zona=self.zona1)
        self.crear(direccion="C", zona=self.zona1)

        self.assertEqual(self.avance()[0]["vivienda__zona__nombre"], "Zona 1")

    def test_no_incluye_zonas_donde_no_tiene_encuestas(self):
        self.crear(direccion="A", zona=self.zona1)

        nombres = [fila["vivienda__zona__nombre"] for fila in self.avance()]

        self.assertEqual(nombres, ["Zona 1"])


# ==========================================================================
# 11. RENDIMIENTO
# ==========================================================================


class ConsultasTest(BaseEncuestaTest):
    """El listado no debe pagar una consulta por encuesta.

    Mismo enfoque que la HU-05 y la HU-06: no se fija un número exacto —sería
    frágil— sino que se comprueba que el coste NO CREZCA con la cantidad de datos.
    """

    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def poblar(self, cuantas, desde=0):
        for numero in range(desde, desde + cuantas):
            self.crear(direccion=f"Calle {numero:02d}")

    def contar_consultas(self, url=None):
        with CaptureQueriesContext(connection) as captura:
            self.client.get(url or self.url_lista)
        return len(captura.captured_queries)

    def test_el_numero_de_consultas_no_crece_con_las_encuestas(self):
        self.poblar(3)
        con_tres = self.contar_consultas()

        self.poblar(6, desde=3)
        con_nueve = self.contar_consultas()

        self.assertEqual(con_tres, con_nueve)

    def test_el_numero_de_consultas_no_crece_con_las_zonas(self):
        self.crear(direccion="A", zona=self.zona1)
        con_una_zona = self.contar_consultas()

        self.crear(direccion="B", zona=self.zona2)
        self.crear(direccion="C", zona=self.zona_norte)
        con_tres_zonas = self.contar_consultas()

        self.assertEqual(con_una_zona, con_tres_zonas)

    def test_la_ficha_resuelve_el_territorio_de_una_vez(self):
        encuesta = self.crear()

        with CaptureQueriesContext(connection) as captura:
            self.client.get(self.url_detalle(encuesta))

        # Umbral holgado y deliberado: lo que se vigila es que la ficha no dispare
        # una consulta por cada nivel del territorio, no el número exacto, que
        # cambia con cualquier retoque de la plantilla base.
        self.assertLess(len(captura.captured_queries), 20)


# ==========================================================================
# 12. LA FICHA DE UNA ENCUESTA
# ==========================================================================


class DetalleTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.mia = self.crear(direccion="Av. Central 100", referencia="Reja verde")
        self.ajena = self.crear(direccion="Calle de Juan 1", censista=self.juan)

    def test_el_encuestador_abre_su_ficha(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Av. Central 100")

    def test_muestra_la_referencia_y_el_territorio(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertContains(respuesta, "Reja verde")
        self.assertContains(respuesta, "Los Boldos")
        self.assertContains(respuesta, "Zona 1")

    def test_muestra_las_indicaciones_del_supervisor(self):
        self.mia.observaciones = "Pasar después de las 19:00"
        self.mia.save()
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertContains(respuesta, "Pasar después de las 19:00")

    def test_la_ficha_ajena_responde_404_y_no_403(self):
        """Un 403 confirmaría que esa ficha existe. Con identificadores en
        secuencia, esa diferencia deja contar el padrón sin ver ninguna ficha."""
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.ajena))

        self.assertEqual(respuesta.status_code, 404)

    def test_una_ficha_inexistente_responde_lo_mismo_que_una_ajena(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(
            reverse("fichas:encuesta_detalle", kwargs={"pk": 999999})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_el_supervisor_abre_la_ficha_de_cualquiera(self):
        """Tiene fichas.ver_todas desde el reparto inicial de la HU-04."""
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertEqual(respuesta.status_code, 200)

    def test_al_supervisor_se_le_avisa_de_que_no_es_suya(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertFalse(respuesta.context["es_propia"])
        self.assertContains(respuesta, "no tuya")

    def test_al_dueno_no_se_le_avisa_de_nada(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertTrue(respuesta.context["es_propia"])
        self.assertNotContains(respuesta, "no tuya")

    def test_sin_ver_todas_no_se_abre_la_ajena_aunque_haya_ver_propias(self):
        permiso = Permiso.objects.get(codigo="fichas.ver_todas")
        self.rol_supervisor.permisos.remove(permiso)
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertEqual(respuesta.status_code, 404)

    def test_avisa_de_los_otros_hogares_de_la_misma_vivienda(self):
        """Dos familias en la misma casa: hay que verlo antes de ir."""
        self.crear(vivienda=self.mia.vivienda)
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertContains(respuesta, "Otros hogares en esta misma vivienda")

    def test_no_avisa_si_la_vivienda_tiene_un_solo_hogar(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertNotContains(respuesta, "Otros hogares en esta misma vivienda")

    def test_otra_vivienda_con_la_misma_direccion_no_es_otro_hogar(self):
        """Es la diferencia que el modelo de la HU-08 vino a hacer explícita.

        Dos viviendas en la misma dirección —la del fondo y la de adelante— NO son
        dos hogares de una misma casa. Antes de la HU-08 la consulta comparaba
        direcciones y las confundía; ahora pregunta por la clave foránea y no puede.
        """
        self.crear(direccion="Av. Central 100", referencia="La del fondo")
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertEqual(len(respuesta.context["otros_hogares"]), 0)

    def test_destaca_una_ficha_devuelta_por_el_supervisor(self):
        self.devolver(self.mia)
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.mia))

        self.assertContains(respuesta, "devolvió esta ficha")

    def test_se_puede_abrir_una_ficha_de_un_operativo_cerrado(self):
        """No sale en el listado por defecto, pero su enlace sigue siendo válido."""
        zona_vieja = self.operativo_cerrado()
        antigua = self.crear(direccion="Calle antigua 1", zona=zona_vieja)
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_detalle(antigua)).status_code, 200)


# ==========================================================================
# 13. EL PANEL DEL CENSISTA
# ==========================================================================


class PanelCensistaTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.url_panel = reverse("dashboards:censista")

    def test_los_contadores_dejaron_de_ser_marcadores_de_posicion(self):
        self.crear(direccion="A")
        self.crear(direccion="B", estado=EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_panel)
        resumen = respuesta.context["resumen_encuestas"]

        self.assertEqual(resumen["total"], 2)
        self.assertEqual(resumen["por_trabajar"], 1)

    def test_muestra_las_proximas_encuestas(self):
        self.crear(direccion="Av. Central 100")

        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, "Av. Central 100")

    def test_no_muestra_las_ya_cerradas_entre_las_proximas(self):
        self.crear(direccion="Ya validada", estado=EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(len(respuesta.context["proximas_encuestas"]), 0)

    def test_las_proximas_van_en_orden_de_urgencia(self):
        self.crear(direccion="A", estado=EstadoEncuesta.PENDIENTE)
        self.crear(direccion="B", estado=EstadoEncuesta.OBSERVADA)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["proximas_encuestas"][0].direccion, "B")

    def test_muestra_como_maximo_cinco(self):
        for numero in range(8):
            self.crear(direccion=f"Calle {numero}")

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(len(respuesta.context["proximas_encuestas"]), 5)

    def test_avisa_de_las_fichas_devueltas(self):
        self.crear(direccion="A", estado=EstadoEncuesta.OBSERVADA)

        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, "con observaciones")

    def test_no_cuenta_las_de_otra_persona(self):
        self.crear(direccion="De Juan", censista=self.juan)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["resumen_encuestas"]["total"], 0)

    def test_no_cuenta_las_de_operativos_cerrados(self):
        zona_vieja = self.operativo_cerrado()
        self.crear(direccion="Antigua", zona=zona_vieja)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["resumen_encuestas"]["total"], 0)

    def test_sin_encuestas_explica_que_todavia_no_hay_padron(self):
        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, "Todavía no hay encuestas cargadas")

    def test_con_el_padron_terminado_lo_dice_como_una_buena_noticia(self):
        self.crear(direccion="A", estado=EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, "No te queda ninguna encuesta")


# ==========================================================================
# 14. EL MENÚ
# ==========================================================================


class MenuTest(BaseEncuestaTest):
    def test_el_enlace_aparece_para_quien_tiene_el_permiso(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(reverse("dashboards:censista"))

        self.assertContains(respuesta, reverse("fichas:mis_encuestas"))

    def test_el_enlace_no_aparece_sin_el_permiso(self):
        """Ocultarlo es comodidad, no seguridad: la vista lo comprueba igual."""
        permiso = Permiso.objects.get(codigo="fichas.ver_propias")
        self.rol_censista.permisos.remove(permiso)
        self.client.force_login(self.marta)

        respuesta = self.client.get(reverse("dashboards:censista"))

        self.assertNotContains(respuesta, reverse("fichas:mis_encuestas"))


# ==========================================================================
# 15. EL COMANDO DE DEMOSTRACIÓN
# ==========================================================================


class ComandoDemoTest(BaseEncuestaTest):
    def test_crea_encuestas_en_todos_los_estados(self):
        call_command("crear_encuestas_demo", censista="marta@opso.cl", verbosity=0)

        estados = set(
            Encuesta.objects.filter(censista=self.marta).values_list(
                "estado", flat=True
            )
        )

        self.assertEqual(estados, set(EstadoEncuesta.values))

    def test_todas_las_encuestas_creadas_son_coherentes(self):
        """Si el comando produjera filas incoherentes, la base las habría rechazado."""
        call_command("crear_encuestas_demo", censista="marta@opso.cl", verbosity=0)

        for encuesta in Encuesta.objects.all():
            with self.subTest(direccion=encuesta.direccion):
                encuesta.full_clean(exclude=["zona", "censista"])

    def test_asigna_el_sector_para_que_la_demostracion_sea_coherente(self):
        call_command("crear_encuestas_demo", censista="marta@opso.cl", verbosity=0)

        self.assertTrue(
            AsignacionSector.objects.filter(censista=self.marta, activa=True).exists()
        )

    def test_se_puede_ejecutar_dos_veces_sin_duplicar(self):
        call_command("crear_encuestas_demo", censista="marta@opso.cl", verbosity=0)
        despues_de_una = Encuesta.objects.count()

        call_command("crear_encuestas_demo", censista="marta@opso.cl", verbosity=0)

        self.assertEqual(Encuesta.objects.count(), despues_de_una)

    def test_falla_con_un_mensaje_util_si_la_cuenta_no_existe(self):
        with self.assertRaises(CommandError) as error:
            call_command("crear_encuestas_demo", censista="nadie@opso.cl", verbosity=0)

        self.assertIn("crear_usuarios_demo", str(error.exception))


# ==========================================================================
# 16. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionTest(BaseEncuestaTest):
    """De la asignación del sector (HU-06) a la jornada del encuestador (HU-07)."""

    def test_recorrido_completo(self):
        # 1. El supervisor reparte el sector, como en la HU-06.
        AsignacionSector.objects.create(
            sector=self.boldos, censista=self.marta, asignado_por=self.supervisor
        )

        # 2. Se carga el padrón de esa zona.
        pendiente = self.crear(direccion="Av. Central 100")
        self.crear(direccion="Av. Central 118", estado=EstadoEncuesta.OBSERVADA)
        self.crear(direccion="Av. Central 132", estado=EstadoEncuesta.VALIDADA)

        # 3. La encuestadora entra y ve su jornada ordenada por urgencia.
        self.client.force_login(self.marta)
        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.context["resumen"]["total"], 3)
        self.assertEqual(respuesta.context["resumen"]["por_trabajar"], 2)
        self.assertEqual(
            respuesta.context["encuestas"][0].estado, EstadoEncuesta.OBSERVADA
        )

        # 4. Filtra lo que le queda por hacer.
        respuesta = self.client.get(self.url_lista, {"estado": "ABIERTAS"})
        self.assertEqual(len(respuesta.context["encuestas"]), 2)

        # 5. Abre una ficha para ver dónde queda.
        respuesta = self.client.get(self.url_detalle(pendiente))
        self.assertEqual(respuesta.status_code, 200)

        # 6. La empieza en terreno y la deja a medias: el avance lo refleja.
        pendiente.cambiar_estado(EstadoEncuesta.BORRADOR)
        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.context["resumen"]["borradores"], 1)
        self.assertEqual(respuesta.context["resumen"]["avance"], 33)

        # 7. Y Juan sigue sin ver nada de todo esto.
        self.client.force_login(self.juan)
        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.context["resumen"]["total"], 0)
        self.assertContains(respuesta, "Todavía no tienes encuestas")


# ==========================================================================
# HU-08 — 17. EL MODELO Vivienda
# ==========================================================================


class ViviendaModeloTest(BaseEncuestaTest):
    def test_una_vivienda_se_crea_sin_describir(self):
        """La columna admite el vacío: hay padrón heredado que no está descrito."""
        vivienda = Vivienda.objects.create(zona=self.zona1, direccion="Calle 1")

        self.assertEqual(vivienda.tipo, "")
        self.assertIsNone(vivienda.tiene_electricidad)
        self.assertFalse(vivienda.datos_completos)

    def test_una_vivienda_descrita_lo_dice(self):
        self.assertTrue(self.crear_vivienda().datos_completos)

    def test_falta_una_caracteristica_y_ya_no_esta_completa(self):
        """Las seis, o no está descrita: media descripción no sirve para calcular."""
        for campo in Vivienda.CARACTERISTICAS:
            with self.subTest(campo=campo):
                vivienda = self.crear_vivienda(direccion=f"Calle {campo}", **{campo: ""})
                self.assertFalse(vivienda.datos_completos)

    def test_sin_saber_si_tiene_luz_tampoco_esta_completa(self):
        vivienda = self.crear_vivienda(tiene_electricidad=None)

        self.assertFalse(vivienda.datos_completos)

    def test_el_texto_es_la_direccion(self):
        self.assertEqual(str(self.crear_vivienda(direccion="Calle 9")), "Calle 9")

    def test_el_nombre_completo_lleva_zona_y_sector(self):
        vivienda = self.crear_vivienda(direccion="Calle 9")

        self.assertEqual(vivienda.nombre_completo, "Calle 9 · Zona 1 · Los Boldos")

    def test_atajos_a_la_jerarquia_territorial(self):
        vivienda = self.crear_vivienda()

        self.assertEqual(vivienda.sector, self.boldos)
        self.assertEqual(vivienda.comuna, self.concepcion)
        self.assertEqual(vivienda.operativo, self.operativo)

    def test_cuenta_sus_hogares(self):
        vivienda = self.crear_vivienda()
        self.crear(vivienda=vivienda)
        self.crear(vivienda=vivienda)

        self.assertEqual(vivienda.total_hogares(), 2)
        self.assertTrue(vivienda.tiene_varios_hogares)

    def test_con_un_solo_hogar_no_es_compartida(self):
        vivienda = self.crear_vivienda()
        self.crear(vivienda=vivienda)

        self.assertFalse(vivienda.tiene_varios_hogares)

    def test_borrar_la_vivienda_se_lleva_sus_encuestas(self):
        """CASCADE: una encuesta no significa nada sin la casa que levanta."""
        vivienda = self.crear_vivienda()
        self.crear(vivienda=vivienda)

        vivienda.delete()

        self.assertEqual(Encuesta.objects.count(), 0)

    def test_un_tipo_inventado_lo_rechaza_la_base_de_datos(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Vivienda.objects.create(
                    zona=self.zona1, direccion="Calle 1", tipo="MANSION"
                )

    def test_una_tenencia_inventada_la_rechaza_la_base_de_datos(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Vivienda.objects.create(
                    zona=self.zona1, direccion="Calle 1", tenencia="HEREDADA"
                )

    def test_el_vacio_si_se_admite(self):
        """La contrapartida: sin describir es un valor legítimo."""
        Vivienda.objects.create(zona=self.zona1, direccion="Calle 1", tipo="")

        self.assertEqual(Vivienda.objects.count(), 1)


class PuedeRegistrarseTrabajoTest(BaseEncuestaTest):
    """Las tres condiciones del territorio, cada una con su motivo explicado."""

    def test_en_un_operativo_en_curso_se_puede(self):
        permitido, motivo = self.crear_vivienda().puede_registrarse_trabajo()

        self.assertTrue(permitido)
        self.assertEqual(motivo, "")

    def test_en_un_operativo_cerrado_no(self):
        vivienda = self.crear_vivienda(zona=self.operativo_cerrado())

        permitido, motivo = vivienda.puede_registrarse_trabajo()

        self.assertFalse(permitido)
        self.assertIn("cerrado", motivo)

    def test_en_un_sector_desactivado_no(self):
        self.boldos.activo = False
        self.boldos.save()

        permitido, motivo = self.crear_vivienda().puede_registrarse_trabajo()

        self.assertFalse(permitido)
        self.assertIn("desactivado", motivo)

    def test_en_una_zona_desactivada_no(self):
        self.zona1.activa = False
        self.zona1.save()

        permitido, motivo = self.crear_vivienda().puede_registrarse_trabajo()

        self.assertFalse(permitido)
        self.assertIn("desactivada", motivo)

    def test_el_motivo_nombra_el_obstaculo_concreto(self):
        """«No se puede» obliga a adivinar; el motivo dice qué arreglar."""
        vivienda = self.crear_vivienda(zona=self.operativo_cerrado())

        _, motivo = vivienda.puede_registrarse_trabajo()

        self.assertIn("Censo 2025", motivo)


class PuedeRegistrarseEncuestaTest(BaseEncuestaTest):
    def test_una_encuesta_abierta_admite_cambios(self):
        for estado in ESTADOS_ABIERTOS:
            with self.subTest(estado=estado):
                encuesta = self.crear(direccion=f"Calle {estado}", estado=estado)
                permitido, _ = encuesta.puede_registrarse()
                self.assertTrue(permitido)

    def test_una_encuesta_cerrada_no_admite_cambios(self):
        for estado in ESTADOS_CERRADOS:
            with self.subTest(estado=estado):
                encuesta = self.crear(direccion=f"Calle {estado}", estado=estado)
                permitido, motivo = encuesta.puede_registrarse()
                self.assertFalse(permitido)
                self.assertIn("supervisor", motivo)

    def test_el_territorio_manda_sobre_el_estado(self):
        """Aunque la encuesta esté abierta, un operativo cerrado la congela."""
        encuesta = self.crear(zona=self.operativo_cerrado())

        permitido, motivo = encuesta.puede_registrarse()

        self.assertFalse(permitido)
        self.assertIn("cerrado", motivo)


# ==========================================================================
# HU-08 — 18. EL MODELO GrupoFamiliar
# ==========================================================================


class GrupoFamiliarModeloTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)

    def hogar(self, **extra):
        datos = {
            "encuesta": self.encuesta,
            "jefe_hogar_nombre": "Rosa Millán",
            "integrantes_declarados": 4,
        }
        datos.update(extra)
        return GrupoFamiliar.objects.create(**datos)

    def test_se_crea_asociado_a_su_encuesta(self):
        hogar = self.hogar()

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.grupo_familiar, hogar)

    def test_la_encuesta_sabe_si_ya_tiene_hogar(self):
        self.assertFalse(self.encuesta.tiene_grupo_familiar)

        self.hogar()

        self.assertTrue(Encuesta.objects.get(pk=self.encuesta.pk).tiene_grupo_familiar)

    def test_el_texto_nombra_al_jefe_de_hogar(self):
        self.assertEqual(str(self.hogar()), "Hogar de Rosa Millán")

    def test_no_puede_haber_dos_hogares_en_la_misma_encuesta(self):
        """Uno a uno: el segundo hogar va en OTRA encuesta de la misma vivienda."""
        self.hogar()

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                GrupoFamiliar.objects.create(
                    encuesta=self.encuesta,
                    jefe_hogar_nombre="Otro",
                    integrantes_declarados=2,
                )

    def test_un_hogar_no_puede_tener_cero_personas(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.hogar(integrantes_declarados=0)

    def test_borrar_la_encuesta_se_lleva_el_hogar(self):
        """CASCADE: el dato no es «el hogar de la casa», es «lo que respondió esta
        encuesta». Sin la encuesta no significa nada."""
        self.hogar()

        self.encuesta.delete()

        self.assertEqual(GrupoFamiliar.objects.count(), 0)

    def test_el_rut_se_normaliza_al_guardar(self):
        hogar = self.hogar(jefe_hogar_rut="12.345.678-5")

        hogar.refresh_from_db()
        self.assertEqual(hogar.jefe_hogar_rut, "12345678-5")

    def test_el_rut_puede_quedar_vacio(self):
        hogar = self.hogar()

        self.assertEqual(hogar.jefe_hogar_rut, "")

    def test_calcula_el_ingreso_por_persona(self):
        hogar = self.hogar(integrantes_declarados=4, ingreso_mensual=800000)

        self.assertEqual(hogar.ingreso_por_persona, 200000)

    def test_sin_ingreso_declarado_no_hay_ingreso_por_persona(self):
        """None y no cero: no declarar no es lo mismo que no tener."""
        self.assertIsNone(self.hogar().ingreso_por_persona)


# ==========================================================================
# HU-08 — 19. QUÉ ZONAS PUEDE USAR CADA PERSONA
# ==========================================================================


class ZonasDisponiblesTest(BaseEncuestaTest):
    """La regla de negocio central de la HU-08, comprobada aparte de las vistas."""

    def setUp(self):
        super().setUp()
        self.asignacion = AsignacionSector.objects.create(
            sector=self.boldos, censista=self.marta, asignado_por=self.supervisor
        )

    def test_ofrece_las_zonas_del_sector_asignado(self):
        self.assertEqual(
            set(zonas_disponibles(self.marta)), {self.zona1, self.zona2}
        )

    def test_no_ofrece_zonas_de_sectores_ajenos(self):
        self.assertNotIn(self.zona_norte, zonas_disponibles(self.marta))

    def test_sin_asignacion_no_hay_ninguna_zona(self):
        self.assertEqual(list(zonas_disponibles(self.juan)), [])

    def test_una_asignacion_retirada_deja_de_dar_acceso(self):
        """El reparto de la HU-06 no es informativo: es una regla de seguridad."""
        self.asignacion.desactivar()

        self.assertEqual(list(zonas_disponibles(self.marta)), [])

    def test_un_operativo_cerrado_no_ofrece_sus_zonas(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.assertEqual(list(zonas_disponibles(self.marta)), [])

    def test_un_sector_desactivado_no_ofrece_sus_zonas(self):
        self.boldos.activo = False
        self.boldos.save()

        self.assertEqual(list(zonas_disponibles(self.marta)), [])

    def test_una_zona_desactivada_no_se_ofrece(self):
        self.zona1.activa = False
        self.zona1.save()

        self.assertEqual(list(zonas_disponibles(self.marta)), [self.zona2])

    def test_el_supervisor_no_tiene_zonas_donde_registrar(self):
        """Separación de funciones: quien valida no levanta."""
        self.assertEqual(list(zonas_disponibles(self.supervisor)), [])


# ==========================================================================
# HU-08 — 20. EL FORMULARIO DE LA VIVIENDA
# ==========================================================================


class ViviendaFormTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)

    def datos(self, **extra):
        base = {
            "zona": self.zona1.pk,
            "direccion": "Pasaje Nuevo 10",
            "referencia": "Casa amarilla",
            "tipo": TipoVivienda.CASA,
            "tenencia": TenenciaVivienda.ARRENDADA,
            "materialidad_muros": MaterialidadMuros.ALBANILERIA,
            "origen_agua": OrigenAgua.RED_PUBLICA,
            "sistema_sanitario": SistemaSanitario.ALCANTARILLADO,
            "tiene_electricidad": True,
        }
        base.update(extra)
        return base

    def test_un_formulario_completo_es_valido(self):
        formulario = ViviendaForm(self.datos(), censista=self.marta)

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_desplegable_solo_ofrece_las_zonas_asignadas(self):
        formulario = ViviendaForm(censista=self.marta)

        self.assertEqual(
            set(formulario.fields["zona"].queryset), {self.zona1, self.zona2}
        )

    def test_enviar_una_zona_ajena_no_sirve_de_nada(self):
        """Si la opción no está en el formulario, mandarla a mano no la hace válida."""
        formulario = ViviendaForm(
            self.datos(zona=self.zona_norte.pk), censista=self.marta
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("zona", formulario.errors)

    def test_las_seis_caracteristicas_son_obligatorias(self):
        for campo in ViviendaForm.OBLIGATORIOS:
            with self.subTest(campo=campo):
                formulario = ViviendaForm(
                    self.datos(**{campo: ""}), censista=self.marta
                )
                self.assertFalse(formulario.is_valid())
                self.assertIn(campo, formulario.errors)

    def test_la_referencia_y_las_observaciones_son_opcionales(self):
        formulario = ViviendaForm(
            self.datos(referencia="", observaciones=""), censista=self.marta
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_la_direccion_se_recorta(self):
        formulario = ViviendaForm(
            self.datos(direccion="   Calle 5   "), censista=self.marta
        )

        self.assertTrue(formulario.is_valid())
        self.assertEqual(formulario.cleaned_data["direccion"], "Calle 5")

    def test_registra_quien_la_dio_de_alta(self):
        formulario = ViviendaForm(self.datos(), censista=self.marta)
        formulario.is_valid()

        vivienda = formulario.save()

        self.assertEqual(vivienda.registrada_por, self.marta)

    def test_editar_no_cambia_quien_la_registro(self):
        vivienda = self.crear_vivienda(direccion="Calle 5")
        vivienda.registrada_por = self.juan
        vivienda.save()

        formulario = ViviendaForm(
            self.datos(direccion="Calle 5"), censista=self.marta, instance=vivienda
        )
        formulario.is_valid()
        formulario.save()

        vivienda.refresh_from_db()
        self.assertEqual(vivienda.registrada_por, self.juan)

    # -- el aviso de duplicado --------------------------------------------

    def test_una_direccion_repetida_pide_confirmacion(self):
        self.crear_vivienda(direccion="Pasaje Nuevo 10")

        formulario = ViviendaForm(self.datos(), censista=self.marta)

        self.assertFalse(formulario.is_valid())
        self.assertIn("confirmar_duplicado", formulario.errors)

    def test_con_la_casilla_marcada_se_guarda(self):
        """Avisar y no bloquear: dos viviendas en un sitio es un caso real."""
        self.crear_vivienda(direccion="Pasaje Nuevo 10")

        formulario = ViviendaForm(
            self.datos(confirmar_duplicado=True), censista=self.marta
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_aviso_no_distingue_mayusculas(self):
        self.crear_vivienda(direccion="PASAJE NUEVO 10")

        formulario = ViviendaForm(self.datos(), censista=self.marta)

        self.assertFalse(formulario.is_valid())

    def test_una_direccion_igual_en_otra_zona_no_es_duplicado(self):
        self.crear_vivienda(direccion="Pasaje Nuevo 10", zona=self.zona2)

        formulario = ViviendaForm(self.datos(), censista=self.marta)

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_al_editar_una_vivienda_no_es_duplicado_de_si_misma(self):
        vivienda = self.crear_vivienda(direccion="Pasaje Nuevo 10")

        formulario = ViviendaForm(
            self.datos(), censista=self.marta, instance=vivienda
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)


# ==========================================================================
# HU-08 — 21. EL FORMULARIO DEL GRUPO FAMILIAR
# ==========================================================================


class GrupoFamiliarFormTest(BaseEncuestaTest):
    def datos(self, **extra):
        base = {
            "jefe_hogar_nombre": "Rosa Elena Millán",
            "jefe_hogar_rut": "",
            "telefono_contacto": "",
            "integrantes_declarados": 3,
            "ingreso_mensual": "",
            "observaciones": "",
        }
        base.update(extra)
        return base

    def test_lo_minimo_es_el_nombre_y_cuantos_son(self):
        formulario = GrupoFamiliarForm(self.datos())

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_nombre_es_obligatorio(self):
        formulario = GrupoFamiliarForm(self.datos(jefe_hogar_nombre=""))

        self.assertFalse(formulario.is_valid())

    def test_una_inicial_no_es_un_nombre(self):
        formulario = GrupoFamiliarForm(self.datos(jefe_hogar_nombre="R."))

        self.assertFalse(formulario.is_valid())
        self.assertIn("jefe_hogar_nombre", formulario.errors)

    def test_el_nombre_se_recorta(self):
        formulario = GrupoFamiliarForm(self.datos(jefe_hogar_nombre="  Rosa Millán "))

        self.assertTrue(formulario.is_valid())
        self.assertEqual(formulario.cleaned_data["jefe_hogar_nombre"], "Rosa Millán")

    def test_el_numero_de_personas_es_obligatorio(self):
        formulario = GrupoFamiliarForm(self.datos(integrantes_declarados=""))

        self.assertFalse(formulario.is_valid())

    def test_cero_personas_no_es_un_hogar(self):
        formulario = GrupoFamiliarForm(self.datos(integrantes_declarados=0))

        self.assertFalse(formulario.is_valid())
        self.assertIn("integrantes_declarados", formulario.errors)

    def test_un_rut_valido_se_acepta(self):
        formulario = GrupoFamiliarForm(self.datos(jefe_hogar_rut="12345678-5"))

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_un_rut_con_digito_verificador_equivocado_se_rechaza(self):
        """Un RUT mal escrito es peor que ninguno: parece identificar y no lo hace."""
        formulario = GrupoFamiliarForm(self.datos(jefe_hogar_rut="12345678-9"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("jefe_hogar_rut", formulario.errors)

    def test_sin_rut_tambien_se_guarda(self):
        """No se puede condicionar el registro a entregar un dato personal."""
        formulario = GrupoFamiliarForm(self.datos(jefe_hogar_rut=""))

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_un_ingreso_normal_se_acepta(self):
        formulario = GrupoFamiliarForm(self.datos(ingreso_mensual=650000))

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_un_ingreso_con_un_digito_de_mas_se_rechaza(self):
        formulario = GrupoFamiliarForm(self.datos(ingreso_mensual=650000000))

        self.assertFalse(formulario.is_valid())
        self.assertIn("ingreso_mensual", formulario.errors)

    def test_sin_ingreso_tambien_se_guarda(self):
        formulario = GrupoFamiliarForm(self.datos(ingreso_mensual=""))

        self.assertTrue(formulario.is_valid(), formulario.errors)


# ==========================================================================
# HU-08 — 22. REGISTRAR UNA VIVIENDA (LA PANTALLA)
# ==========================================================================


class RegistrarViviendaTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.url = reverse("fichas:vivienda_registrar")
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "zona": self.zona1.pk,
            "direccion": "Pasaje Nuevo 10",
            "referencia": "",
            "tipo": TipoVivienda.CASA,
            "tenencia": TenenciaVivienda.ARRENDADA,
            "materialidad_muros": MaterialidadMuros.ALBANILERIA,
            "origen_agua": OrigenAgua.RED_PUBLICA,
            "sistema_sanitario": SistemaSanitario.ALCANTARILLADO,
            "tiene_electricidad": True,
            "observaciones": "",
        }
        base.update(extra)
        return base

    def test_el_encuestador_ve_el_formulario(self):
        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Registrar una vivienda")

    def test_guardar_crea_la_vivienda(self):
        self.client.post(self.url, self.datos())

        self.assertEqual(Vivienda.objects.count(), 1)

    def test_guardar_crea_ademas_su_encuesta(self):
        """Nadie registra una vivienda «por si acaso»: se registra porque se está ahí."""
        self.client.post(self.url, self.datos())

        encuesta = Encuesta.objects.get()
        self.assertEqual(encuesta.censista, self.marta)
        self.assertEqual(encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_la_encuesta_nueva_no_queda_como_pendiente(self):
        """PENDIENTE significa «sin visitar», y la visita acaba de ocurrir."""
        self.client.post(self.url, self.datos())

        self.assertIsNotNone(Encuesta.objects.get().iniciada_en)

    def test_despues_de_guardar_lleva_al_formulario_del_hogar(self):
        respuesta = self.client.post(self.url, self.datos())

        encuesta = Encuesta.objects.get()
        self.assertRedirects(
            respuesta, reverse("fichas:registrar_hogar", kwargs={"pk": encuesta.pk})
        )

    def test_un_formulario_incompleto_no_crea_nada(self):
        self.client.post(self.url, self.datos(tipo=""))

        self.assertEqual(Vivienda.objects.count(), 0)
        self.assertEqual(Encuesta.objects.count(), 0)

    def test_una_zona_ajena_no_crea_nada(self):
        self.client.post(self.url, self.datos(zona=self.zona_norte.pk))

        self.assertEqual(Vivienda.objects.count(), 0)

    def test_sin_territorio_asignado_no_se_ofrece_el_formulario(self):
        self.client.force_login(self.juan)

        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "No tienes territorio donde registrar")

    def test_sin_territorio_asignado_tampoco_se_puede_enviar(self):
        self.client.force_login(self.juan)

        self.client.post(self.url, self.datos())

        self.assertEqual(Vivienda.objects.count(), 0)

    def test_con_el_operativo_cerrado_no_hay_donde_registrar(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "No tienes territorio donde registrar")

    def test_el_supervisor_no_puede_registrar(self):
        """No tiene fichas.crear: la separación de funciones la aplica la matriz."""
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 302)

    def test_sin_el_permiso_de_crear_no_se_entra(self):
        self.rol_censista.permisos.remove(Permiso.objects.get(codigo="fichas.crear"))
        self.rol_censista.permisos.remove(Permiso.objects.get(codigo="fichas.editar"))

        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 302)

    def test_una_direccion_repetida_no_se_guarda_a_la_primera(self):
        self.crear_vivienda(direccion="Pasaje Nuevo 10")

        self.client.post(self.url, self.datos())

        self.assertEqual(Vivienda.objects.count(), 1)

    def test_una_direccion_repetida_se_guarda_al_confirmar(self):
        self.crear_vivienda(direccion="Pasaje Nuevo 10")

        self.client.post(self.url, self.datos(confirmar_duplicado="on"))

        self.assertEqual(Vivienda.objects.count(), 2)

    def test_el_aviso_enlaza_la_vivienda_que_ya_existe(self):
        otra = self.crear_vivienda(direccion="Pasaje Nuevo 10")

        respuesta = self.client.post(self.url, self.datos())

        self.assertContains(
            respuesta, reverse("fichas:vivienda_detalle", kwargs={"pk": otra.pk})
        )


# ==========================================================================
# HU-08 — 23. EDITAR UNA VIVIENDA
# ==========================================================================


class EditarViviendaTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.vivienda = Vivienda.objects.create(
            zona=self.zona1, direccion="Calle del padrón 1"
        )
        self.url = reverse("fichas:vivienda_editar", kwargs={"pk": self.vivienda.pk})
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "zona": self.zona1.pk,
            "direccion": "Calle del padrón 1",
            "referencia": "",
            "tipo": TipoVivienda.MEDIAGUA,
            "tenencia": TenenciaVivienda.IRREGULAR,
            "materialidad_muros": MaterialidadMuros.PRECARIO,
            "origen_agua": OrigenAgua.CAMION,
            "sistema_sanitario": SistemaSanitario.NO_TIENE,
            "tiene_electricidad": False,
            "observaciones": "",
        }
        base.update(extra)
        return base

    def test_completa_una_vivienda_del_padron_antiguo(self):
        """El caso de uso que justifica la pantalla: la migración no inventó datos."""
        self.assertFalse(self.vivienda.datos_completos)

        self.client.post(self.url, self.datos())

        self.vivienda.refresh_from_db()
        self.assertTrue(self.vivienda.datos_completos)

    def test_avisa_de_que_todavia_no_esta_descrita(self):
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "todavía no está descrita")

    def test_guarda_los_cambios(self):
        self.client.post(self.url, self.datos())

        self.vivienda.refresh_from_db()
        self.assertEqual(self.vivienda.tipo, TipoVivienda.MEDIAGUA)
        self.assertFalse(self.vivienda.tiene_electricidad)

    def test_lleva_a_la_ficha_de_la_vivienda(self):
        respuesta = self.client.post(self.url, self.datos())

        self.assertRedirects(
            respuesta,
            reverse("fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk}),
        )

    def test_una_vivienda_fuera_de_mi_territorio_responde_404(self):
        ajena = Vivienda.objects.create(zona=self.zona_norte, direccion="Otra 1")

        respuesta = self.client.get(
            reverse("fichas:vivienda_editar", kwargs={"pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_un_compañero_del_mismo_sector_si_puede_completarla(self):
        """El sector puede estar repartido y la casa es la misma para todos."""
        AsignacionSector.objects.create(sector=self.boldos, censista=self.juan)
        self.client.force_login(self.juan)

        self.assertEqual(self.client.get(self.url).status_code, 200)


# ==========================================================================
# HU-08 — 24. REGISTRAR EL GRUPO FAMILIAR
# ==========================================================================


class RegistrarHogarTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        self.url = reverse("fichas:registrar_hogar", kwargs={"pk": self.encuesta.pk})
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "jefe_hogar_nombre": "Rosa Elena Millán",
            "jefe_hogar_rut": "12345678-5",
            "telefono_contacto": "+56 9 1234 5678",
            "integrantes_declarados": 4,
            "ingreso_mensual": 700000,
            "observaciones": "",
        }
        base.update(extra)
        return base

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Registrar el grupo familiar")

    def test_recuerda_de_que_vivienda_se_trata(self):
        """Con dos hogares abiertos en la misma casa, sin esto se confunden."""
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, self.encuesta.direccion)

    def test_guarda_el_hogar(self):
        self.client.post(self.url, self.datos())

        self.encuesta.refresh_from_db()
        self.assertEqual(
            self.encuesta.grupo_familiar.jefe_hogar_nombre, "Rosa Elena Millán"
        )

    def test_normaliza_el_rut(self):
        self.client.post(self.url, self.datos(jefe_hogar_rut="12.345.678-5"))

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.grupo_familiar.jefe_hogar_rut, "12345678-5")

    def test_lleva_a_la_ficha_de_la_encuesta(self):
        respuesta = self.client.post(self.url, self.datos())

        self.assertRedirects(respuesta, self.url_detalle(self.encuesta))

    def test_una_encuesta_pendiente_pasa_a_borrador(self):
        pendiente = self.crear(direccion="Calle 2")
        url = reverse("fichas:registrar_hogar", kwargs={"pk": pendiente.pk})

        self.client.post(url, self.datos())

        pendiente.refresh_from_db()
        self.assertEqual(pendiente.estado, EstadoEncuesta.BORRADOR)

    def test_no_la_deja_completada(self):
        """Faltan los integrantes: darla por terminada engañaría al supervisor."""
        self.client.post(self.url, self.datos())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_volver_a_entrar_muestra_lo_ya_escrito(self):
        self.client.post(self.url, self.datos())

        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "Rosa Elena Millán")
        self.assertContains(respuesta, "Editar el grupo familiar")

    def test_editar_no_crea_un_segundo_hogar(self):
        self.client.post(self.url, self.datos())
        self.client.post(self.url, self.datos(jefe_hogar_nombre="Otra Persona"))

        self.assertEqual(GrupoFamiliar.objects.count(), 1)

    def test_datos_invalidos_no_guardan_nada(self):
        self.client.post(self.url, self.datos(integrantes_declarados=0))

        self.assertEqual(GrupoFamiliar.objects.count(), 0)

    def test_no_se_puede_escribir_en_la_encuesta_de_otra_persona(self):
        """Escribir en nombre de otro dejaría el dato atribuido a quien no estuvo."""
        ajena = self.crear(direccion="Calle de Juan", censista=self.juan)

        respuesta = self.client.post(
            reverse("fichas:registrar_hogar", kwargs={"pk": ajena.pk}), self.datos()
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_el_supervisor_tampoco_puede_aunque_vea_todas(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 302)

    def test_una_encuesta_validada_no_admite_cambios(self):
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url)

        self.assertRedirects(respuesta, self.url_detalle(self.encuesta))

    def test_una_encuesta_validada_tampoco_por_POST(self):
        """Ocultar el botón no es una validación: la URL se escribe a mano."""
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        self.client.post(self.url, self.datos())

        self.assertEqual(GrupoFamiliar.objects.count(), 0)

    def test_una_encuesta_observada_si_admite_cambios(self):
        """Es justamente el estado que existe para poder corregir."""
        self.devolver(self.encuesta)

        self.assertEqual(self.client.get(self.url).status_code, 200)

    def test_con_el_operativo_cerrado_no_se_puede_escribir(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.client.post(self.url, self.datos())

        self.assertEqual(GrupoFamiliar.objects.count(), 0)


# ==========================================================================
# HU-08 — 25. UN SEGUNDO HOGAR EN LA MISMA VIVIENDA
# ==========================================================================


class AgregarHogarTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        self.vivienda = self.encuesta.vivienda
        self.url = reverse("fichas:hogar_agregar", kwargs={"pk": self.vivienda.pk})
        self.client.force_login(self.marta)

    def test_crea_una_encuesta_mas_en_la_misma_vivienda(self):
        self.client.post(self.url)

        self.assertEqual(self.vivienda.encuestas.count(), 2)

    def test_no_duplica_la_vivienda(self):
        """Es el punto: la casa se describe una vez para los dos hogares."""
        self.client.post(self.url)

        self.assertEqual(Vivienda.objects.count(), 1)

    def test_lleva_al_formulario_del_hogar_nuevo(self):
        respuesta = self.client.post(self.url)

        nueva = self.vivienda.encuestas.exclude(pk=self.encuesta.pk).get()
        self.assertRedirects(
            respuesta, reverse("fichas:registrar_hogar", kwargs={"pk": nueva.pk})
        )

    def test_un_GET_no_crea_nada(self):
        """Con un GET, un <img src="..."> ajeno llenaría la base de encuestas."""
        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 405)
        self.assertEqual(self.vivienda.encuestas.count(), 1)

    def test_no_se_puede_en_una_vivienda_fuera_de_mi_territorio(self):
        ajena = Vivienda.objects.create(zona=self.zona_norte, direccion="Otra 1")

        respuesta = self.client.post(
            reverse("fichas:hogar_agregar", kwargs={"pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_no_se_puede_con_el_operativo_cerrado(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.client.post(self.url)

        self.assertEqual(self.vivienda.encuestas.count(), 1)


# ==========================================================================
# HU-08 — 26. LA FICHA DE LA VIVIENDA
# ==========================================================================


class ViviendaDetalleTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        self.vivienda = self.encuesta.vivienda
        self.url = reverse("fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk})
        self.client.force_login(self.marta)

    def test_muestra_las_caracteristicas(self):
        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Albañilería")

    def test_muestra_los_hogares(self):
        GrupoFamiliar.objects.create(
            encuesta=self.encuesta,
            jefe_hogar_nombre="Rosa Millán",
            integrantes_declarados=3,
        )

        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "Rosa Millán")

    def test_muestra_los_dos_hogares_de_una_vivienda_compartida(self):
        self.crear(vivienda=self.vivienda, censista=self.juan)

        respuesta = self.client.get(self.url)

        self.assertEqual(len(respuesta.context["hogares"]), 2)

    def test_una_vivienda_sin_describir_lo_dice(self):
        vivienda = Vivienda.objects.create(zona=self.zona1, direccion="Calle 9")

        respuesta = self.client.get(
            reverse("fichas:vivienda_detalle", kwargs={"pk": vivienda.pk})
        )

        self.assertContains(respuesta, "Sin describir")

    def test_ofrece_agregar_otro_hogar(self):
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "Agregar otro hogar")

    def test_con_el_operativo_cerrado_no_lo_ofrece(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        respuesta = self.client.get(self.url)

        self.assertNotContains(respuesta, "Agregar otro hogar")

    def test_el_supervisor_la_ve_pero_no_puede_registrar(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 200)
        self.assertNotContains(respuesta, "Agregar otro hogar")

    def test_una_vivienda_ajena_responde_404(self):
        ajena = Vivienda.objects.create(zona=self.zona_norte, direccion="Otra 1")

        respuesta = self.client.get(
            reverse("fichas:vivienda_detalle", kwargs={"pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)


# ==========================================================================
# HU-08 — 27. LO QUE LA HU-07 GANA CON LA HU-08
# ==========================================================================


class FichaConHogarTest(BaseEncuestaTest):
    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        self.client.force_login(self.marta)

    def test_la_ficha_ofrece_registrar_el_hogar_si_no_lo_tiene(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Registrar el grupo familiar")

    def test_la_ficha_muestra_el_hogar_cuando_existe(self):
        GrupoFamiliar.objects.create(
            encuesta=self.encuesta,
            jefe_hogar_nombre="Rosa Millán",
            integrantes_declarados=3,
            ingreso_mensual=600000,
        )

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Rosa Millán")
        self.assertContains(respuesta, "Editar los datos del hogar")

    def test_una_encuesta_cerrada_no_ofrece_editar(self):
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertNotContains(respuesta, "Editar los datos del hogar")

    def test_el_listado_ofrece_registrar_una_vivienda(self):
        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, reverse("fichas:vivienda_registrar"))

    def test_al_supervisor_no_se_le_ofrece_registrar(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_lista)

        self.assertNotContains(respuesta, reverse("fichas:vivienda_registrar"))


# ==========================================================================
# HU-08 — 28. RECORRIDO COMPLETO DE LA HISTORIA
# ==========================================================================


class IntegracionHU08Test(BaseEncuestaTest):
    """De la asignación del sector (HU-06) al hogar registrado (HU-08)."""

    def test_recorrido_completo(self):
        # 1. El supervisor reparte el sector (HU-06).
        AsignacionSector.objects.create(
            sector=self.boldos, censista=self.marta, asignado_por=self.supervisor
        )
        self.client.force_login(self.marta)

        # 2. La encuestadora registra una vivienda al llegar a la puerta.
        respuesta = self.client.post(
            reverse("fichas:vivienda_registrar"),
            {
                "zona": self.zona1.pk,
                "direccion": "Pasaje Los Robles 47",
                "referencia": "Sitio con dos casas",
                "tipo": TipoVivienda.CASA,
                "tenencia": TenenciaVivienda.PROPIA_PAGADA,
                "materialidad_muros": MaterialidadMuros.ALBANILERIA,
                "origen_agua": OrigenAgua.RED_PUBLICA,
                "sistema_sanitario": SistemaSanitario.ALCANTARILLADO,
                "tiene_electricidad": True,
                "observaciones": "",
            },
        )
        vivienda = Vivienda.objects.get()
        encuesta = Encuesta.objects.get()
        self.assertRedirects(
            respuesta, reverse("fichas:registrar_hogar", kwargs={"pk": encuesta.pk})
        )

        # 3. Registra el hogar.
        self.client.post(
            reverse("fichas:registrar_hogar", kwargs={"pk": encuesta.pk}),
            {
                "jefe_hogar_nombre": "Rosa Elena Millán",
                "jefe_hogar_rut": "12345678-5",
                "telefono_contacto": "",
                "integrantes_declarados": 5,
                "ingreso_mensual": 750000,
                "observaciones": "",
            },
        )
        encuesta.refresh_from_db()
        self.assertEqual(encuesta.grupo_familiar.integrantes_declarados, 5)
        self.assertEqual(encuesta.estado, EstadoEncuesta.BORRADOR)

        # 4. Resulta que en la misma casa vive una segunda familia.
        self.client.post(
            reverse("fichas:hogar_agregar", kwargs={"pk": vivienda.pk})
        )
        segunda = vivienda.encuestas.exclude(pk=encuesta.pk).get()

        self.assertEqual(Vivienda.objects.count(), 1)
        self.assertEqual(vivienda.encuestas.count(), 2)

        # 5. Las dos aparecen en su listado, y la casa se describió una sola vez.
        respuesta = self.client.get(self.url_lista)
        self.assertEqual(respuesta.context["resumen"]["total"], 2)
        self.assertEqual(segunda.vivienda, encuesta.vivienda)

        # 6. La ficha de una avisa de la otra.
        respuesta = self.client.get(self.url_detalle(encuesta))
        self.assertContains(respuesta, "Otros hogares en esta misma vivienda")

        # 7. Y Juan, que no tiene el sector, no puede tocar nada de esto.
        self.client.force_login(self.juan)
        self.assertEqual(
            self.client.get(
                reverse("fichas:vivienda_detalle", kwargs={"pk": vivienda.pk})
            ).status_code,
            404,
        )


# ==========================================================================
# HU-09 — 29. EL MODELO Integrante
# ==========================================================================


class BaseIntegranteTest(BaseEncuestaTest):
    """Escenario común: Marta con su sector asignado y una encuesta con hogar."""

    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        self.hogar = GrupoFamiliar.objects.create(
            encuesta=self.encuesta,
            jefe_hogar_nombre="Rosa Elena Millán",
            integrantes_declarados=4,
        )
        self.url_lista = reverse(
            "fichas:integrantes", kwargs={"encuesta_pk": self.encuesta.pk}
        )
        self.url_nuevo = reverse(
            "fichas:integrante_nuevo", kwargs={"encuesta_pk": self.encuesta.pk}
        )

    def nacido_hace(self, anios):
        """Una fecha de nacimiento que da esa edad hoy, sin caer en el cumpleaños."""
        return timezone.localdate() - timedelta(days=anios * 365 + 100)

    def persona(self, **extra):
        datos = {
            "grupo_familiar": self.hogar,
            "parentesco": Parentesco.HIJO,
            "nombres": "Camila Andrea",
            "apellidos": "Riquelme Soto",
            "sexo": Sexo.FEMENINO,
            "fecha_nacimiento": self.nacido_hace(20),
            "nivel_educacional": NivelEducacional.MEDIA_COMPLETA,
            "situacion_ocupacional": SituacionOcupacional.ESTUDIA,
        }
        datos.update(extra)
        return Integrante.objects.create(**datos)

    def jefe(self, **extra):
        datos = {"parentesco": Parentesco.JEFE_HOGAR, "nombres": "Rosa Elena"}
        datos.update(extra)
        return self.persona(**datos)

    def url_editar(self, integrante):
        return reverse(
            "fichas:integrante_editar",
            kwargs={"encuesta_pk": self.encuesta.pk, "pk": integrante.pk},
        )

    def url_quitar(self, integrante):
        return reverse(
            "fichas:integrante_quitar",
            kwargs={"encuesta_pk": self.encuesta.pk, "pk": integrante.pk},
        )


class IntegranteModeloTest(BaseIntegranteTest):
    def test_se_crea_asociado_a_su_hogar(self):
        persona = self.persona()

        self.assertIn(persona, self.hogar.integrantes.all())

    def test_el_nombre_completo_junta_nombres_y_apellidos(self):
        persona = self.persona(nombres="Ana", apellidos="Rojas")

        self.assertEqual(persona.nombre_completo, "Ana Rojas")
        self.assertEqual(str(persona), "Ana Rojas")

    def test_calcula_la_edad_desde_la_fecha_de_nacimiento(self):
        self.assertEqual(self.persona(fecha_nacimiento=self.nacido_hace(30)).edad(), 30)

    def test_la_edad_no_se_adelanta_antes_del_cumpleanos(self):
        """El caso borde que se escribe mal una vez por proyecto."""
        hoy = timezone.localdate()
        # Nació el mismo día del año pero un día después: todavía no los cumple.
        fecha = date(hoy.year - 30, hoy.month, hoy.day) + timedelta(days=1)

        persona = self.persona(fecha_nacimiento=fecha)

        self.assertEqual(persona.edad(), 29)

    def test_la_edad_se_puede_calcular_a_una_fecha_dada(self):
        """Guardar la fecha y no la edad permite mirar hacia atrás."""
        persona = self.persona(fecha_nacimiento=date(2000, 6, 15))

        self.assertEqual(persona.edad(a_fecha=date(2020, 6, 14)), 19)
        self.assertEqual(persona.edad(a_fecha=date(2020, 6, 15)), 20)

    def test_sabe_si_es_menor_de_edad(self):
        self.assertTrue(self.persona(fecha_nacimiento=self.nacido_hace(10)).es_menor_de_edad)
        self.assertFalse(
            self.persona(
                apellidos="Otro", fecha_nacimiento=self.nacido_hace(40)
            ).es_menor_de_edad
        )

    def test_sabe_a_quien_se_le_pregunta_escolaridad_y_ocupacion(self):
        bebe = Integrante(fecha_nacimiento=self.nacido_hace(2))
        escolar = Integrante(fecha_nacimiento=self.nacido_hace(9))
        adulto = Integrante(fecha_nacimiento=self.nacido_hace(30))

        self.assertFalse(bebe.se_le_pregunta_escolaridad)
        self.assertTrue(escolar.se_le_pregunta_escolaridad)
        self.assertFalse(escolar.se_le_pregunta_ocupacion)
        self.assertTrue(adulto.se_le_pregunta_ocupacion)

    def test_sabe_quien_es_el_jefe_de_hogar(self):
        jefe = self.jefe()

        self.assertTrue(jefe.es_jefe_hogar)
        self.assertFalse(self.persona(apellidos="Otro").es_jefe_hogar)

    def test_el_rut_se_normaliza_al_guardar(self):
        persona = self.persona(rut="12.345.678-5")

        persona.refresh_from_db()
        self.assertEqual(persona.rut, "12345678-5")

    def test_borrar_el_hogar_se_lleva_a_sus_integrantes(self):
        self.persona()

        self.hogar.delete()

        self.assertEqual(Integrante.objects.count(), 0)

    def test_borrar_la_encuesta_tambien_los_borra(self):
        """La cadena completa: encuesta -> hogar -> personas."""
        self.persona()

        self.encuesta.delete()

        self.assertEqual(Integrante.objects.count(), 0)


class IntegranteRestriccionesTest(BaseIntegranteTest):
    def test_no_puede_haber_dos_jefes_de_hogar(self):
        """Con dos, el parentesco de todos los demás dejaría de significar algo."""
        self.jefe()

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.persona(parentesco=Parentesco.JEFE_HOGAR, apellidos="Otra")

    def test_si_puede_haber_dos_hijos(self):
        """La unicidad es PARCIAL: solo afecta al jefe de hogar."""
        self.persona(nombres="Camila")
        self.persona(nombres="Matías", apellidos="Riquelme Soto")

        self.assertEqual(self.hogar.integrantes.count(), 2)

    def test_otro_hogar_puede_tener_su_propio_jefe(self):
        self.jefe()
        otra = self.crear(direccion="Calle 2", estado=EstadoEncuesta.BORRADOR)
        otro_hogar = GrupoFamiliar.objects.create(
            encuesta=otra, jefe_hogar_nombre="Otro", integrantes_declarados=1
        )

        Integrante.objects.create(
            grupo_familiar=otro_hogar,
            parentesco=Parentesco.JEFE_HOGAR,
            nombres="Otro",
            apellidos="Jefe",
            sexo=Sexo.MASCULINO,
            fecha_nacimiento=self.nacido_hace(40),
            nivel_educacional=NivelEducacional.MEDIA_COMPLETA,
            situacion_ocupacional=SituacionOcupacional.TRABAJA,
        )

        self.assertEqual(Integrante.objects.filter(parentesco="JEFE_HOGAR").count(), 2)

    def test_el_rut_no_se_puede_repetir_en_el_mismo_hogar(self):
        self.persona(rut="12345678-5")

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.persona(rut="12345678-5", apellidos="Otra")

    def test_dos_personas_sin_rut_no_chocan(self):
        """Sin la condición parcial, la cadena vacía las haría chocar entre sí."""
        self.persona(rut="", nombres="Una")
        self.persona(rut="", nombres="Otra")

        self.assertEqual(self.hogar.integrantes.count(), 2)

    def test_un_parentesco_inventado_lo_rechaza_la_base_de_datos(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.persona(parentesco="PRIMO_SEGUNDO")

    def test_un_sexo_inventado_lo_rechaza_la_base_de_datos(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.persona(sexo="INVENTADO")

    def test_una_fecha_futura_la_rechaza_la_validacion(self):
        """No puede estar en un CheckConstraint: dependería de la fecha de hoy."""
        persona = Integrante(
            grupo_familiar=self.hogar,
            parentesco=Parentesco.HIJO,
            nombres="Futura",
            apellidos="Persona",
            sexo=Sexo.FEMENINO,
            fecha_nacimiento=timezone.localdate() + timedelta(days=1),
        )

        with self.assertRaises(ValidationError) as error:
            persona.full_clean()

        self.assertIn("fecha_nacimiento", error.exception.message_dict)

    def test_mas_de_120_anios_se_rechaza(self):
        persona = Integrante(
            grupo_familiar=self.hogar,
            parentesco=Parentesco.HIJO,
            nombres="Muy",
            apellidos="Mayor",
            sexo=Sexo.FEMENINO,
            fecha_nacimiento=self.nacido_hace(130),
        )

        with self.assertRaises(ValidationError):
            persona.full_clean()


# ==========================================================================
# HU-09 — 30. EL RECUENTO DEL HOGAR
# ==========================================================================


class RecuentoDelHogarTest(BaseIntegranteTest):
    """Aquí es donde `integrantes_declarados` de la HU-08 cobra sentido."""

    def test_sin_nadie_registrado_faltan_todas(self):
        self.assertEqual(self.hogar.total_integrantes(), 0)
        self.assertEqual(self.hogar.integrantes_pendientes, 4)
        self.assertFalse(self.hogar.esta_completo)

    def test_cuenta_las_registradas(self):
        self.persona()
        self.persona(apellidos="Otra")

        self.assertEqual(self.hogar.total_integrantes(), 2)
        self.assertEqual(self.hogar.integrantes_pendientes, 2)

    def test_con_todas_registradas_el_hogar_esta_completo(self):
        for numero in range(4):
            self.persona(apellidos=f"Apellido {numero}")

        self.assertTrue(self.hogar.esta_completo)
        self.assertEqual(self.hogar.integrantes_pendientes, 0)

    def test_los_pendientes_nunca_son_negativos(self):
        """Devolver -2 obligaría a cada plantilla a acordarse del caso."""
        for numero in range(6):
            self.persona(apellidos=f"Apellido {numero}")

        self.assertEqual(self.hogar.integrantes_pendientes, 0)

    def test_registrar_mas_de_las_declaradas_es_una_discrepancia(self):
        """No es un error: aparece la abuela de la pieza del fondo."""
        for numero in range(6):
            self.persona(apellidos=f"Apellido {numero}")

        self.assertTrue(self.hogar.hay_discrepancia)
        self.assertTrue(self.hogar.esta_completo)

    def test_sin_discrepancia_cuando_coinciden(self):
        for numero in range(4):
            self.persona(apellidos=f"Apellido {numero}")

        self.assertFalse(self.hogar.hay_discrepancia)

    def test_el_jefe_de_hogar_registrado_se_encuentra(self):
        jefe = self.jefe()
        self.persona(apellidos="Otra")

        self.assertEqual(self.hogar.jefe_hogar_registrado, jefe)

    def test_sin_jefe_registrado_devuelve_none(self):
        self.persona()

        self.assertIsNone(self.hogar.jefe_hogar_registrado)

    def test_el_jefe_va_primero_en_la_lista(self):
        """El parentesco de los demás se entiende respecto a esa persona."""
        self.persona(nombres="Mayor", fecha_nacimiento=self.nacido_hace(80))
        jefe = self.jefe(fecha_nacimiento=self.nacido_hace(35))

        self.assertEqual(self.hogar.integrantes_ordenados().first(), jefe)

    def test_despues_del_jefe_van_de_mayor_a_menor(self):
        self.jefe(fecha_nacimiento=self.nacido_hace(35))
        self.persona(nombres="Joven", fecha_nacimiento=self.nacido_hace(10))
        self.persona(nombres="Mayor", apellidos="Otra", fecha_nacimiento=self.nacido_hace(70))

        nombres = [p.nombres for p in self.hogar.integrantes_ordenados()]

        self.assertEqual(nombres, ["Rosa Elena", "Mayor", "Joven"])


class NombreDelJefeTest(BaseIntegranteTest):
    """Los dos sitios donde vive el nombre del jefe de hogar, y su coherencia."""

    def test_sin_jefe_registrado_no_hay_nada_que_contradecir(self):
        self.assertTrue(self.hogar.nombre_del_jefe_coincide)

    def test_coincide_cuando_es_el_mismo_nombre(self):
        self.jefe(nombres="Rosa Elena", apellidos="Millán")

        self.assertTrue(self.hogar.nombre_del_jefe_coincide)

    def test_no_distingue_mayusculas_ni_espacios_de_mas(self):
        self.jefe(nombres="ROSA   elena", apellidos="millán")

        self.assertTrue(self.hogar.nombre_del_jefe_coincide)

    def test_avisa_cuando_no_coinciden(self):
        self.jefe(nombres="Otra", apellidos="Persona")

        self.assertFalse(self.hogar.nombre_del_jefe_coincide)


# ==========================================================================
# HU-09 — 31. EL FORMULARIO
# ==========================================================================


class IntegranteFormTest(BaseIntegranteTest):
    def datos(self, **extra):
        base = {
            "parentesco": Parentesco.HIJO,
            "nombres": "Camila Andrea",
            "apellidos": "Riquelme Soto",
            "rut": "",
            "sexo": Sexo.FEMENINO,
            "fecha_nacimiento": self.nacido_hace(20).isoformat(),
            "nivel_educacional": NivelEducacional.MEDIA_COMPLETA,
            "situacion_ocupacional": SituacionOcupacional.ESTUDIA,
            "pueblo_originario": PuebloOriginario.NINGUNO,
            "observaciones": "",
        }
        base.update(extra)
        return base

    def formulario(self, datos=None, **kwargs):
        return IntegranteForm(datos, grupo_familiar=self.hogar, **kwargs)

    def test_un_formulario_completo_es_valido(self):
        formulario = self.formulario(self.datos())

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_nombre_y_el_apellido_son_obligatorios(self):
        for campo in ("nombres", "apellidos"):
            with self.subTest(campo=campo):
                formulario = self.formulario(self.datos(**{campo: ""}))
                self.assertFalse(formulario.is_valid())

    def test_una_inicial_no_es_un_nombre(self):
        formulario = self.formulario(self.datos(nombres="C"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("nombres", formulario.errors)

    def test_la_fecha_de_nacimiento_es_obligatoria(self):
        formulario = self.formulario(self.datos(fecha_nacimiento=""))

        self.assertFalse(formulario.is_valid())

    def test_una_fecha_futura_se_rechaza(self):
        futura = (timezone.localdate() + timedelta(days=1)).isoformat()

        formulario = self.formulario(self.datos(fecha_nacimiento=futura))

        self.assertFalse(formulario.is_valid())
        self.assertIn("fecha_nacimiento", formulario.errors)

    # -- las reglas que dependen de la edad --------------------------------

    def test_a_un_bebe_no_se_le_pide_escolaridad_ni_ocupacion(self):
        formulario = self.formulario(
            self.datos(
                fecha_nacimiento=self.nacido_hace(2).isoformat(),
                nivel_educacional="",
                situacion_ocupacional="",
            )
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_desde_los_cinco_anios_la_escolaridad_es_obligatoria(self):
        formulario = self.formulario(
            self.datos(
                fecha_nacimiento=self.nacido_hace(9).isoformat(),
                nivel_educacional="",
                situacion_ocupacional="",
            )
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("nivel_educacional", formulario.errors)

    def test_a_un_nino_de_nueve_no_se_le_pide_ocupacion(self):
        formulario = self.formulario(
            self.datos(
                fecha_nacimiento=self.nacido_hace(9).isoformat(),
                nivel_educacional=NivelEducacional.BASICA_INCOMPLETA,
                situacion_ocupacional="",
            )
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_desde_los_quince_la_ocupacion_es_obligatoria(self):
        formulario = self.formulario(
            self.datos(
                fecha_nacimiento=self.nacido_hace(17).isoformat(),
                situacion_ocupacional="",
            )
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("situacion_ocupacional", formulario.errors)

    def test_sin_fecha_no_se_exigen_los_campos_que_dependen_de_ella(self):
        """El error de la fecha ya se informó: no hay que apilar tres más."""
        formulario = self.formulario(
            self.datos(
                fecha_nacimiento="", nivel_educacional="", situacion_ocupacional=""
            )
        )

        self.assertFalse(formulario.is_valid())
        self.assertNotIn("nivel_educacional", formulario.errors)

    # -- el jefe de hogar --------------------------------------------------

    def test_el_primero_puede_ser_jefe_de_hogar(self):
        valores = [v for v, _ in self.formulario().fields["parentesco"].choices]

        self.assertIn(Parentesco.JEFE_HOGAR, valores)

    def test_con_jefe_ya_registrado_la_opcion_desaparece(self):
        self.jefe()

        valores = [v for v, _ in self.formulario().fields["parentesco"].choices]

        self.assertNotIn(Parentesco.JEFE_HOGAR, valores)

    def test_enviar_jefe_de_hogar_a_mano_no_sirve(self):
        self.jefe()

        formulario = self.formulario(self.datos(parentesco=Parentesco.JEFE_HOGAR))

        self.assertFalse(formulario.is_valid())

    def test_al_editar_al_jefe_la_opcion_sigue_disponible(self):
        """Si no, guardar cualquier otro cambio lo dejaría sin parentesco."""
        jefe = self.jefe()

        formulario = self.formulario(instance=jefe)
        valores = [v for v, _ in formulario.fields["parentesco"].choices]

        self.assertIn(Parentesco.JEFE_HOGAR, valores)

    # -- el RUT ------------------------------------------------------------

    def test_un_rut_invalido_se_rechaza(self):
        formulario = self.formulario(self.datos(rut="12345678-9"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("rut", formulario.errors)

    def test_un_rut_repetido_en_el_hogar_se_rechaza_con_mensaje(self):
        self.persona(rut="12345678-5")

        formulario = self.formulario(self.datos(rut="12345678-5"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("rut", formulario.errors)

    def test_el_rut_repetido_se_detecta_aunque_venga_con_puntos(self):
        """Sin normalizar antes de comparar, pasaría el formulario y reventaría la base."""
        self.persona(rut="12345678-5")

        formulario = self.formulario(self.datos(rut="12.345.678-5"))

        self.assertFalse(formulario.is_valid())

    def test_editar_a_una_persona_no_choca_con_su_propio_rut(self):
        persona = self.persona(rut="12345678-5")

        formulario = self.formulario(self.datos(rut="12345678-5"), instance=persona)

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_mismo_rut_en_otro_hogar_no_estorba(self):
        self.persona(rut="12345678-5")
        otra = self.crear(direccion="Calle 2", estado=EstadoEncuesta.BORRADOR)
        otro_hogar = GrupoFamiliar.objects.create(
            encuesta=otra, jefe_hogar_nombre="Otro", integrantes_declarados=1
        )

        formulario = IntegranteForm(
            self.datos(rut="12345678-5"), grupo_familiar=otro_hogar
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)


# ==========================================================================
# HU-09 — 32. LA PANTALLA DE INTEGRANTES
# ==========================================================================


class IntegrantesPantallaTest(BaseIntegranteTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def test_muestra_la_lista(self):
        self.persona(nombres="Camila Andrea")

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Camila Andrea")

    def test_muestra_el_avance(self):
        self.persona()

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "1 de 4")
        self.assertContains(respuesta, "Faltan")

    def test_avisa_cuando_el_hogar_esta_completo(self):
        for numero in range(4):
            self.persona(apellidos=f"Apellido {numero}")

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "El hogar está completo")

    def test_avisa_de_la_discrepancia(self):
        for numero in range(6):
            self.persona(apellidos=f"Apellido {numero}")

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "personas registradas y la")

    def test_sin_nadie_invita_a_empezar_por_el_jefe_de_hogar(self):
        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "Todavía no hay nadie registrado")
        self.assertContains(respuesta, "Rosa Elena Millán")

    def test_avisa_si_el_nombre_del_jefe_no_coincide(self):
        self.jefe(nombres="Otra", apellidos="Persona")

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "está registrado a nombre de")

    def test_sin_hogar_registrado_manda_a_registrarlo_primero(self):
        """El parentesco se declara respecto al jefe de hogar: hay un orden real."""
        otra = self.crear(direccion="Calle 2", estado=EstadoEncuesta.BORRADOR)
        url = reverse("fichas:integrantes", kwargs={"encuesta_pk": otra.pk})

        respuesta = self.client.get(url)

        self.assertRedirects(
            respuesta, reverse("fichas:registrar_hogar", kwargs={"pk": otra.pk})
        )

    def test_la_encuesta_de_otra_persona_responde_404(self):
        ajena = self.crear(
            direccion="Calle de Juan", censista=self.juan, estado=EstadoEncuesta.BORRADOR
        )
        GrupoFamiliar.objects.create(
            encuesta=ajena, jefe_hogar_nombre="Otro", integrantes_declarados=1
        )

        respuesta = self.client.get(
            reverse("fichas:integrantes", kwargs={"encuesta_pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_el_supervisor_no_entra_aunque_vea_todas(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.status_code, 302)

    def test_con_la_encuesta_cerrada_no_se_ofrecen_acciones(self):
        self.persona()
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_lista)

        self.assertNotContains(respuesta, "Agregar una persona")


# ==========================================================================
# HU-09 — 33. AGREGAR UNA PERSONA
# ==========================================================================


class RegistrarIntegranteTest(BaseIntegranteTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "parentesco": Parentesco.HIJO,
            "nombres": "Camila Andrea",
            "apellidos": "Riquelme Soto",
            "rut": "",
            "sexo": Sexo.FEMENINO,
            "fecha_nacimiento": self.nacido_hace(20).isoformat(),
            "nivel_educacional": NivelEducacional.MEDIA_COMPLETA,
            "situacion_ocupacional": SituacionOcupacional.ESTUDIA,
            "pueblo_originario": PuebloOriginario.NINGUNO,
            "observaciones": "",
        }
        base.update(extra)
        return base

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url_nuevo)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Agregar una persona")

    def test_guarda_a_la_persona(self):
        self.client.post(self.url_nuevo, self.datos())

        self.assertEqual(self.hogar.integrantes.count(), 1)

    def test_la_asocia_al_hogar_correcto(self):
        self.client.post(self.url_nuevo, self.datos())

        self.assertEqual(Integrante.objects.get().grupo_familiar, self.hogar)

    def test_vuelve_a_la_lista(self):
        respuesta = self.client.post(self.url_nuevo, self.datos())

        self.assertRedirects(respuesta, self.url_lista)

    def test_el_mensaje_dice_cuantas_faltan(self):
        respuesta = self.client.post(self.url_nuevo, self.datos(), follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("Faltan 3" in m for m in mensajes))

    def test_el_mensaje_avisa_cuando_el_hogar_queda_completo(self):
        for numero in range(3):
            self.persona(apellidos=f"Apellido {numero}")

        respuesta = self.client.post(self.url_nuevo, self.datos(), follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("completo" in m for m in mensajes))

    def test_guardar_y_seguir_devuelve_al_formulario(self):
        """Registrar seis personas seguidas no puede costar dieciocho toques."""
        respuesta = self.client.post(
            self.url_nuevo, {**self.datos(), "guardar_y_seguir": "1"}
        )

        self.assertRedirects(respuesta, self.url_nuevo)
        self.assertEqual(self.hogar.integrantes.count(), 1)

    def test_datos_invalidos_no_guardan_nada(self):
        self.client.post(self.url_nuevo, self.datos(nombres=""))

        self.assertEqual(Integrante.objects.count(), 0)

    # -- el prellenado de la primera persona -------------------------------

    def test_la_primera_persona_viene_prellenada_como_jefa_de_hogar(self):
        respuesta = self.client.get(self.url_nuevo)
        inicial = respuesta.context["form"].initial

        self.assertEqual(inicial["parentesco"], Parentesco.JEFE_HOGAR)
        self.assertIn("Rosa", inicial["nombres"])

    def test_el_prellenado_arrastra_el_rut_del_hogar(self):
        self.hogar.jefe_hogar_rut = "12345678-5"
        self.hogar.save()

        respuesta = self.client.get(self.url_nuevo)

        self.assertEqual(respuesta.context["form"].initial["rut"], "12345678-5")

    def test_con_jefe_ya_registrado_no_se_prellena_nada(self):
        self.jefe()

        respuesta = self.client.get(self.url_nuevo)

        self.assertEqual(respuesta.context["form"].initial, {})

    def test_se_avisa_de_que_el_prellenado_es_un_borrador(self):
        respuesta = self.client.get(self.url_nuevo)

        self.assertContains(respuesta, "Corrige lo que haga falta")

    # -- acceso ------------------------------------------------------------

    def test_no_se_puede_agregar_a_la_encuesta_de_otra_persona(self):
        ajena = self.crear(
            direccion="Calle de Juan", censista=self.juan, estado=EstadoEncuesta.BORRADOR
        )
        GrupoFamiliar.objects.create(
            encuesta=ajena, jefe_hogar_nombre="Otro", integrantes_declarados=1
        )

        respuesta = self.client.post(
            reverse("fichas:integrante_nuevo", kwargs={"encuesta_pk": ajena.pk}),
            self.datos(),
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_con_la_encuesta_validada_no_se_puede_agregar(self):
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        self.client.post(self.url_nuevo, self.datos())

        self.assertEqual(Integrante.objects.count(), 0)

    def test_con_una_encuesta_observada_si_se_puede(self):
        self.devolver(self.encuesta)

        self.client.post(self.url_nuevo, self.datos())

        self.assertEqual(Integrante.objects.count(), 1)

    def test_con_el_operativo_cerrado_no_se_puede(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.client.post(self.url_nuevo, self.datos())

        self.assertEqual(Integrante.objects.count(), 0)

    def test_el_supervisor_no_puede_agregar_personas(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.post(self.url_nuevo, self.datos())

        self.assertEqual(respuesta.status_code, 302)
        self.assertEqual(Integrante.objects.count(), 0)


# ==========================================================================
# HU-09 — 34. EDITAR Y QUITAR
# ==========================================================================


class EditarIntegranteTest(BaseIntegranteTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.persona_creada = self.persona()

    def datos(self, **extra):
        base = {
            "parentesco": Parentesco.HIJO,
            "nombres": "Camila Corregida",
            "apellidos": "Riquelme Soto",
            "rut": "",
            "sexo": Sexo.FEMENINO,
            "fecha_nacimiento": self.nacido_hace(21).isoformat(),
            "nivel_educacional": NivelEducacional.TECNICA,
            "situacion_ocupacional": SituacionOcupacional.TRABAJA,
            "pueblo_originario": PuebloOriginario.MAPUCHE,
            "observaciones": "",
        }
        base.update(extra)
        return base

    def test_muestra_los_datos_actuales(self):
        respuesta = self.client.get(self.url_editar(self.persona_creada))

        self.assertContains(respuesta, "Camila Andrea")

    def test_guarda_los_cambios(self):
        self.client.post(self.url_editar(self.persona_creada), self.datos())

        self.persona_creada.refresh_from_db()
        self.assertEqual(self.persona_creada.nombres, "Camila Corregida")
        self.assertEqual(
            self.persona_creada.pueblo_originario, PuebloOriginario.MAPUCHE
        )

    def test_no_crea_una_persona_nueva(self):
        self.client.post(self.url_editar(self.persona_creada), self.datos())

        self.assertEqual(Integrante.objects.count(), 1)

    def test_una_persona_de_otro_hogar_responde_404(self):
        """La URL lleva la encuesta, así que el filtro por dueño va siempre."""
        otra = self.crear(
            direccion="Calle de Juan", censista=self.juan, estado=EstadoEncuesta.BORRADOR
        )
        otro_hogar = GrupoFamiliar.objects.create(
            encuesta=otra, jefe_hogar_nombre="Otro", integrantes_declarados=1
        )
        ajena = Integrante.objects.create(
            grupo_familiar=otro_hogar,
            parentesco=Parentesco.JEFE_HOGAR,
            nombres="Ajena",
            apellidos="Persona",
            sexo=Sexo.FEMENINO,
            fecha_nacimiento=self.nacido_hace(40),
            nivel_educacional=NivelEducacional.MEDIA_COMPLETA,
            situacion_ocupacional=SituacionOcupacional.TRABAJA,
        )

        respuesta = self.client.get(
            reverse(
                "fichas:integrante_editar",
                kwargs={"encuesta_pk": self.encuesta.pk, "pk": ajena.pk},
            )
        )

        self.assertEqual(respuesta.status_code, 404)


class QuitarIntegranteTest(BaseIntegranteTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.persona_creada = self.persona()

    def test_el_get_solo_pide_confirmacion(self):
        """Si un GET borrara, un <img src> ajeno lo ejecutaría con tu sesión."""
        respuesta = self.client.get(self.url_quitar(self.persona_creada))

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(Integrante.objects.count(), 1)

    def test_avisa_de_que_los_datos_se_borran(self):
        respuesta = self.client.get(self.url_quitar(self.persona_creada))

        self.assertContains(respuesta, "se borran, no se archivan")

    def test_el_post_la_quita(self):
        self.client.post(self.url_quitar(self.persona_creada))

        self.assertEqual(Integrante.objects.count(), 0)

    def test_vuelve_a_la_lista(self):
        respuesta = self.client.post(self.url_quitar(self.persona_creada))

        self.assertRedirects(respuesta, self.url_lista)

    def test_quitar_al_jefe_avisa_de_que_el_hogar_queda_sin_jefe(self):
        jefe = self.jefe(apellidos="Millán")

        respuesta = self.client.post(self.url_quitar(jefe), follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("sin jefe de hogar" in m for m in mensajes))

    def test_al_confirmar_se_avisa_si_es_el_jefe(self):
        jefe = self.jefe(apellidos="Millán")

        respuesta = self.client.get(self.url_quitar(jefe))

        self.assertContains(respuesta, "jefa de hogar")

    def test_con_la_encuesta_cerrada_no_se_puede_quitar(self):
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        self.client.post(self.url_quitar(self.persona_creada))

        self.assertEqual(Integrante.objects.count(), 1)

    def test_no_se_puede_quitar_a_alguien_de_otro_encuestador(self):
        otra = self.crear(
            direccion="Calle de Juan", censista=self.juan, estado=EstadoEncuesta.BORRADOR
        )
        otro_hogar = GrupoFamiliar.objects.create(
            encuesta=otra, jefe_hogar_nombre="Otro", integrantes_declarados=1
        )
        ajena = Integrante.objects.create(
            grupo_familiar=otro_hogar,
            parentesco=Parentesco.JEFE_HOGAR,
            nombres="Ajena",
            apellidos="Persona",
            sexo=Sexo.FEMENINO,
            fecha_nacimiento=self.nacido_hace(40),
            nivel_educacional=NivelEducacional.MEDIA_COMPLETA,
            situacion_ocupacional=SituacionOcupacional.TRABAJA,
        )

        respuesta = self.client.post(
            reverse(
                "fichas:integrante_quitar",
                kwargs={"encuesta_pk": otra.pk, "pk": ajena.pk},
            )
        )

        self.assertEqual(respuesta.status_code, 404)
        self.assertEqual(Integrante.objects.count(), 2)


# ==========================================================================
# HU-09 — 35. LO QUE GANAN LAS PANTALLAS ANTERIORES
# ==========================================================================


class FichaConIntegrantesTest(BaseIntegranteTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def test_la_ficha_muestra_cuantas_personas_van(self):
        self.persona()

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "1 de")
        self.assertContains(respuesta, "faltan 3")

    def test_la_ficha_enlaza_la_pantalla_de_integrantes(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, self.url_lista)

    def test_la_ficha_dice_cuando_el_hogar_esta_completo(self):
        for numero in range(4):
            self.persona(apellidos=f"Apellido {numero}")

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "completo")


class ConsultasIntegrantesTest(BaseIntegranteTest):
    """La lista no debe pagar una consulta por persona."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def contar(self):
        with CaptureQueriesContext(connection) as captura:
            self.client.get(self.url_lista)
        return len(captura.captured_queries)

    def test_el_numero_de_consultas_no_crece_con_las_personas(self):
        self.persona(apellidos="Uno")
        con_una = self.contar()

        self.persona(apellidos="Dos")
        self.persona(apellidos="Tres")
        con_tres = self.contar()

        self.assertEqual(con_una, con_tres)


# ==========================================================================
# HU-09 — 36. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionHU09Test(BaseIntegranteTest):
    """Del hogar registrado (HU-08) al hogar enumerado (HU-09)."""

    def test_recorrido_completo(self):
        self.client.force_login(self.marta)

        # 1. Al entrar no hay nadie y el sistema propone empezar por el jefe.
        respuesta = self.client.get(self.url_lista)
        self.assertContains(respuesta, "Todavía no hay nadie registrado")

        # 2. La primera persona viene prellenada como jefa de hogar.
        respuesta = self.client.get(self.url_nuevo)
        self.assertEqual(
            respuesta.context["form"].initial["parentesco"], Parentesco.JEFE_HOGAR
        )

        # 3. Se registra a la jefa de hogar.
        self.client.post(
            self.url_nuevo,
            {
                "parentesco": Parentesco.JEFE_HOGAR,
                "nombres": "Rosa Elena",
                "apellidos": "Millán",
                "rut": "12345678-5",
                "sexo": Sexo.FEMENINO,
                "fecha_nacimiento": self.nacido_hace(42).isoformat(),
                "nivel_educacional": NivelEducacional.MEDIA_COMPLETA,
                "situacion_ocupacional": SituacionOcupacional.TRABAJA,
                "pueblo_originario": PuebloOriginario.MAPUCHE,
                "observaciones": "",
                "guardar_y_seguir": "1",
            },
        )
        self.assertTrue(self.hogar.nombre_del_jefe_coincide)

        # 4. Y a una hija de 3 años, a quien no se le pide escolaridad.
        self.client.post(
            self.url_nuevo,
            {
                "parentesco": Parentesco.HIJO,
                "nombres": "Emilia Paz",
                "apellidos": "Millán",
                "rut": "",
                "sexo": Sexo.FEMENINO,
                "fecha_nacimiento": self.nacido_hace(3).isoformat(),
                "nivel_educacional": "",
                "situacion_ocupacional": "",
                "pueblo_originario": PuebloOriginario.NINGUNO,
                "observaciones": "",
            },
        )
        self.assertEqual(self.hogar.total_integrantes(), 2)

        # 5. Ya no se puede registrar a un segundo jefe de hogar.
        valores = [
            v
            for v, _ in IntegranteForm(grupo_familiar=self.hogar)
            .fields["parentesco"]
            .choices
        ]
        self.assertNotIn(Parentesco.JEFE_HOGAR, valores)

        # 6. La lista pone a la jefa primero y avisa de lo que falta.
        respuesta = self.client.get(self.url_lista)
        self.assertEqual(
            respuesta.context["integrantes"][0].parentesco, Parentesco.JEFE_HOGAR
        )
        self.assertContains(respuesta, "Faltan")

        # 7. Se quita a alguien agregado por error y el recuento lo refleja.
        sobrante = self.persona(nombres="Agregada", apellidos="Por Error")
        self.assertEqual(self.hogar.total_integrantes(), 3)
        self.client.post(self.url_quitar(sobrante))
        self.assertEqual(self.hogar.total_integrantes(), 2)

        # 8. Y Juan sigue sin poder ver ni tocar nada de este hogar.
        self.client.force_login(self.juan)
        self.assertEqual(self.client.get(self.url_lista).status_code, 404)


# ==========================================================================
# HU-10 — 37. ¿QUÉ FALTA PARA TERMINAR?
# ==========================================================================


class BaseBorradorTest(BaseEncuestaTest):
    """Escenario común: una encuesta en borrador que se va completando por pasos."""

    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)
        self.url_borrador = reverse(
            "fichas:guardar_borrador", kwargs={"pk": self.encuesta.pk}
        )
        self.url_completar = reverse(
            "fichas:completar_encuesta", kwargs={"pk": self.encuesta.pk}
        )
        self.url_cerrar = reverse(
            "fichas:cerrar_encuesta", kwargs={"pk": self.encuesta.pk}
        )

    def nacido_hace(self, anios):
        return timezone.localdate() - timedelta(days=anios * 365 + 100)

    def con_hogar(self, declarados=2):
        return GrupoFamiliar.objects.create(
            encuesta=self.encuesta,
            jefe_hogar_nombre="Rosa Elena Millán",
            integrantes_declarados=declarados,
        )

    def con_persona(self, hogar, parentesco=Parentesco.JEFE_HOGAR, **extra):
        datos = {
            "grupo_familiar": hogar,
            "parentesco": parentesco,
            "nombres": "Rosa Elena",
            "apellidos": "Millán",
            "sexo": Sexo.FEMENINO,
            "fecha_nacimiento": self.nacido_hace(40),
            "nivel_educacional": NivelEducacional.MEDIA_COMPLETA,
            "situacion_ocupacional": SituacionOcupacional.TRABAJA,
        }
        datos.update(extra)
        return Integrante.objects.create(**datos)

    def completar_todo(self, declarados=1):
        """Deja la encuesta lista para poder terminarse."""
        hogar = self.con_hogar(declarados=declarados)
        self.con_persona(hogar)
        return hogar


class PasosPendientesTest(BaseBorradorTest):
    def test_una_encuesta_recien_creada_tiene_pasos_pendientes(self):
        self.assertTrue(self.encuesta.pasos_pendientes())
        self.assertFalse(self.encuesta.puede_completarse)

    def test_el_primer_paso_es_describir_la_vivienda_si_falta(self):
        vivienda = Vivienda.objects.create(zona=self.zona1, direccion="Sin describir 1")
        encuesta = self.crear(vivienda=vivienda, estado=EstadoEncuesta.BORRADOR)

        pasos = encuesta.pasos_pendientes()

        self.assertIn("Describir la vivienda", pasos[0]["texto"])

    def test_con_la_vivienda_descrita_el_paso_es_el_hogar(self):
        pasos = self.encuesta.pasos_pendientes()

        self.assertEqual(len(pasos), 1)
        self.assertIn("hogar", pasos[0]["texto"])

    def test_sin_hogar_no_se_listan_pasos_de_personas(self):
        """Tres pasos que dicen lo mismo no ayudan: se corta en el que bloquea."""
        pasos = self.encuesta.pasos_pendientes()

        self.assertEqual(len(pasos), 1)

    def test_con_hogar_sin_jefe_pide_el_jefe(self):
        hogar = self.con_hogar()
        self.con_persona(hogar, parentesco=Parentesco.HIJO, nombres="Hija")

        textos = [p["texto"] for p in self.encuesta.pasos_pendientes()]

        self.assertTrue(any("jefe de hogar" in t for t in textos))

    def test_con_personas_faltantes_dice_cuantas(self):
        hogar = self.con_hogar(declarados=4)
        self.con_persona(hogar)

        textos = [p["texto"] for p in self.encuesta.pasos_pendientes()]

        self.assertTrue(any("3 personas" in t for t in textos))

    def test_con_todo_registrado_no_falta_nada(self):
        self.completar_todo()

        self.assertEqual(self.encuesta.pasos_pendientes(), [])
        self.assertTrue(self.encuesta.puede_completarse)

    def test_cada_paso_trae_su_ruta_y_su_argumento(self):
        """Es lo que permite enlazar «Ir» en vez de decir «no puedes»."""
        for paso in self.encuesta.pasos_pendientes():
            with self.subTest(paso=paso["texto"]):
                self.assertIn("ruta", paso)
                self.assertIsNotNone(paso["argumento"])
                # La ruta tiene que existir de verdad.
                reverse(paso["ruta"], args=[paso["argumento"]])

    def test_registrar_mas_personas_de_las_declaradas_no_deja_pasos(self):
        hogar = self.con_hogar(declarados=1)
        self.con_persona(hogar)
        self.con_persona(hogar, parentesco=Parentesco.HIJO, nombres="Extra")

        self.assertTrue(self.encuesta.puede_completarse)

    def test_puede_completarse_no_mira_el_estado(self):
        """Datos completos y encuesta cerrada son preguntas distintas."""
        self.completar_todo()
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        self.assertTrue(self.encuesta.puede_completarse)
        permitido, _ = self.encuesta.puede_registrarse()
        self.assertFalse(permitido)


class VisitaVencidaTest(BaseBorradorTest):
    def test_sin_fecha_anotada_no_hay_visita_vencida(self):
        self.assertFalse(self.encuesta.visita_pendiente_vencida)

    def test_una_fecha_futura_no_esta_vencida(self):
        self.encuesta.proxima_visita = timezone.localdate() + timedelta(days=2)

        self.assertFalse(self.encuesta.visita_pendiente_vencida)

    def test_la_fecha_de_hoy_ya_cuenta_como_vencida(self):
        """«Vuelvo esta tarde» tiene que aparecer en la lista de hoy."""
        self.encuesta.proxima_visita = timezone.localdate()

        self.assertTrue(self.encuesta.visita_pendiente_vencida)

    def test_una_fecha_pasada_esta_vencida(self):
        self.encuesta.proxima_visita = timezone.localdate() - timedelta(days=3)

        self.assertTrue(self.encuesta.visita_pendiente_vencida)

    def test_una_encuesta_cerrada_no_avisa_de_visitas(self):
        """Ya no espera a nadie: seguir avisando sería ruido."""
        self.completar_todo()
        self.encuesta.proxima_visita = timezone.localdate() - timedelta(days=3)
        self.encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)

        self.assertFalse(self.encuesta.visita_pendiente_vencida)


# ==========================================================================
# HU-10 — 38. LA RESTRICCIÓN DEL MOTIVO DE CIERRE
# ==========================================================================


class MotivoDeCierreTest(BaseBorradorTest):
    def test_cerrar_sin_levantar_exige_motivo(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(direccion="Otra 1"),
                    censista=self.marta,
                    estado=EstadoEncuesta.NO_UBICADA,
                    iniciada_en=timezone.now(),
                    cerrada_en=timezone.now(),
                    motivo_cierre="",
                )

    def test_con_motivo_si_se_puede_guardar(self):
        encuesta = Encuesta.objects.create(
            vivienda=self.crear_vivienda(direccion="Otra 1"),
            censista=self.marta,
            estado=EstadoEncuesta.NO_UBICADA,
            iniciada_en=timezone.now(),
            cerrada_en=timezone.now(),
            motivo_cierre="La dirección no existe.",
        )

        self.assertEqual(encuesta.estado, EstadoEncuesta.NO_UBICADA)

    def test_los_demas_estados_no_exigen_motivo(self):
        """Solo NO_UBICADA y RECHAZADA: una completada no necesita explicación."""
        for estado in (
            EstadoEncuesta.PENDIENTE,
            EstadoEncuesta.BORRADOR,
            EstadoEncuesta.COMPLETADA,
            EstadoEncuesta.VALIDADA,
            EstadoEncuesta.OBSERVADA,
        ):
            with self.subTest(estado=estado):
                encuesta = self.crear(direccion=f"Calle {estado}", estado=estado)
                self.assertEqual(encuesta.motivo_cierre, "")

    def test_rechazada_tambien_exige_motivo(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(direccion="Otra 2"),
                    censista=self.marta,
                    estado=EstadoEncuesta.RECHAZADA,
                    iniciada_en=timezone.now(),
                    cerrada_en=timezone.now(),
                )


# ==========================================================================
# HU-10 — 39. LOS FORMULARIOS
# ==========================================================================


class BorradorFormTest(BaseBorradorTest):
    def test_los_dos_campos_son_opcionales(self):
        """Guardar el borrador sin escribir nada no es un error."""
        formulario = BorradorForm(
            {"nota_avance": "", "proxima_visita": ""}, instance=self.encuesta
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_guarda_la_nota(self):
        formulario = BorradorForm(
            {"nota_avance": "Falta el módulo de ingresos.", "proxima_visita": ""},
            instance=self.encuesta,
        )
        formulario.is_valid()
        formulario.save()

        self.encuesta.refresh_from_db()
        self.assertIn("ingresos", self.encuesta.nota_avance)

    def test_acepta_una_fecha_futura(self):
        manana = (timezone.localdate() + timedelta(days=1)).isoformat()

        formulario = BorradorForm(
            {"nota_avance": "", "proxima_visita": manana}, instance=self.encuesta
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_acepta_hoy(self):
        """«Vuelvo esta tarde» es un caso real."""
        formulario = BorradorForm(
            {"nota_avance": "", "proxima_visita": timezone.localdate().isoformat()},
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_rechaza_una_fecha_pasada(self):
        """Una fecha pasada no es una cita, es un olvido."""
        ayer = (timezone.localdate() - timedelta(days=1)).isoformat()

        formulario = BorradorForm(
            {"nota_avance": "", "proxima_visita": ayer}, instance=self.encuesta
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("proxima_visita", formulario.errors)


class CerrarSinDatosFormTest(BaseBorradorTest):
    def test_solo_ofrece_los_dos_estados_sin_levantar(self):
        formulario = CerrarSinDatosForm()
        valores = [valor for valor, _ in formulario.fields["estado"].choices]

        self.assertEqual(set(valores), set(ESTADOS_SIN_LEVANTAR))

    def test_no_ofrece_completada_ni_validada(self):
        """Esta pantalla no es un atajo para terminar la encuesta."""
        formulario = CerrarSinDatosForm()
        valores = [valor for valor, _ in formulario.fields["estado"].choices]

        self.assertNotIn(EstadoEncuesta.COMPLETADA, valores)
        self.assertNotIn(EstadoEncuesta.VALIDADA, valores)

    def test_un_motivo_completo_es_valido(self):
        formulario = CerrarSinDatosForm(
            {
                "estado": EstadoEncuesta.NO_UBICADA,
                "motivo_cierre": "La dirección no existe, el pasaje llega hasta el 40.",
            },
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_motivo_es_obligatorio(self):
        formulario = CerrarSinDatosForm(
            {"estado": EstadoEncuesta.NO_UBICADA, "motivo_cierre": ""},
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("motivo_cierre", formulario.errors)

    def test_un_motivo_de_una_letra_no_sirve(self):
        """La restricción de la base solo exige no vacío; el lector exige legible."""
        formulario = CerrarSinDatosForm(
            {"estado": EstadoEncuesta.NO_UBICADA, "motivo_cierre": "x"},
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("motivo_cierre", formulario.errors)

    def test_el_estado_es_obligatorio(self):
        formulario = CerrarSinDatosForm(
            {"estado": "", "motivo_cierre": "La dirección no existe en el pasaje."},
        )

        self.assertFalse(formulario.is_valid())


# ==========================================================================
# HU-10 — 40. GUARDAR EL BORRADOR
# ==========================================================================


class GuardarBorradorTest(BaseBorradorTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {"nota_avance": "Falta el módulo de ingresos.", "proxima_visita": ""}
        base.update(extra)
        return base

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url_borrador)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Guardar para continuar")

    def test_explica_que_los_datos_ya_estaban_guardados(self):
        """Es el malentendido que la pantalla existe para evitar."""
        respuesta = self.client.get(self.url_borrador)

        self.assertContains(respuesta, "ya escribiste está guardado")

    def test_muestra_lo_que_falta(self):
        respuesta = self.client.get(self.url_borrador)

        self.assertContains(respuesta, "Lo que falta en esta encuesta")

    def test_guarda_la_nota(self):
        self.client.post(self.url_borrador, self.datos())

        self.encuesta.refresh_from_db()
        self.assertIn("ingresos", self.encuesta.nota_avance)

    def test_guarda_la_proxima_visita(self):
        manana = timezone.localdate() + timedelta(days=1)

        self.client.post(self.url_borrador, self.datos(proxima_visita=manana.isoformat()))

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.proxima_visita, manana)

    def test_el_mensaje_dice_cuando_hay_que_volver(self):
        manana = timezone.localdate() + timedelta(days=1)

        respuesta = self.client.post(
            self.url_borrador, self.datos(proxima_visita=manana.isoformat()), follow=True
        )
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("volver el" in m for m in mensajes))

    def test_una_pendiente_pasa_a_borrador(self):
        """Dejar una nota implica haber estado ahí: «pendiente» ya no es verdad."""
        pendiente = self.crear(direccion="Calle 2")
        url = reverse("fichas:guardar_borrador", kwargs={"pk": pendiente.pk})

        self.client.post(url, self.datos())

        pendiente.refresh_from_db()
        self.assertEqual(pendiente.estado, EstadoEncuesta.BORRADOR)

    def test_una_observada_sigue_observada(self):
        """Bajarla a borrador borraría el aviso más urgente del encuestador."""
        self.devolver(self.encuesta)

        self.client.post(self.url_borrador, self.datos())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.OBSERVADA)

    def test_una_fecha_pasada_no_se_guarda(self):
        ayer = (timezone.localdate() - timedelta(days=1)).isoformat()

        self.client.post(self.url_borrador, self.datos(proxima_visita=ayer))

        self.encuesta.refresh_from_db()
        self.assertIsNone(self.encuesta.proxima_visita)

    def test_la_encuesta_de_otra_persona_responde_404(self):
        ajena = self.crear(direccion="De Juan", censista=self.juan)

        respuesta = self.client.post(
            reverse("fichas:guardar_borrador", kwargs={"pk": ajena.pk}), self.datos()
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_una_encuesta_cerrada_no_admite_nota(self):
        self.completar_todo()
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        self.client.post(self.url_borrador, self.datos())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.nota_avance, "")

    def test_muestra_aparte_las_indicaciones_recibidas(self):
        """Dos campos distintos desde la HU-10, para que no se pisen."""
        self.encuesta.observaciones = "Pasar después de las 19:00."
        self.encuesta.save()

        respuesta = self.client.get(self.url_borrador)

        self.assertContains(respuesta, "Indicaciones que tienes")
        self.assertContains(respuesta, "después de las 19:00")


# ==========================================================================
# HU-10 — 41. TERMINAR LA ENCUESTA
# ==========================================================================


class CompletarEncuestaTest(BaseBorradorTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def test_con_pasos_pendientes_muestra_la_lista(self):
        respuesta = self.client.get(self.url_completar)

        self.assertContains(respuesta, "Todavía no se puede terminar")
        self.assertContains(respuesta, "Falta esto por registrar")

    def test_con_pasos_pendientes_no_ofrece_el_boton(self):
        respuesta = self.client.get(self.url_completar)

        self.assertNotContains(respuesta, "terminar y enviar a revisión")

    def test_con_pasos_pendientes_el_post_no_la_completa(self):
        """Ocultar el botón no es una validación: la URL se escribe a mano."""
        self.client.post(self.url_completar)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_el_post_rechazado_explica_que_falta(self):
        respuesta = self.client.post(self.url_completar, follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("falta" in m.lower() for m in mensajes))

    def test_completa_cuando_no_falta_nada(self):
        self.completar_todo()

        self.client.post(self.url_completar)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_al_completar_se_marca_la_fecha_de_cierre(self):
        self.completar_todo()

        self.client.post(self.url_completar)

        self.encuesta.refresh_from_db()
        self.assertIsNotNone(self.encuesta.cerrada_en)

    def test_despues_de_completar_vuelve_al_listado(self):
        self.completar_todo()

        respuesta = self.client.post(self.url_completar)

        self.assertRedirects(respuesta, reverse("fichas:mis_encuestas"))

    def test_el_mensaje_dice_que_va_a_revision(self):
        self.completar_todo()

        respuesta = self.client.post(self.url_completar, follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("revisión" in m for m in mensajes))

    def test_muestra_el_resumen_antes_de_enviar(self):
        """Última oportunidad de ver un dato mal escrito."""
        self.completar_todo()

        respuesta = self.client.get(self.url_completar)

        self.assertContains(respuesta, "Rosa Elena Millán")
        self.assertContains(respuesta, "La vivienda")
        self.assertContains(respuesta, "Las personas")

    def test_avisa_de_que_despues_no_se_puede_modificar(self):
        self.completar_todo()

        respuesta = self.client.get(self.url_completar)

        self.assertContains(respuesta, "ya no podrás modificarla")

    def test_una_completada_no_la_puede_reabrir_el_encuestador(self):
        """El camino de vuelta es del supervisor: devolverla como observada."""
        self.completar_todo()
        self.encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)

        respuesta = self.client.get(self.url_completar)

        self.assertRedirects(respuesta, self.url_detalle(self.encuesta))

    def test_una_observada_si_se_puede_volver_a_completar(self):
        """Es justamente para lo que existe ese estado."""
        self.completar_todo()
        self.devolver(self.encuesta)

        self.client.post(self.url_completar)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_la_encuesta_de_otra_persona_responde_404(self):
        ajena = self.crear(direccion="De Juan", censista=self.juan)

        respuesta = self.client.post(
            reverse("fichas:completar_encuesta", kwargs={"pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_el_supervisor_no_puede_completar_encuestas(self):
        self.completar_todo()
        self.client.force_login(self.supervisor)

        respuesta = self.client.post(self.url_completar)

        self.assertEqual(respuesta.status_code, 302)
        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_con_el_operativo_cerrado_no_se_puede_completar(self):
        self.completar_todo()
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.client.post(self.url_completar)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.BORRADOR)


# ==========================================================================
# HU-10 — 42. CERRAR SIN PODER LEVANTAR
# ==========================================================================


class CerrarSinDatosTest(BaseBorradorTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "estado": EstadoEncuesta.NO_UBICADA,
            "motivo_cierre": "La dirección no existe, el pasaje llega hasta el 40.",
        }
        base.update(extra)
        return base

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url_cerrar)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Cerrar sin levantar")

    def test_explica_que_no_es_un_fracaso(self):
        respuesta = self.client.get(self.url_cerrar)

        self.assertContains(respuesta, "no es un fracaso, es un resultado")

    def test_cierra_como_no_ubicada(self):
        self.client.post(self.url_cerrar, self.datos())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.NO_UBICADA)

    def test_cierra_como_rechazada(self):
        self.client.post(
            self.url_cerrar, self.datos(estado=EstadoEncuesta.RECHAZADA)
        )

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.RECHAZADA)

    def test_guarda_el_motivo(self):
        self.client.post(self.url_cerrar, self.datos())

        self.encuesta.refresh_from_db()
        self.assertIn("no existe", self.encuesta.motivo_cierre)

    def test_marca_las_dos_fechas(self):
        self.client.post(self.url_cerrar, self.datos())

        self.encuesta.refresh_from_db()
        self.assertIsNotNone(self.encuesta.iniciada_en)
        self.assertIsNotNone(self.encuesta.cerrada_en)

    def test_borra_la_proxima_visita(self):
        """Ya no espera a nadie: seguir avisando sería ruido."""
        self.encuesta.proxima_visita = timezone.localdate() + timedelta(days=2)
        self.encuesta.save()

        self.client.post(self.url_cerrar, self.datos())

        self.encuesta.refresh_from_db()
        self.assertIsNone(self.encuesta.proxima_visita)

    def test_sin_motivo_no_cierra(self):
        self.client.post(self.url_cerrar, self.datos(motivo_cierre=""))

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_con_un_motivo_de_una_letra_no_cierra(self):
        self.client.post(self.url_cerrar, self.datos(motivo_cierre="x"))

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.BORRADOR)

    def test_se_puede_cerrar_una_pendiente_sin_nada_registrado(self):
        """Es el caso más frecuente: se llega, no hay nadie, se cierra."""
        pendiente = self.crear(direccion="Calle 2")
        url = reverse("fichas:cerrar_encuesta", kwargs={"pk": pendiente.pk})

        self.client.post(url, self.datos())

        pendiente.refresh_from_db()
        self.assertEqual(pendiente.estado, EstadoEncuesta.NO_UBICADA)

    def test_avisa_si_ya_habia_datos_del_hogar(self):
        self.con_hogar()

        respuesta = self.client.get(self.url_cerrar)

        self.assertContains(respuesta, "ya tiene datos registrados")

    def test_cerrar_con_datos_conserva_el_hogar(self):
        """Si la familia se arrepintió a mitad, lo levantado no se tira."""
        hogar = self.con_hogar()

        self.client.post(self.url_cerrar, self.datos(estado=EstadoEncuesta.RECHAZADA))

        self.assertTrue(GrupoFamiliar.objects.filter(pk=hogar.pk).exists())

    def test_vuelve_al_listado(self):
        respuesta = self.client.post(self.url_cerrar, self.datos())

        self.assertRedirects(respuesta, reverse("fichas:mis_encuestas"))

    def test_la_encuesta_de_otra_persona_responde_404(self):
        ajena = self.crear(direccion="De Juan", censista=self.juan)

        respuesta = self.client.post(
            reverse("fichas:cerrar_encuesta", kwargs={"pk": ajena.pk}), self.datos()
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_una_ya_cerrada_no_se_puede_volver_a_cerrar(self):
        self.encuesta.motivo_cierre = "Motivo anterior que no se debe pisar."
        self.encuesta.save()
        self.encuesta.cambiar_estado(EstadoEncuesta.RECHAZADA)

        self.client.post(self.url_cerrar, self.datos())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.RECHAZADA)
        self.assertIn("anterior", self.encuesta.motivo_cierre)


# ==========================================================================
# HU-10 — 43. LO QUE GANAN LAS PANTALLAS ANTERIORES
# ==========================================================================


class BorradorEnLasPantallasTest(BaseBorradorTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def test_la_ficha_muestra_la_nota_de_avance(self):
        self.encuesta.nota_avance = "Falta el módulo de ingresos."
        self.encuesta.save()

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Por dónde ibas")
        self.assertContains(respuesta, "módulo de ingresos")

    def test_la_ficha_avisa_de_una_visita_vencida(self):
        self.encuesta.proxima_visita = timezone.localdate() - timedelta(days=2)
        self.encuesta.save()

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "esa fecha ya pasó")

    def test_la_ficha_muestra_el_motivo_del_cierre(self):
        self.encuesta.motivo_cierre = "La dirección no existe en ese pasaje."
        self.encuesta.save()
        self.encuesta.cambiar_estado(EstadoEncuesta.NO_UBICADA)

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "no existe en ese pasaje")

    def test_la_ficha_ofrece_terminar_cuando_esta_completa(self):
        self.completar_todo()

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Terminar y enviar a revisión")

    def test_la_ficha_ofrece_ver_que_falta_cuando_no_lo_esta(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Ver qué falta para terminar")

    def test_la_ficha_ofrece_las_tres_salidas(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, self.url_borrador)
        self.assertContains(respuesta, self.url_completar)
        self.assertContains(respuesta, self.url_cerrar)

    def test_una_encuesta_cerrada_no_ofrece_ninguna(self):
        self.completar_todo()
        self.encuesta.cambiar_estado(EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertNotContains(respuesta, self.url_cerrar)

    def test_el_listado_cuenta_las_visitas_vencidas(self):
        self.encuesta.proxima_visita = timezone.localdate() - timedelta(days=1)
        self.encuesta.save()

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.context["resumen"]["visitas_vencidas"], 1)
        self.assertContains(respuesta, "visita anotada")

    def test_una_visita_futura_no_cuenta_como_vencida(self):
        self.encuesta.proxima_visita = timezone.localdate() + timedelta(days=5)
        self.encuesta.save()

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.context["resumen"]["visitas_vencidas"], 0)

    def test_el_listado_muestra_la_fecha_de_la_visita(self):
        self.encuesta.proxima_visita = timezone.localdate() + timedelta(days=5)
        self.encuesta.save()

        respuesta = self.client.get(self.url_lista)

        self.assertContains(respuesta, "volver el")

    def test_las_visitas_de_encuestas_cerradas_no_se_cuentan(self):
        self.completar_todo()
        self.encuesta.proxima_visita = timezone.localdate() - timedelta(days=1)
        self.encuesta.save()
        self.encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)

        respuesta = self.client.get(self.url_lista)

        self.assertEqual(respuesta.context["resumen"]["visitas_vencidas"], 0)


# ==========================================================================
# HU-10 — 44. RECORRIDO COMPLETO DEL CICLO DE VIDA
# ==========================================================================


class IntegracionHU10Test(BaseEncuestaTest):
    """De la vivienda registrada al envío a revisión, pasando por el borrador."""

    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.client.force_login(self.marta)

    def nacido_hace(self, anios):
        return timezone.localdate() - timedelta(days=anios * 365 + 100)

    def test_recorrido_completo(self):
        # 1. Se registra la vivienda: la encuesta nace en BORRADOR.
        self.client.post(
            reverse("fichas:vivienda_registrar"),
            {
                "zona": self.zona1.pk,
                "direccion": "Av. Central 100",
                "referencia": "",
                "tipo": TipoVivienda.CASA,
                "tenencia": TenenciaVivienda.ARRENDADA,
                "materialidad_muros": MaterialidadMuros.ALBANILERIA,
                "origen_agua": OrigenAgua.RED_PUBLICA,
                "sistema_sanitario": SistemaSanitario.ALCANTARILLADO,
                "tiene_electricidad": True,
                "observaciones": "",
            },
        )
        encuesta = Encuesta.objects.get()
        self.assertEqual(encuesta.estado, EstadoEncuesta.BORRADOR)

        # 2. Todavía no se puede terminar: falta el hogar.
        respuesta = self.client.get(
            reverse("fichas:completar_encuesta", kwargs={"pk": encuesta.pk})
        )
        self.assertContains(respuesta, "Todavía no se puede terminar")

        # 3. Se registra el hogar de dos personas.
        self.client.post(
            reverse("fichas:registrar_hogar", kwargs={"pk": encuesta.pk}),
            {
                "jefe_hogar_nombre": "Rosa Elena Millán",
                "jefe_hogar_rut": "",
                "telefono_contacto": "",
                "integrantes_declarados": 2,
                "ingreso_mensual": "",
                "observaciones": "",
            },
        )

        # 4. Se hace de noche: se guarda el borrador con una nota y fecha de vuelta.
        manana = timezone.localdate() + timedelta(days=1)
        self.client.post(
            reverse("fichas:guardar_borrador", kwargs={"pk": encuesta.pk}),
            {
                "nota_avance": "Falta registrar a las dos personas del hogar.",
                "proxima_visita": manana.isoformat(),
            },
        )
        encuesta.refresh_from_db()
        self.assertEqual(encuesta.proxima_visita, manana)

        # 5. Al día siguiente, la nota está a la vista en la ficha.
        respuesta = self.client.get(
            reverse("fichas:encuesta_detalle", kwargs={"pk": encuesta.pk})
        )
        self.assertContains(respuesta, "Falta registrar a las dos personas")

        # 6. Se registran las dos personas.
        for numero, (parentesco, nombres) in enumerate(
            [(Parentesco.JEFE_HOGAR, "Rosa Elena"), (Parentesco.HIJO, "Camila")]
        ):
            self.client.post(
                reverse("fichas:integrante_nuevo", kwargs={"encuesta_pk": encuesta.pk}),
                {
                    "parentesco": parentesco,
                    "nombres": nombres,
                    "apellidos": "Millán",
                    "rut": "",
                    "sexo": Sexo.FEMENINO,
                    "fecha_nacimiento": self.nacido_hace(40 - numero * 20).isoformat(),
                    "nivel_educacional": NivelEducacional.MEDIA_COMPLETA,
                    "situacion_ocupacional": SituacionOcupacional.TRABAJA,
                    "pueblo_originario": PuebloOriginario.NINGUNO,
                    "observaciones": "",
                },
            )

        # 7. Ahora sí se puede terminar, y la pantalla muestra el resumen.
        encuesta.refresh_from_db()
        self.assertTrue(encuesta.puede_completarse)
        respuesta = self.client.get(
            reverse("fichas:completar_encuesta", kwargs={"pk": encuesta.pk})
        )
        self.assertContains(respuesta, "Rosa Elena Millán")

        # 8. Se envía a revisión.
        self.client.post(
            reverse("fichas:completar_encuesta", kwargs={"pk": encuesta.pk})
        )
        encuesta.refresh_from_db()
        self.assertEqual(encuesta.estado, EstadoEncuesta.COMPLETADA)
        self.assertIsNotNone(encuesta.cerrada_en)

        # 9. Y ya no se puede modificar: el camino de vuelta es del supervisor.
        respuesta = self.client.get(
            reverse("fichas:guardar_borrador", kwargs={"pk": encuesta.pk})
        )
        self.assertEqual(respuesta.status_code, 302)

        # 10. El listado lo refleja: una completada y ninguna por trabajar.
        respuesta = self.client.get(self.url_lista)
        self.assertEqual(respuesta.context["resumen"]["completadas"], 1)
        self.assertEqual(respuesta.context["resumen"]["por_trabajar"], 0)

    def test_recorrido_de_una_puerta_que_no_se_pudo(self):
        """El otro final posible, y el que la HU-07 dejó sin motivo dónde escribir."""
        encuesta = self.crear(direccion="Calle El Canelo 302")

        self.client.post(
            reverse("fichas:cerrar_encuesta", kwargs={"pk": encuesta.pk}),
            {
                "estado": EstadoEncuesta.NO_UBICADA,
                "motivo_cierre": (
                    "La vivienda está deshabitada desde hace meses según dos vecinos."
                ),
            },
        )

        encuesta.refresh_from_db()
        self.assertEqual(encuesta.estado, EstadoEncuesta.NO_UBICADA)
        self.assertIn("deshabitada", encuesta.motivo_cierre)
        self.assertFalse(encuesta.requiere_trabajo)

        # Y el motivo se lee en la ficha, que es lo que necesita el supervisor.
        respuesta = self.client.get(self.url_detalle(encuesta))
        self.assertContains(respuesta, "deshabitada")


# ==========================================================================
# HU-11 — 45. LA UBICACIÓN EN EL MODELO
# ==========================================================================


class BaseUbicacionTest(BaseEncuestaTest):
    """Escenario común: Marta con su sector y una vivienda a la que ubicar."""

    #: Un punto real dentro de Concepción, para no inventar coordenadas.
    LAT = Decimal("-36.826700")
    LON = Decimal("-73.049700")

    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.vivienda = self.crear_vivienda(direccion="Av. Central 100")
        self.encuesta = self.crear(
            vivienda=self.vivienda, estado=EstadoEncuesta.BORRADOR
        )
        self.url_ubicacion = reverse(
            "fichas:capturar_ubicacion", kwargs={"pk": self.vivienda.pk}
        )

    def ubicar(self, vivienda, lat=None, lon=None, precision=8, manual=False):
        vivienda.latitud = self.LAT if lat is None else Decimal(str(lat))
        vivienda.longitud = self.LON if lon is None else Decimal(str(lon))
        vivienda.precision_metros = precision
        vivienda.ubicacion_capturada_en = timezone.now()
        vivienda.ubicacion_manual = manual
        vivienda.save()
        return vivienda


class UbicacionModeloTest(BaseUbicacionTest):
    def test_una_vivienda_nace_sin_ubicacion(self):
        self.assertFalse(self.vivienda.tiene_ubicacion)
        self.assertIsNone(self.vivienda.coordenadas)

    def test_con_las_dos_coordenadas_tiene_ubicacion(self):
        self.ubicar(self.vivienda)

        self.assertTrue(self.vivienda.tiene_ubicacion)

    def test_las_coordenadas_se_muestran_con_seis_decimales(self):
        """Media pantalla con cuatro y la otra con seis parece dos datos distintos."""
        self.ubicar(self.vivienda)

        self.assertEqual(self.vivienda.coordenadas, "-36.826700, -73.049700")

    def test_una_precision_pequena_es_aceptable(self):
        self.ubicar(self.vivienda, precision=8)

        self.assertTrue(self.vivienda.precision_aceptable)

    def test_una_precision_grande_no_lo_es(self):
        self.ubicar(self.vivienda, precision=150)

        self.assertFalse(self.vivienda.precision_aceptable)

    def test_el_limite_de_precision_es_inclusivo(self):
        self.ubicar(self.vivienda, precision=Vivienda.PRECISION_ACEPTABLE)

        self.assertTrue(self.vivienda.precision_aceptable)

    def test_sin_dato_de_precision_no_se_da_por_buena(self):
        """De un punto del que no se sabe el error no se puede decir que sirve."""
        self.vivienda.latitud = self.LAT
        self.vivienda.longitud = self.LON
        self.vivienda.save()

        self.assertFalse(self.vivienda.precision_aceptable)

    def test_marca_las_ubicaciones_escritas_a_mano(self):
        self.ubicar(self.vivienda, manual=True)

        self.assertTrue(self.vivienda.ubicacion_manual)


class DistanciaTest(BaseUbicacionTest):
    def test_la_distancia_a_si_misma_es_cero(self):
        self.ubicar(self.vivienda)

        self.assertAlmostEqual(
            self.vivienda.distancia_a(self.LAT, self.LON), 0, places=3
        )

    def test_un_grado_de_latitud_son_unos_111_kilometros(self):
        """Comprobación contra un valor conocido: el haversine no puede ir a ojo."""
        self.ubicar(self.vivienda)

        metros = self.vivienda.distancia_a(self.LAT + Decimal("1"), self.LON)

        self.assertAlmostEqual(metros / 1000, 111.2, delta=0.5)

    def test_una_diezmilesima_de_grado_son_unos_once_metros(self):
        self.ubicar(self.vivienda)

        metros = self.vivienda.distancia_a(self.LAT + Decimal("0.0001"), self.LON)

        self.assertAlmostEqual(metros, 11.1, delta=1)

    def test_sin_ubicacion_no_hay_distancia(self):
        self.assertIsNone(self.vivienda.distancia_a(self.LAT, self.LON))

    def test_el_centro_de_una_zona_sin_ubicadas_es_none(self):
        """La primera casa de la jornada no tiene contra qué compararse."""
        self.assertIsNone(Vivienda.centro_de_la_zona(self.zona1))

    def test_el_centro_promedia_las_ubicadas(self):
        self.ubicar(self.vivienda, lat="-36.800000", lon="-73.000000")
        self.ubicar(
            self.crear_vivienda(direccion="Otra 1"),
            lat="-36.900000",
            lon="-73.100000",
        )

        lat, lon = Vivienda.centro_de_la_zona(self.zona1)

        self.assertAlmostEqual(float(lat), -36.85, places=4)
        self.assertAlmostEqual(float(lon), -73.05, places=4)

    def test_el_centro_puede_excluir_una_vivienda(self):
        """Al recolocar una casa, ella misma no puede ser su propia referencia."""
        self.ubicar(self.vivienda, lat="-36.800000", lon="-73.000000")
        otra = self.ubicar(
            self.crear_vivienda(direccion="Otra 1"),
            lat="-36.900000",
            lon="-73.100000",
        )

        lat, lon = Vivienda.centro_de_la_zona(self.zona1, excluir=otra.pk)

        self.assertAlmostEqual(float(lat), -36.80, places=4)

    def test_el_centro_no_mezcla_zonas(self):
        self.ubicar(self.vivienda)
        self.ubicar(self.crear_vivienda(direccion="En otra zona", zona=self.zona2))

        puntos = Vivienda.objects.filter(zona=self.zona1, latitud__isnull=False)

        self.assertEqual(puntos.count(), 1)


# ==========================================================================
# HU-11 — 46. LO QUE GARANTIZA LA BASE DE DATOS
# ==========================================================================


class RestriccionesUbicacionTest(BaseUbicacionTest):
    def crear_con(self, **extra):
        return Vivienda.objects.create(
            zona=self.zona1, direccion="Prueba 1", **extra
        )

    def test_no_se_puede_guardar_media_coordenada(self):
        """Una latitud sin longitud es una línea que cruza el planeta."""
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_con(latitud=self.LAT)

    def test_tampoco_al_reves(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_con(longitud=self.LON)

    def test_las_dos_vacias_si_se_admiten(self):
        vivienda = self.crear_con()

        self.assertFalse(vivienda.tiene_ubicacion)

    def test_una_latitud_positiva_se_rechaza(self):
        """El error más común al escribir a mano: olvidar el signo."""
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_con(latitud=Decimal("36.826700"), longitud=self.LON)

    def test_una_longitud_fuera_de_chile_se_rechaza(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_con(latitud=self.LAT, longitud=Decimal("-3.700000"))

    def test_intercambiar_latitud_y_longitud_se_rechaza(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_con(latitud=self.LON, longitud=self.LAT)

    def test_rapa_nui_se_acepta(self):
        """Si el rango fuera solo el continental, quedaría fuera territorio chileno."""
        vivienda = self.crear_con(
            latitud=Decimal("-27.150000"), longitud=Decimal("-109.433300")
        )

        self.assertTrue(vivienda.tiene_ubicacion)

    def test_el_extremo_norte_se_acepta(self):
        vivienda = self.crear_con(
            latitud=Decimal("-17.500000"), longitud=Decimal("-70.100000")
        )

        self.assertTrue(vivienda.tiene_ubicacion)

    def test_el_extremo_sur_se_acepta(self):
        vivienda = self.crear_con(
            latitud=Decimal("-56.500000"), longitud=Decimal("-68.700000")
        )

        self.assertTrue(vivienda.tiene_ubicacion)


# ==========================================================================
# HU-11 — 47. EL FORMULARIO
# ==========================================================================


class UbicacionFormTest(BaseUbicacionTest):
    def datos(self, **extra):
        base = {
            "latitud": "-36.826700",
            "longitud": "-73.049700",
            "precision_metros": 8,
        }
        base.update(extra)
        return base

    def formulario(self, datos=None, vivienda=None):
        return UbicacionForm(datos, instance=vivienda or self.vivienda)

    def test_un_punto_valido_se_acepta(self):
        formulario = self.formulario(self.datos())

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_las_coordenadas_son_obligatorias_en_esta_pantalla(self):
        """Una pantalla que se llama «capturar ubicación» sin ubicación no hizo nada."""
        for campo in ("latitud", "longitud"):
            with self.subTest(campo=campo):
                formulario = self.formulario(self.datos(**{campo: ""}))
                self.assertFalse(formulario.is_valid())
                self.assertIn(campo, formulario.errors)

    def test_la_precision_es_opcional(self):
        formulario = self.formulario(self.datos(precision_metros=""))

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_una_latitud_positiva_se_rechaza_con_mensaje(self):
        formulario = self.formulario(self.datos(latitud="36.826700"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("latitud", formulario.errors)
        self.assertIn("signo", str(formulario.errors["latitud"]))

    def test_una_longitud_fuera_de_chile_se_rechaza_con_mensaje(self):
        formulario = self.formulario(self.datos(longitud="-3.700000"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("longitud", formulario.errors)

    def test_marca_como_manual_si_no_vino_del_aparato(self):
        """La suposición prudente: sin la marca del script, se escribió a mano."""
        formulario = self.formulario(self.datos())
        formulario.is_valid()
        vivienda = formulario.save()

        self.assertTrue(vivienda.ubicacion_manual)

    def test_no_la_marca_como_manual_si_la_capturo_el_aparato(self):
        formulario = self.formulario(self.datos(capturada="1"))
        formulario.is_valid()
        vivienda = formulario.save()

        self.assertFalse(vivienda.ubicacion_manual)

    def test_registra_cuando_se_capturo(self):
        formulario = self.formulario(self.datos())
        formulario.is_valid()
        vivienda = formulario.save()

        self.assertIsNotNone(vivienda.ubicacion_capturada_en)

    # -- la tercera capa: la lejanía --------------------------------------

    def test_sin_otras_viviendas_ubicadas_no_hay_nada_que_comparar(self):
        formulario = self.formulario(self.datos())

        self.assertTrue(formulario.is_valid(), formulario.errors)
        self.assertIsNone(formulario.distancia_al_resto)

    def test_un_punto_cercano_al_resto_pasa_sin_preguntar(self):
        self.ubicar(self.crear_vivienda(direccion="Vecina 1"))

        formulario = self.formulario(self.datos(latitud="-36.826800"))

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_un_punto_lejano_pide_confirmacion(self):
        """El caso real: el teléfono devolvió la posición de hace media hora."""
        self.ubicar(self.crear_vivienda(direccion="Vecina 1"))

        formulario = self.formulario(self.datos(latitud="-36.900000"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("confirmar_lejania", formulario.errors)

    def test_con_la_casilla_marcada_el_punto_lejano_se_acepta(self):
        """Una parcela apartada puede pertenecer de verdad a la zona."""
        self.ubicar(self.crear_vivienda(direccion="Vecina 1"))

        formulario = self.formulario(
            self.datos(latitud="-36.900000", confirmar_lejania=True)
        )

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_el_aviso_dice_a_cuantos_metros_esta(self):
        self.ubicar(self.crear_vivienda(direccion="Vecina 1"))

        formulario = self.formulario(self.datos(latitud="-36.900000"))
        formulario.is_valid()

        self.assertIn("m del resto", str(formulario.errors["confirmar_lejania"]))

    def test_recolocar_una_vivienda_no_la_compara_consigo_misma(self):
        """Si se comparara, mover una casa 600 m siempre pediría confirmación."""
        self.ubicar(self.vivienda)

        formulario = self.formulario(self.datos(latitud="-36.900000"))

        self.assertTrue(formulario.is_valid(), formulario.errors)


# ==========================================================================
# HU-11 — 48. LA PANTALLA DE CAPTURA
# ==========================================================================


class CapturarUbicacionTest(BaseUbicacionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "latitud": "-36.826700",
            "longitud": "-73.049700",
            "precision_metros": 8,
        }
        base.update(extra)
        return base

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url_ubicacion)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Capturar mi ubicación")

    def test_incluye_el_script_de_geolocalizacion(self):
        """Es la única pantalla del proyecto con JavaScript propio."""
        respuesta = self.client.get(self.url_ubicacion)

        self.assertContains(respuesta, "navigator.geolocation")

    def test_los_campos_son_editables_sin_javascript(self):
        """Progressive enhancement: el script mejora la pantalla, no la sostiene."""
        respuesta = self.client.get(self.url_ubicacion)

        self.assertContains(respuesta, 'name="latitud"')
        self.assertContains(respuesta, 'name="longitud"')

    def test_no_pide_nada_a_un_servidor_externo(self):
        """Ni mapas ni teselas: las coordenadas de una familia no salen de OPSO."""
        respuesta = self.client.get(self.url_ubicacion)
        html = respuesta.content.decode()

        self.assertNotIn("openstreetmap", html.lower())
        self.assertNotIn("googleapis", html.lower())
        self.assertNotIn("mapbox", html.lower())

    def test_guarda_la_ubicacion(self):
        self.client.post(self.url_ubicacion, self.datos())

        self.vivienda.refresh_from_db()
        self.assertTrue(self.vivienda.tiene_ubicacion)

    def test_vuelve_a_la_ficha_de_la_vivienda(self):
        respuesta = self.client.post(self.url_ubicacion, self.datos())

        self.assertRedirects(
            respuesta,
            reverse("fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk}),
        )

    def test_el_mensaje_confirma_la_precision(self):
        respuesta = self.client.post(self.url_ubicacion, self.datos(), follow=True)
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("Precisión: 8 m" in m for m in mensajes))

    def test_el_mensaje_avisa_si_la_precision_es_mala(self):
        """«Guardado» a secas escondería que el punto no sirve."""
        respuesta = self.client.post(
            self.url_ubicacion, self.datos(precision_metros=300), follow=True
        )
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("mucho" in m for m in mensajes))

    def test_el_mensaje_avisa_si_no_hay_dato_de_precision(self):
        respuesta = self.client.post(
            self.url_ubicacion, self.datos(precision_metros=""), follow=True
        )
        mensajes = [str(m) for m in respuesta.context["messages"]]

        self.assertTrue(any("precisión" in m for m in mensajes))

    def test_un_punto_invalido_no_guarda_nada(self):
        self.client.post(self.url_ubicacion, self.datos(latitud="36.826700"))

        self.vivienda.refresh_from_db()
        self.assertFalse(self.vivienda.tiene_ubicacion)

    def test_volver_a_capturar_reemplaza_el_punto(self):
        self.ubicar(self.vivienda)

        self.client.post(self.url_ubicacion, self.datos(latitud="-36.826800"))

        self.vivienda.refresh_from_db()
        self.assertEqual(str(self.vivienda.latitud), "-36.826800")

    # -- acceso ------------------------------------------------------------

    def test_una_vivienda_fuera_de_mi_territorio_responde_404(self):
        ajena = Vivienda.objects.create(zona=self.zona_norte, direccion="Otra 1")

        respuesta = self.client.get(
            reverse("fichas:capturar_ubicacion", kwargs={"pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_un_companero_del_mismo_sector_si_puede(self):
        """La ubicación describe el inmueble, no el trabajo de nadie."""
        AsignacionSector.objects.create(sector=self.boldos, censista=self.juan)
        self.client.force_login(self.juan)

        self.assertEqual(self.client.get(self.url_ubicacion).status_code, 200)

    def test_el_supervisor_no_puede_capturar(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_ubicacion)

        self.assertEqual(respuesta.status_code, 302)

    def test_con_el_operativo_cerrado_no_se_puede(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.client.post(self.url_ubicacion, self.datos())

        self.vivienda.refresh_from_db()
        self.assertFalse(self.vivienda.tiene_ubicacion)


# ==========================================================================
# HU-11 — 49. LA UBICACIÓN EN LAS DEMÁS PANTALLAS
# ==========================================================================


class UbicacionEnLasPantallasTest(BaseUbicacionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.url_vivienda = reverse(
            "fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk}
        )
        self.url_completar = reverse(
            "fichas:completar_encuesta", kwargs={"pk": self.encuesta.pk}
        )

    def completar_datos(self):
        """Deja la encuesta lista para terminarse, sin tocar la ubicación."""
        hogar = GrupoFamiliar.objects.create(
            encuesta=self.encuesta,
            jefe_hogar_nombre="Rosa Elena Millán",
            integrantes_declarados=1,
        )
        Integrante.objects.create(
            grupo_familiar=hogar,
            parentesco=Parentesco.JEFE_HOGAR,
            nombres="Rosa Elena",
            apellidos="Millán",
            sexo=Sexo.FEMENINO,
            fecha_nacimiento=timezone.localdate() - timedelta(days=40 * 365),
            nivel_educacional=NivelEducacional.MEDIA_COMPLETA,
            situacion_ocupacional=SituacionOcupacional.TRABAJA,
        )

    def test_la_ficha_de_la_vivienda_dice_que_no_hay_ubicacion(self):
        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(respuesta, "Sin ubicación capturada")

    def test_la_ficha_muestra_las_coordenadas(self):
        self.ubicar(self.vivienda)

        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(respuesta, "-36.826700, -73.049700")

    def test_la_ficha_marca_una_precision_mala(self):
        self.ubicar(self.vivienda, precision=300)

        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(respuesta, "poco precisa")

    def test_la_ficha_marca_las_escritas_a_mano(self):
        self.ubicar(self.vivienda, manual=True)

        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(respuesta, "escrita a mano")

    def test_la_ubicacion_no_es_un_paso_pendiente(self):
        """Depende de la señal, y eso no lo controla el encuestador."""
        self.completar_datos()

        self.assertTrue(self.encuesta.puede_completarse)

    def test_terminar_avisa_si_no_hay_ubicacion(self):
        self.completar_datos()

        respuesta = self.client.get(self.url_completar)

        self.assertContains(respuesta, "no tiene ubicación capturada")

    def test_terminar_avisa_si_la_precision_es_mala(self):
        self.completar_datos()
        self.ubicar(self.vivienda, precision=300)

        respuesta = self.client.get(self.url_completar)

        self.assertContains(respuesta, "poca precisión")

    def test_se_puede_terminar_sin_ubicacion(self):
        """El aviso no bloquea: una ficha correcta no se queda atrapada por el GPS."""
        self.completar_datos()

        self.client.post(self.url_completar)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)


# ==========================================================================
# HU-11 — 50. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionHU11Test(BaseUbicacionTest):
    def test_recorrido_completo(self):
        self.client.force_login(self.marta)

        # 1. La vivienda todavía no tiene punto.
        respuesta = self.client.get(
            reverse("fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk})
        )
        self.assertContains(respuesta, "Sin ubicación capturada")

        # 2. Se captura desde el aparato (el script marca `capturada=1`).
        self.client.post(
            self.url_ubicacion,
            {
                "latitud": "-36.826700",
                "longitud": "-73.049700",
                "precision_metros": 7,
                "capturada": "1",
            },
        )
        self.vivienda.refresh_from_db()
        self.assertTrue(self.vivienda.tiene_ubicacion)
        self.assertFalse(self.vivienda.ubicacion_manual)
        self.assertTrue(self.vivienda.precision_aceptable)

        # 3. La casa de al lado se captura cerca y no molesta a nadie.
        vecina = self.crear_vivienda(direccion="Av. Central 118")
        respuesta = self.client.post(
            reverse("fichas:capturar_ubicacion", kwargs={"pk": vecina.pk}),
            {
                "latitud": "-36.826810",
                "longitud": "-73.049700",
                "precision_metros": 9,
                "capturada": "1",
            },
        )
        vecina.refresh_from_db()
        self.assertTrue(vecina.tiene_ubicacion)

        # 4. Una tercera, con el teléfono devolviendo la posición de hace rato:
        #    el sistema lo detecta y pide confirmar.
        lejana = self.crear_vivienda(direccion="Av. Central 132")
        url_lejana = reverse("fichas:capturar_ubicacion", kwargs={"pk": lejana.pk})
        self.client.post(
            url_lejana,
            {
                "latitud": "-36.900000",
                "longitud": "-73.049700",
                "precision_metros": 9,
                "capturada": "1",
            },
        )
        lejana.refresh_from_db()
        self.assertFalse(lejana.tiene_ubicacion)

        # 5. Confirmando, se guarda: puede ser una parcela apartada de verdad.
        self.client.post(
            url_lejana,
            {
                "latitud": "-36.900000",
                "longitud": "-73.049700",
                "precision_metros": 9,
                "capturada": "1",
                "confirmar_lejania": "on",
            },
        )
        lejana.refresh_from_db()
        self.assertTrue(lejana.tiene_ubicacion)

        # 6. Y un punto imposible se rechaza siempre, con o sin confirmación.
        self.client.post(
            url_lejana,
            {
                "latitud": "36.900000",
                "longitud": "-73.049700",
                "precision_metros": 9,
                "capturada": "1",
                "confirmar_lejania": "on",
            },
        )
        lejana.refresh_from_db()
        self.assertEqual(str(lejana.latitud), "-36.900000")


# ==========================================================================
# HU-12 — 51. LAS FOTOGRAFÍAS
# ==========================================================================


def imagen_de_prueba(nombre="foto.jpg", formato="JPEG", medidas=(60, 40)):
    """Una imagen de verdad, en memoria, para las pruebas de subida.

    Tiene que ser una imagen REAL y no unos bytes cualesquiera, porque ImageField la
    decodifica con Pillow: ese es justamente el control que la HU-12 quiere probar.
    """
    buffer = BytesIO()
    Image.new("RGB", medidas, "red").save(buffer, format=formato)
    buffer.seek(0)

    return SimpleUploadedFile(nombre, buffer.read(), content_type="image/jpeg")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp(prefix="opso-pruebas-"))
class BaseFotografiaTest(BaseEncuestaTest):
    """Escenario común, con MEDIA_ROOT apuntando a una carpeta temporal.

    Sin `override_settings`, las pruebas escribirían en la carpeta media real del
    proyecto y la irían llenando de archivos que nadie borra. La carpeta temporal se
    elimina al terminar la clase.
    """

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(settings.MEDIA_ROOT, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.vivienda = self.crear_vivienda(direccion="Av. Central 100")
        self.encuesta = self.crear(
            vivienda=self.vivienda, estado=EstadoEncuesta.BORRADOR
        )
        self.url_subir = reverse(
            "fichas:subir_fotografia", kwargs={"pk": self.vivienda.pk}
        )

    def crear_foto(self, vivienda=None, tipo=TipoFotografia.FACHADA, **extra):
        datos = {
            "vivienda": vivienda or self.vivienda,
            "imagen": imagen_de_prueba(),
            "tipo": tipo,
            "descripcion": "El número de la casa está borrado.",
            "tomada_por": self.marta,
        }
        datos.update(extra)
        return Fotografia.objects.create(**datos)


class FotografiaModeloTest(BaseFotografiaTest):
    def test_se_asocia_a_su_vivienda(self):
        foto = self.crear_foto()

        self.assertIn(foto, self.vivienda.fotografias.all())

    def test_el_nombre_original_no_se_conserva(self):
        """Lo elige quien sube: no se usa para nada."""
        foto = self.crear_foto(imagen=imagen_de_prueba(nombre="IMG_0001.jpg"))

        self.assertNotIn("IMG_0001", foto.imagen.name)

    def test_el_archivo_se_guarda_con_un_nombre_impredecible(self):
        foto = self.crear_foto()
        nombre = Path(foto.imagen.name).stem

        # Un UUID en hexadecimal: 32 caracteres.
        self.assertEqual(len(nombre), 32)

    def test_se_reparte_por_ano_y_mes(self):
        foto = self.crear_foto()

        self.assertTrue(foto.imagen.name.startswith("fichas/"))
        self.assertIn(str(timezone.localdate().year), foto.imagen.name)

    def test_una_extension_rara_se_reemplaza(self):
        """La extensión del disco no es la fuente de verdad sobre el contenido."""
        foto = self.crear_foto(imagen=imagen_de_prueba(nombre="foto.raro"))

        self.assertTrue(foto.imagen.name.endswith(".jpg"))

    def test_una_extension_admitida_se_conserva(self):
        foto = self.crear_foto(imagen=imagen_de_prueba(nombre="foto.png"))

        self.assertTrue(foto.imagen.name.endswith(".png"))

    def test_el_texto_dice_el_tipo_y_la_direccion(self):
        foto = self.crear_foto()

        self.assertIn("Fachada", str(foto))
        self.assertIn("Av. Central 100", str(foto))

    def test_la_url_apunta_a_la_vista_y_no_al_archivo(self):
        """`imagen.url` daría la ruta directa, que OPSO no sirve a propósito."""
        foto = self.crear_foto()

        self.assertEqual(
            foto.get_absolute_url(),
            reverse("fichas:ver_fotografia", kwargs={"pk": foto.pk}),
        )

    def test_informa_el_tamano(self):
        self.assertGreater(self.crear_foto().tamano_kb, 0)

    def test_sin_archivo_el_tamano_es_none_y_no_revienta(self):
        """Una fila cuyo archivo se perdió tiene que poder mostrarse igual."""
        foto = self.crear_foto()
        Path(foto.imagen.path).unlink()

        self.assertIsNone(foto.tamano_kb)

    def test_borrar_la_vivienda_se_lleva_las_fotos(self):
        self.crear_foto()

        self.vivienda.delete()

        self.assertEqual(Fotografia.objects.count(), 0)

    def test_borrar_archivo_elimina_la_fila_y_el_archivo(self):
        """Django no borra el archivo solo: hay que hacerlo explícitamente."""
        foto = self.crear_foto()
        ruta = Path(foto.imagen.path)
        self.assertTrue(ruta.exists())

        foto.borrar_archivo()

        self.assertEqual(Fotografia.objects.count(), 0)
        self.assertFalse(ruta.exists())


class RestriccionesFotografiaTest(BaseFotografiaTest):
    def test_un_tipo_inventado_lo_rechaza_la_base_de_datos(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_foto(tipo="SELFIE")

    def test_una_descripcion_vacia_la_rechaza_la_base_de_datos(self):
        """Una foto sin explicación no es evidencia de nada."""
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.crear_foto(descripcion="")


# ==========================================================================
# HU-12 — 52. EL FORMULARIO
# ==========================================================================


class FotografiaFormTest(BaseFotografiaTest):
    def datos(self, **extra):
        base = {
            "tipo": TipoFotografia.FACHADA,
            "descripcion": "El número de la casa está borrado.",
        }
        base.update(extra)
        return base

    def formulario(self, datos=None, archivo=None, vivienda=None):
        archivos = {"imagen": archivo} if archivo is not None else None
        return FotografiaForm(
            datos, archivos, vivienda=vivienda or self.vivienda
        )

    def test_una_imagen_valida_se_acepta(self):
        formulario = self.formulario(self.datos(), imagen_de_prueba())

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_la_imagen_es_obligatoria(self):
        formulario = self.formulario(self.datos())

        self.assertFalse(formulario.is_valid())
        self.assertIn("imagen", formulario.errors)

    def test_un_archivo_que_no_es_imagen_se_rechaza(self):
        """Lo detecta Pillow al decodificar: mirar la extensión no valida nada."""
        falso = SimpleUploadedFile(
            "foto.jpg", b"esto no es una imagen", content_type="image/jpeg"
        )

        formulario = self.formulario(self.datos(), falso)

        self.assertFalse(formulario.is_valid())
        self.assertIn("imagen", formulario.errors)

    def test_un_formato_no_admitido_se_rechaza(self):
        """Pillow abre docenas de formatos; se aceptan solo los tres de un teléfono."""
        formulario = self.formulario(
            self.datos(), imagen_de_prueba(nombre="foto.gif", formato="GIF")
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("imagen", formulario.errors)

    def test_png_y_webp_se_aceptan(self):
        for formato, nombre in (("PNG", "f.png"), ("WEBP", "f.webp")):
            with self.subTest(formato=formato):
                formulario = self.formulario(
                    self.datos(), imagen_de_prueba(nombre=nombre, formato=formato)
                )
                self.assertTrue(formulario.is_valid(), formulario.errors)

    @override_settings(OPSO_TAMANO_MAXIMO_FOTO=500)
    def test_una_imagen_demasiado_pesada_se_rechaza(self):
        formulario = self.formulario(self.datos(), imagen_de_prueba(medidas=(800, 600)))

        self.assertFalse(formulario.is_valid())
        self.assertIn("imagen", formulario.errors)

    def test_la_descripcion_es_obligatoria(self):
        formulario = self.formulario(self.datos(descripcion=""), imagen_de_prueba())

        self.assertFalse(formulario.is_valid())
        self.assertIn("descripcion", formulario.errors)

    def test_una_descripcion_de_dos_palabras_no_basta(self):
        formulario = self.formulario(self.datos(descripcion="la casa"), imagen_de_prueba())

        self.assertFalse(formulario.is_valid())

    @override_settings(OPSO_MAXIMO_FOTOS_POR_VIVIENDA=2)
    def test_no_se_pasa_del_tope_de_fotografias(self):
        """«Cuando sea necesario»: la ficha no puede volverse un álbum."""
        self.crear_foto()
        self.crear_foto()

        formulario = self.formulario(self.datos(), imagen_de_prueba())

        self.assertFalse(formulario.is_valid())

    @override_settings(OPSO_MAXIMO_FOTOS_POR_VIVIENDA=2)
    def test_el_tope_es_por_vivienda(self):
        self.crear_foto()
        self.crear_foto()
        otra = self.crear_vivienda(direccion="Otra 1")

        formulario = self.formulario(self.datos(), imagen_de_prueba(), vivienda=otra)

        self.assertTrue(formulario.is_valid(), formulario.errors)


# ==========================================================================
# HU-12 — 53. SUBIR Y QUITAR
# ==========================================================================


class SubirFotografiaTest(BaseFotografiaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)

    def datos(self, **extra):
        base = {
            "tipo": TipoFotografia.FACHADA,
            "descripcion": "El número de la casa está borrado.",
            "imagen": imagen_de_prueba(),
        }
        base.update(extra)
        return base

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url_subir)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Adjuntar una fotografía")

    def test_advierte_de_no_fotografiar_personas(self):
        """Es la regla más importante de la historia, y va antes del campo."""
        respuesta = self.client.get(self.url_subir)

        self.assertContains(respuesta, "No fotografíes a personas")

    def test_el_formulario_admite_archivos(self):
        """Sin enctype, el navegador manda el nombre y no el archivo."""
        respuesta = self.client.get(self.url_subir)

        self.assertContains(respuesta, "multipart/form-data")

    def test_guarda_la_fotografia(self):
        self.client.post(self.url_subir, self.datos())

        self.assertEqual(self.vivienda.fotografias.count(), 1)

    def test_registra_quien_la_subio(self):
        self.client.post(self.url_subir, self.datos())

        self.assertEqual(Fotografia.objects.get().tomada_por, self.marta)

    def test_vuelve_a_la_ficha_de_la_vivienda(self):
        respuesta = self.client.post(self.url_subir, self.datos())

        self.assertRedirects(
            respuesta,
            reverse("fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk}),
        )

    def test_un_archivo_invalido_no_guarda_nada(self):
        falso = SimpleUploadedFile("x.jpg", b"no soy una imagen")

        self.client.post(self.url_subir, self.datos(imagen=falso))

        self.assertEqual(Fotografia.objects.count(), 0)

    def test_una_vivienda_fuera_de_mi_territorio_responde_404(self):
        ajena = Vivienda.objects.create(zona=self.zona_norte, direccion="Otra 1")

        respuesta = self.client.get(
            reverse("fichas:subir_fotografia", kwargs={"pk": ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)

    def test_el_supervisor_no_puede_subir(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_subir)

        self.assertEqual(respuesta.status_code, 302)

    def test_con_el_operativo_cerrado_no_se_puede(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        self.client.post(self.url_subir, self.datos())

        self.assertEqual(Fotografia.objects.count(), 0)


class QuitarFotografiaTest(BaseFotografiaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.foto = self.crear_foto()
        self.url_quitar = reverse(
            "fichas:quitar_fotografia", kwargs={"pk": self.foto.pk}
        )

    def test_el_get_solo_confirma(self):
        respuesta = self.client.get(self.url_quitar)

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(Fotografia.objects.count(), 1)

    def test_avisa_de_que_el_archivo_se_borra(self):
        respuesta = self.client.get(self.url_quitar)

        self.assertContains(respuesta, "se borra del disco")

    def test_el_post_borra_la_fila(self):
        self.client.post(self.url_quitar)

        self.assertEqual(Fotografia.objects.count(), 0)

    def test_el_post_borra_tambien_el_archivo(self):
        """Si no, el disco acumula fotos de familias que ya nadie puede ver."""
        ruta = Path(self.foto.imagen.path)

        self.client.post(self.url_quitar)

        self.assertFalse(ruta.exists())

    def test_una_foto_fuera_de_mi_territorio_responde_404(self):
        ajena = Vivienda.objects.create(zona=self.zona_norte, direccion="Otra 1")
        foto_ajena = self.crear_foto(vivienda=ajena)

        respuesta = self.client.post(
            reverse("fichas:quitar_fotografia", kwargs={"pk": foto_ajena.pk})
        )

        self.assertEqual(respuesta.status_code, 404)
        self.assertEqual(Fotografia.objects.count(), 2)


# ==========================================================================
# HU-12 — 54. SERVIR EL ARCHIVO: LA DECISIÓN DE SEGURIDAD
# ==========================================================================


class ServirFotografiaTest(BaseFotografiaTest):
    def setUp(self):
        super().setUp()
        self.foto = self.crear_foto()
        self.url_ver = reverse("fichas:ver_fotografia", kwargs={"pk": self.foto.pk})

    def test_el_encuestador_ve_su_foto(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_ver)

        self.assertEqual(respuesta.status_code, 200)

    def test_entrega_el_contenido_del_archivo(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_ver)
        contenido = b"".join(respuesta.streaming_content)
        respuesta.close()

        self.assertGreater(len(contenido), 0)

    def test_un_visitante_anonimo_no_puede(self):
        """El punto entero de la historia: el archivo NO es público."""
        respuesta = self.client.get(self.url_ver)

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(reverse("usuarios:login"), respuesta.url)

    def test_un_encuestador_ajeno_recibe_404(self):
        self.client.force_login(self.juan)

        respuesta = self.client.get(self.url_ver)

        self.assertEqual(respuesta.status_code, 404)

    def test_el_supervisor_si_puede_con_ver_todas(self):
        """Es lo que necesita para revisar el trabajo."""
        self.client.force_login(self.supervisor)

        self.assertEqual(self.client.get(self.url_ver).status_code, 200)

    def test_sin_ver_todas_el_supervisor_no_puede(self):
        self.rol_supervisor.permisos.remove(
            Permiso.objects.get(codigo="fichas.ver_todas")
        )
        self.client.force_login(self.supervisor)

        self.assertEqual(self.client.get(self.url_ver).status_code, 404)

    def test_un_companero_del_mismo_sector_si_puede(self):
        AsignacionSector.objects.create(sector=self.boldos, censista=self.juan)
        self.client.force_login(self.juan)

        self.assertEqual(self.client.get(self.url_ver).status_code, 200)

    def test_prohibe_adivinar_el_tipo_de_contenido(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_ver)

        self.assertEqual(respuesta["X-Content-Type-Options"], "nosniff")

    def test_prohibe_guardarla_en_cachés_compartidas(self):
        """Son datos personales: no pueden quedar en una caché intermedia."""
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_ver)

        self.assertIn("no-store", respuesta["Cache-Control"])

    def test_si_el_archivo_falta_responde_404(self):
        """Un respaldo restaurado a medias no puede tumbar la pantalla."""
        Path(self.foto.imagen.path).unlink()
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_ver).status_code, 404)

    def test_la_carpeta_media_no_se_sirve_como_estatica(self):
        """La comprobación que resume la historia: no hay ruta pública al archivo.

        Si algún día alguien agregara `static(MEDIA_URL, ...)` a las URLs, esta
        prueba fallaría, que es exactamente lo que tiene que pasar.
        """
        self.client.force_login(self.marta)

        respuesta = self.client.get(f"/{settings.MEDIA_URL}{self.foto.imagen.name}")

        self.assertNotEqual(respuesta.status_code, 200)


# ==========================================================================
# HU-12 — 55. LAS FOTOS EN LA FICHA DE LA VIVIENDA
# ==========================================================================


class FotografiasEnLaFichaTest(BaseFotografiaTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.marta)
        self.url_vivienda = reverse(
            "fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk}
        )

    def test_sin_fotos_explica_cuando_hacen_falta(self):
        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(respuesta, "Sin fotografías")
        self.assertContains(respuesta, "solo cuando hacen falta")

    def test_muestra_las_fotos_adjuntas(self):
        foto = self.crear_foto()

        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(
            respuesta, reverse("fichas:ver_fotografia", kwargs={"pk": foto.pk})
        )

    def test_nunca_enlaza_el_archivo_directo(self):
        self.crear_foto()

        respuesta = self.client.get(self.url_vivienda)

        self.assertNotContains(respuesta, "/media/fichas/")

    def test_ofrece_adjuntar(self):
        respuesta = self.client.get(self.url_vivienda)

        self.assertContains(respuesta, self.url_subir)

    def test_con_el_operativo_cerrado_no_ofrece_adjuntar(self):
        self.operativo.estado = EstadoOperativo.CERRADO
        self.operativo.save()

        respuesta = self.client.get(self.url_vivienda)

        self.assertNotContains(respuesta, self.url_subir)


# ==========================================================================
# HU-12 — 56. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionHU12Test(BaseFotografiaTest):
    def test_recorrido_completo(self):
        self.client.force_login(self.marta)

        # 1. La vivienda no tiene fotos y la ficha lo dice.
        url_vivienda = reverse(
            "fichas:vivienda_detalle", kwargs={"pk": self.vivienda.pk}
        )
        self.assertContains(self.client.get(url_vivienda), "Sin fotografías")

        # 2. Se adjunta una, con su motivo.
        self.client.post(
            self.url_subir,
            {
                "tipo": TipoFotografia.ACCESO,
                "descripcion": "El pasaje no tiene letrero; es la tercera casa.",
                "imagen": imagen_de_prueba(nombre="IMG_9987.jpg"),
            },
        )
        foto = Fotografia.objects.get()
        self.assertEqual(foto.tomada_por, self.marta)
        self.assertNotIn("IMG_9987", foto.imagen.name)

        # 3. Se ve por la vista protegida, nunca por la ruta del archivo.
        #
        # Se cierra la respuesta a mano: FileResponse deja el archivo ABIERTO hasta
        # que se cierra, y en Windows un archivo abierto no se puede borrar. En
        # producción lo cierra el servidor al terminar de enviarlo; en una prueba
        # que después borra el archivo, hay que hacerlo aquí.
        url_ver = reverse("fichas:ver_fotografia", kwargs={"pk": foto.pk})
        respuesta = self.client.get(url_ver)
        self.assertEqual(respuesta.status_code, 200)
        respuesta.close()

        # 4. Juan, que no tiene ese sector, no la ve.
        self.client.force_login(self.juan)
        self.assertEqual(self.client.get(url_ver).status_code, 404)

        # 5. Un visitante sin sesión, tampoco.
        self.client.logout()
        self.assertEqual(self.client.get(url_ver).status_code, 302)

        # 6. Marta la quita y el archivo desaparece del disco.
        self.client.force_login(self.marta)
        ruta = Path(foto.imagen.path)
        self.client.post(
            reverse("fichas:quitar_fotografia", kwargs={"pk": foto.pk})
        )

        self.assertEqual(Fotografia.objects.count(), 0)
        self.assertFalse(ruta.exists())


# ==========================================================================
# HU-13 — 57. LA REVISIÓN DEL SUPERVISOR
# ==========================================================================


class BaseRevisionTest(BaseEncuestaTest):
    """Escenario común: encuestas de dos personas en varios estados."""

    def setUp(self):
        super().setUp()
        AsignacionSector.objects.create(sector=self.boldos, censista=self.marta)
        self.url_bandeja = reverse("fichas:bandeja_revision")

    def url_revisar(self, encuesta):
        return reverse("fichas:revisar_encuesta", kwargs={"pk": encuesta.pk})

    def recibida(self, direccion="Av. Central 100", hace_dias=0, censista=None, **extra):
        """Una encuesta COMPLETADA, enviada hace N días."""
        encuesta = self.crear(
            direccion=direccion,
            censista=censista,
            estado=EstadoEncuesta.COMPLETADA,
            **extra,
        )
        encuesta.cerrada_en = timezone.now() - timedelta(days=hace_dias)
        encuesta.save(update_fields=["cerrada_en"])
        return encuesta

    def con_hogar(self, encuesta, declarados=2, personas=0):
        hogar = GrupoFamiliar.objects.create(
            encuesta=encuesta,
            jefe_hogar_nombre="Rosa Elena Millán",
            integrantes_declarados=declarados,
        )
        for numero in range(personas):
            Integrante.objects.create(
                grupo_familiar=hogar,
                parentesco=(
                    Parentesco.JEFE_HOGAR if numero == 0 else Parentesco.HIJO
                ),
                nombres=f"Persona {numero}",
                apellidos="Millán",
                sexo=Sexo.FEMENINO,
                fecha_nacimiento=timezone.localdate() - timedelta(days=30 * 365),
                nivel_educacional=NivelEducacional.MEDIA_COMPLETA,
                situacion_ocupacional=SituacionOcupacional.TRABAJA,
            )
        return hogar


class EsperaDeRevisionTest(BaseRevisionTest):
    def test_solo_las_completadas_estan_en_revision(self):
        for estado in EstadoEncuesta.values:
            with self.subTest(estado=estado):
                encuesta = self.crear(direccion=f"Calle {estado}", estado=estado)
                self.assertEqual(
                    encuesta.esta_en_revision, estado == EstadoEncuesta.COMPLETADA
                )

    def test_cuenta_los_dias_desde_que_se_envio(self):
        encuesta = self.recibida(hace_dias=4)

        self.assertEqual(encuesta.dias_esperando(), 4)

    def test_una_recien_enviada_lleva_cero_dias(self):
        """Cero es «llegó hoy», que es distinto de «no está esperando»."""
        self.assertEqual(self.recibida().dias_esperando(), 0)

    def test_una_que_no_espera_no_tiene_dias(self):
        encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)

        self.assertIsNone(encuesta.dias_esperando())

    def test_una_validada_tampoco(self):
        encuesta = self.crear(direccion="Calle 2", estado=EstadoEncuesta.VALIDADA)

        self.assertIsNone(encuesta.dias_esperando())

    def test_la_espera_se_puede_medir_a_una_fecha_dada(self):
        encuesta = self.recibida(hace_dias=2)
        futuro = timezone.localdate() + timedelta(days=5)

        self.assertEqual(encuesta.dias_esperando(a_fecha=futuro), 7)

    def test_una_espera_corta_no_es_prolongada(self):
        self.assertFalse(self.recibida(hace_dias=2).espera_prolongada)

    def test_una_espera_larga_si_lo_es(self):
        """Pasado el umbral, devolverla cuesta otra visita completa."""
        self.assertTrue(self.recibida(hace_dias=10).espera_prolongada)

    def test_el_umbral_es_inclusivo(self):
        self.assertTrue(
            self.recibida(hace_dias=Encuesta.DIAS_ESPERA_PROLONGADA).espera_prolongada
        )

    def test_una_que_no_espera_nunca_es_prolongada(self):
        encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)

        self.assertFalse(encuesta.espera_prolongada)


class ResumenParaRevisionTest(BaseRevisionTest):
    def test_una_encuesta_sin_hogar_lo_dice(self):
        datos = self.recibida().resumen_para_revision()

        self.assertFalse(datos["tiene_hogar"])
        self.assertEqual(datos["personas"], 0)

    def test_cuenta_personas_registradas_y_declaradas(self):
        encuesta = self.recibida()
        self.con_hogar(encuesta, declarados=4, personas=2)

        datos = encuesta.resumen_para_revision()

        self.assertEqual(datos["personas"], 2)
        self.assertEqual(datos["declaradas"], 4)

    def test_informa_si_la_vivienda_esta_descrita(self):
        self.assertTrue(self.recibida().resumen_para_revision()["vivienda_descrita"])

    def test_informa_si_falta_la_descripcion(self):
        vivienda = Vivienda.objects.create(zona=self.zona1, direccion="Sin describir")
        encuesta = self.recibida(vivienda=vivienda, direccion="Sin describir")

        self.assertFalse(encuesta.resumen_para_revision()["vivienda_descrita"])

    def test_informa_si_falta_la_ubicacion(self):
        self.assertFalse(self.recibida().resumen_para_revision()["tiene_ubicacion"])

    def test_cuenta_las_fotografias(self):
        self.assertEqual(self.recibida().resumen_para_revision()["fotografias"], 0)


# ==========================================================================
# HU-13 — 58. LA BANDEJA
# ==========================================================================


class BandejaRevisionTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    def test_el_supervisor_entra(self):
        respuesta = self.client.get(self.url_bandeja)

        self.assertEqual(respuesta.status_code, 200)

    def test_el_administrador_entra(self):
        self.client.force_login(self.admin)

        self.assertEqual(self.client.get(self.url_bandeja).status_code, 200)

    def test_el_encuestador_no_entra(self):
        """Tiene `ver_propias` pero no `ver_todas`: la bandeja no es para él."""
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_bandeja).status_code, 302)

    def test_un_visitante_anonimo_va_al_login(self):
        self.client.logout()

        respuesta = self.client.get(self.url_bandeja)

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(reverse("usuarios:login"), respuesta.url)

    def test_muestra_las_recibidas_por_defecto(self):
        self.recibida(direccion="Av. Central 100")

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "Av. Central 100")

    def test_no_muestra_las_pendientes(self):
        """Una puerta que nadie tocó no es trabajo recibido."""
        self.crear(direccion="Nunca visitada")

        respuesta = self.client.get(self.url_bandeja)

        self.assertNotContains(respuesta, "Nunca visitada")

    def test_no_muestra_los_borradores(self):
        self.crear(direccion="A medias", estado=EstadoEncuesta.BORRADOR)

        respuesta = self.client.get(self.url_bandeja)

        self.assertNotContains(respuesta, "A medias")

    def test_muestra_las_de_todos_los_encuestadores(self):
        """Es la diferencia con «Mis encuestas»: aquí se ve el trabajo de todos."""
        self.recibida(direccion="De Marta", censista=self.marta)
        self.recibida(direccion="De Juan", censista=self.juan)

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "De Marta")
        self.assertContains(respuesta, "De Juan")

    # -- el orden ----------------------------------------------------------

    def test_la_que_lleva_mas_esperando_va_primero(self):
        """Orden de llegada: es la única política justa para una cola."""
        self.recibida(direccion="Reciente", hace_dias=1)
        self.recibida(direccion="Antigua", hace_dias=9)

        respuesta = self.client.get(self.url_bandeja)
        direcciones = [e.direccion for e in respuesta.context["encuestas"]]

        self.assertEqual(direcciones, ["Antigua", "Reciente"])

    def test_destaca_la_mas_antigua_si_lleva_demasiado(self):
        self.recibida(direccion="Antigua", hace_dias=12)

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "12 días")
        self.assertContains(respuesta, "devolverla con observaciones")

    def test_no_avisa_si_la_cola_esta_al_dia(self):
        self.recibida(hace_dias=1)

        respuesta = self.client.get(self.url_bandeja)

        self.assertNotContains(respuesta, "devolverla con observaciones")

    # -- los contadores ----------------------------------------------------

    def test_cuenta_cada_estado(self):
        self.recibida(direccion="A")
        self.crear(direccion="B", estado=EstadoEncuesta.VALIDADA)
        self.crear(direccion="C", estado=EstadoEncuesta.OBSERVADA)
        self.crear(direccion="D", estado=EstadoEncuesta.NO_UBICADA)

        resumen = self.client.get(self.url_bandeja).context["resumen"]

        self.assertEqual(resumen["recibidas"], 1)
        self.assertEqual(resumen["validadas"], 1)
        self.assertEqual(resumen["observadas"], 1)
        self.assertEqual(resumen["sin_levantar"], 1)

    def test_los_contadores_no_cambian_al_filtrar(self):
        self.recibida(direccion="A")
        self.crear(direccion="B", estado=EstadoEncuesta.VALIDADA)

        resumen = self.client.get(
            self.url_bandeja, {"estado": EstadoEncuesta.VALIDADA}
        ).context["resumen"]

        self.assertEqual(resumen["recibidas"], 1)

    def test_sin_nada_por_revisar_lo_dice(self):
        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "Todo lo recibido está revisado")

    # -- las señales de la fila -------------------------------------------

    def test_avisa_de_una_encuesta_sin_hogar(self):
        self.recibida()

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "sin hogar")

    def test_avisa_de_personas_faltantes(self):
        encuesta = self.recibida()
        self.con_hogar(encuesta, declarados=4, personas=1)

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "1 de 4 personas")

    def test_avisa_de_una_vivienda_sin_describir(self):
        vivienda = Vivienda.objects.create(zona=self.zona1, direccion="Sin describir")
        self.recibida(vivienda=vivienda, direccion="Sin describir")

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "vivienda sin describir")

    def test_avisa_de_la_falta_de_gps(self):
        self.recibida()

        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "sin GPS")


class FiltrosRevisionTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)
        self.de_marta = self.recibida(direccion="Av. Central 100", censista=self.marta)
        self.de_juan = self.crear(
            direccion="Calle Juan 1",
            censista=self.juan,
            estado=EstadoEncuesta.VALIDADA,
        )

    def direcciones(self, **filtros):
        respuesta = self.client.get(self.url_bandeja, filtros)
        return [e.direccion for e in respuesta.context["encuestas"]]

    def test_por_defecto_solo_las_recibidas(self):
        self.assertEqual(self.direcciones(), ["Av. Central 100"])

    def test_filtra_por_validadas(self):
        self.assertEqual(
            self.direcciones(estado=EstadoEncuesta.VALIDADA), ["Calle Juan 1"]
        )

    def test_el_grupo_revisadas_junta_validadas_y_devueltas(self):
        self.crear(direccion="Devuelta 1", estado=EstadoEncuesta.OBSERVADA)

        direcciones = self.direcciones(estado="REVISADAS")

        self.assertEqual(set(direcciones), {"Calle Juan 1", "Devuelta 1"})

    def test_el_grupo_todas_incluye_lo_cerrado_sin_levantar(self):
        self.crear(direccion="No ubicada 1", estado=EstadoEncuesta.NO_UBICADA)

        self.assertEqual(len(self.direcciones(estado="TODAS")), 3)

    def test_filtra_por_encuestador(self):
        self.assertEqual(
            self.direcciones(estado="TODAS", censista=self.juan.pk), ["Calle Juan 1"]
        )

    def test_filtra_por_sector(self):
        self.assertEqual(
            len(self.direcciones(estado="TODAS", sector=self.boldos.pk)), 2
        )

    def test_filtra_por_operativo(self):
        self.assertEqual(
            len(self.direcciones(estado="TODAS", operativo=self.operativo.pk)), 2
        )

    def test_busca_por_direccion(self):
        self.assertEqual(self.direcciones(q="central"), ["Av. Central 100"])

    def test_busca_por_jefe_de_hogar(self):
        self.con_hogar(self.de_marta)

        self.assertEqual(self.direcciones(q="Rosa Elena"), ["Av. Central 100"])

    def test_un_estado_inventado_no_rompe_la_pantalla(self):
        respuesta = self.client.get(self.url_bandeja, {"estado": "INVENTADO"})

        self.assertEqual(respuesta.status_code, 200)

    def test_los_desplegables_solo_ofrecen_lo_que_tiene_encuestas(self):
        """Una opción que siempre devuelve una lista vacía es una opción que estorba."""
        formulario = FiltroRevisionForm()

        self.assertEqual(list(formulario.fields["operativo"].queryset), [self.operativo])
        self.assertEqual(
            set(formulario.fields["censista"].queryset), {self.marta, self.juan}
        )

    def test_los_filtros_sobreviven_al_cambiar_de_pagina(self):
        respuesta = self.client.get(self.url_bandeja, {"q": "central"})

        self.assertIn("q=central", respuesta.context["parametros"])


class ConsultasBandejaTest(BaseRevisionTest):
    """La bandeja no debe pagar una consulta por fila."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    def poblar(self, cuantas, desde=0):
        for numero in range(desde, desde + cuantas):
            encuesta = self.recibida(direccion=f"Calle {numero:02d}")
            self.con_hogar(encuesta, declarados=2, personas=1)

    def contar(self):
        with CaptureQueriesContext(connection) as captura:
            self.client.get(self.url_bandeja)
        return len(captura.captured_queries)

    def test_el_numero_de_consultas_no_crece_con_las_encuestas(self):
        self.poblar(2)
        con_dos = self.contar()

        self.poblar(5, desde=2)
        con_siete = self.contar()

        self.assertEqual(con_dos, con_siete)


# ==========================================================================
# HU-13 — 59. LA PANTALLA DE REVISIÓN
# ==========================================================================


class RevisarEncuestaTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)
        self.encuesta = self.recibida(direccion="Av. Central 100", hace_dias=3)
        self.hogar = self.con_hogar(self.encuesta, declarados=2, personas=2)
        self.url = self.url_revisar(self.encuesta)

    def test_el_supervisor_abre_cualquier_encuesta(self):
        respuesta = self.client.get(self.url)

        self.assertEqual(respuesta.status_code, 200)

    def test_el_encuestador_no_entra_aunque_sea_suya(self):
        """Su lectura es la ficha de la HU-07; esta pantalla es para supervisar."""
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url).status_code, 302)

    def test_muestra_la_vivienda_completa(self):
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "Albañilería")
        self.assertContains(respuesta, "Arrendada")

    def test_muestra_el_hogar(self):
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "Rosa Elena Millán")

    def test_muestra_a_las_personas_una_por_una(self):
        respuesta = self.client.get(self.url)

        self.assertEqual(len(respuesta.context["integrantes"]), 2)
        self.assertContains(respuesta, "Persona 0")

    def test_muestra_los_dias_de_espera(self):
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "esperando hace 3 días")

    def test_avisa_si_faltan_personas(self):
        otra = self.recibida(direccion="Calle 2")
        self.con_hogar(otra, declarados=5, personas=1)

        respuesta = self.client.get(self.url_revisar(otra))

        self.assertContains(respuesta, "faltan 4")

    def test_avisa_si_el_nombre_del_jefe_no_coincide(self):
        otra = self.recibida(direccion="Calle 3")
        hogar = self.con_hogar(otra, declarados=1, personas=0)
        Integrante.objects.create(
            grupo_familiar=hogar,
            parentesco=Parentesco.JEFE_HOGAR,
            nombres="Otro",
            apellidos="Nombre",
            sexo=Sexo.MASCULINO,
            fecha_nacimiento=timezone.localdate() - timedelta(days=30 * 365),
            nivel_educacional=NivelEducacional.MEDIA_COMPLETA,
            situacion_ocupacional=SituacionOcupacional.TRABAJA,
        )

        respuesta = self.client.get(self.url_revisar(otra))

        self.assertContains(respuesta, "se registró a nombre de")

    def test_muestra_el_motivo_de_un_cierre_sin_datos(self):
        cerrada = self.crear(
            direccion="No ubicada 1",
            estado=EstadoEncuesta.NO_UBICADA,
            motivo_cierre="La dirección no existe en ese pasaje.",
        )

        respuesta = self.client.get(self.url_revisar(cerrada))

        self.assertContains(respuesta, "no existe en ese pasaje")

    def test_avisa_de_los_otros_hogares_de_la_vivienda(self):
        """Explica un ingreso por persona extraño o gente que no cuadra."""
        self.recibida(vivienda=self.encuesta.vivienda, direccion="Av. Central 100")

        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "hogar")
        self.assertEqual(len(respuesta.context["otros_hogares"]), 1)

    def test_muestra_la_nota_del_encuestador(self):
        self.encuesta.nota_avance = "Volví tres veces hasta encontrarlos."
        self.encuesta.save()

        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, "Volví tres veces")

    def test_ofrece_resolver_a_quien_puede_validar(self):
        """Hasta la HU-13 esta pantalla no ofrecía acciones; la HU-14 las agregó."""
        respuesta = self.client.get(self.url)

        self.assertContains(respuesta, reverse("fichas:validar_encuesta", args=[self.encuesta.pk]))
        self.assertContains(respuesta, reverse("fichas:anular_encuesta", args=[self.encuesta.pk]))

    def test_no_ofrece_resolver_a_quien_solo_puede_mirar(self):
        """Leer y firmar son permisos distintos: es el corte que la HU-13 preparó."""
        self.rol_supervisor.permisos.remove(
            Permiso.objects.get(codigo="fichas.validar")
        )

        respuesta = self.client.get(self.url)

        self.assertNotContains(
            respuesta, reverse("fichas:validar_encuesta", args=[self.encuesta.pk])
        )

    def test_una_encuesta_pendiente_no_se_puede_revisar(self):
        pendiente = self.crear(direccion="Nunca visitada")

        respuesta = self.client.get(self.url_revisar(pendiente))

        self.assertEqual(respuesta.status_code, 404)


# ==========================================================================
# HU-13 — 60. EL MENÚ Y EL PANEL DEL SUPERVISOR
# ==========================================================================


class SupervisionEnElPanelTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.url_panel = reverse("dashboards:supervisor")

    def test_el_enlace_a_la_bandeja_aparece_para_el_supervisor(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, self.url_bandeja)

    def test_el_encuestador_no_ve_el_enlace(self):
        self.client.force_login(self.marta)

        respuesta = self.client.get(reverse("dashboards:censista"))

        self.assertNotContains(respuesta, self.url_bandeja)

    def test_el_panel_cuenta_lo_que_espera_revision(self):
        self.recibida()
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["resumen_revision"]["recibidas"], 1)

    def test_el_panel_muestra_la_cabeza_de_la_cola(self):
        self.recibida(direccion="Antigua", hace_dias=9)
        self.recibida(direccion="Reciente", hace_dias=1)
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["cola_revision"][0].direccion, "Antigua")

    def test_el_panel_muestra_como_maximo_cinco(self):
        for numero in range(8):
            self.recibida(direccion=f"Calle {numero}", hace_dias=numero)
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(len(respuesta.context["cola_revision"]), 5)

    def test_sin_cola_el_panel_lo_dice(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, "No hay encuestas esperando revisión")


# ==========================================================================
# HU-13 — 61. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionHU13Test(BaseRevisionTest):
    def test_del_envio_del_encuestador_a_la_bandeja_del_supervisor(self):
        # 1. Marta envía dos encuestas, una antigua y una de hoy.
        antigua = self.recibida(direccion="Av. Central 100", hace_dias=10)
        self.con_hogar(antigua, declarados=3, personas=1)
        self.recibida(direccion="Av. Central 118", hace_dias=0)

        # 2. El supervisor abre la bandeja y ve las dos, la antigua primero.
        self.client.force_login(self.supervisor)
        respuesta = self.client.get(self.url_bandeja)

        self.assertEqual(respuesta.context["resumen"]["recibidas"], 2)
        self.assertEqual(
            respuesta.context["encuestas"][0].direccion, "Av. Central 100"
        )

        # 3. La bandeja avisa de que la más antigua lleva demasiado.
        self.assertContains(respuesta, "10 días")

        # 4. Y muestra, sin abrirla, que a esa le faltan personas.
        self.assertContains(respuesta, "1 de 3 personas")

        # 5. Al abrirla, ve todo lo levantado en una pantalla.
        respuesta = self.client.get(self.url_revisar(antigua))
        self.assertContains(respuesta, "Rosa Elena Millán")
        self.assertContains(respuesta, "faltan 2")

        # 6. Marta, que la levantó, no entra a la pantalla de revisión.
        self.client.force_login(self.marta)
        self.assertEqual(self.client.get(self.url_revisar(antigua)).status_code, 302)

        # 7. Pero sigue viendo su propia ficha, que es otra lectura del mismo dato.
        self.assertEqual(self.client.get(self.url_detalle(antigua)).status_code, 200)


# ==========================================================================
# HU-14 — 62. APROBAR O ANULAR
# ==========================================================================


class BaseResolucionTest(BaseRevisionTest):
    """Escenario común: una encuesta recibida, lista para resolverse."""

    def setUp(self):
        super().setUp()
        self.encuesta = self.recibida(direccion="Av. Central 100", hace_dias=2)
        self.hogar = self.con_hogar(self.encuesta, declarados=2, personas=2)
        self.url_validar = reverse(
            "fichas:validar_encuesta", kwargs={"pk": self.encuesta.pk}
        )
        self.url_anular = reverse(
            "fichas:anular_encuesta", kwargs={"pk": self.encuesta.pk}
        )

    def datos_anulacion(self, **extra):
        base = {
            "causa": "Duplicada",
            "motivo": "La misma vivienda está levantada en la ficha del Pasaje 4.",
            "confirmar": "on",
        }
        base.update(extra)
        return base


class EstadoAnuladaTest(BaseResolucionTest):
    """El octavo estado y su diferencia con los dos que se le parecen."""

    def test_anulada_es_un_estado_cerrado(self):
        encuesta = self.crear(direccion="Anulada 1", estado=EstadoEncuesta.ANULADA)

        self.assertTrue(encuesta.esta_cerrada)
        self.assertFalse(encuesta.requiere_trabajo)

    def test_los_dos_grupos_siguen_cubriendo_todos_los_estados(self):
        """La partición de la HU-07 tiene que seguir siendo exhaustiva con ocho."""
        agrupados = set(ESTADOS_ABIERTOS) | set(ESTADOS_CERRADOS)

        self.assertEqual(agrupados, set(EstadoEncuesta.values))

    def test_anulada_no_es_lo_mismo_que_rechazada(self):
        """Una la decide el supervisor y la otra la familia: no se pueden mezclar."""
        self.assertNotEqual(EstadoEncuesta.ANULADA, EstadoEncuesta.RECHAZADA)
        self.assertNotIn(EstadoEncuesta.ANULADA, ESTADOS_SIN_LEVANTAR)

    def test_anulada_es_una_resolucion_del_supervisor(self):
        self.assertIn(EstadoEncuesta.ANULADA, ESTADOS_RESUELTOS)
        self.assertIn(EstadoEncuesta.VALIDADA, ESTADOS_RESUELTOS)
        self.assertIn(EstadoEncuesta.OBSERVADA, ESTADOS_RESUELTOS)

    def test_una_encuesta_anulada_no_admite_cambios_del_encuestador(self):
        encuesta = self.crear(direccion="Anulada 2", estado=EstadoEncuesta.ANULADA)

        permitido, _ = encuesta.puede_registrarse()

        self.assertFalse(permitido)

    def test_anular_exige_comentario_en_la_base_de_datos(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(direccion="Otra 1"),
                    censista=self.marta,
                    estado=EstadoEncuesta.ANULADA,
                    iniciada_en=timezone.now(),
                    cerrada_en=timezone.now(),
                    comentario_revision="",
                )

    def test_validar_no_exige_comentario(self):
        """Aprobar es el resultado esperado: exigir un texto daría cientos de «ok»."""
        encuesta = self.crear(direccion="Validada 1", estado=EstadoEncuesta.VALIDADA)

        self.assertEqual(encuesta.comentario_revision, "")


class PuedeResolverlaTest(BaseResolucionTest):
    def test_una_recibida_se_puede_resolver(self):
        permitido, motivo = self.encuesta.puede_resolverla(self.supervisor)

        self.assertTrue(permitido)
        self.assertEqual(motivo, "")

    def test_un_borrador_no_se_puede_resolver(self):
        """No lo ha enviado nadie todavía."""
        encuesta = self.crear(direccion="Borrador 1", estado=EstadoEncuesta.BORRADOR)

        permitido, motivo = encuesta.puede_resolverla(self.supervisor)

        self.assertFalse(permitido)
        self.assertIn("no está esperando revisión", motivo)

    def test_una_ya_validada_no_se_vuelve_a_resolver(self):
        encuesta = self.crear(direccion="Validada 1", estado=EstadoEncuesta.VALIDADA)

        permitido, _ = encuesta.puede_resolverla(self.supervisor)

        self.assertFalse(permitido)

    def test_una_observada_tampoco(self):
        """Volvió al encuestador: hay que esperar a que la reenvíe."""
        encuesta = self.crear(direccion="Observada 1", estado=EstadoEncuesta.OBSERVADA)

        permitido, _ = encuesta.puede_resolverla(self.supervisor)

        self.assertFalse(permitido)

    def test_nadie_valida_su_propia_encuesta(self):
        """La separación de funciones de la HU-03, comprobada en el modelo."""
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        permitido, motivo = propia.puede_resolverla(self.supervisor)

        self.assertFalse(permitido)
        self.assertIn("control cruzado", motivo)

    def test_ni_siquiera_el_administrador_valida_la_suya(self):
        """Tiene todos los permisos por definición, y aun así no puede."""
        propia = self.recibida(direccion="Del admin", censista=self.admin)

        permitido, _ = propia.puede_resolverla(self.admin)

        self.assertFalse(permitido)

    def test_otra_persona_si_puede_resolverla(self):
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        permitido, _ = propia.puede_resolverla(self.admin)

        self.assertTrue(permitido)


class ResolverTest(BaseResolucionTest):
    def test_resolver_deja_constancia_de_quien_y_cuando(self):
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.revisada_por, self.supervisor)
        self.assertIsNotNone(self.encuesta.revisada_en)

    def test_resolver_cambia_el_estado(self):
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.VALIDADA)

    def test_resolver_guarda_el_comentario_recortado(self):
        self.encuesta.resolver(
            EstadoEncuesta.VALIDADA, usuario=self.supervisor, comentario="  Revisada  "
        )

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.comentario_revision, "Revisada")

    def test_anular_con_comentario_funciona(self):
        self.encuesta.resolver(
            EstadoEncuesta.ANULADA,
            usuario=self.supervisor,
            comentario="Duplicada: ya está en otra ficha.",
        )

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.ANULADA)

    def test_resolver_marca_la_encuesta_como_resuelta(self):
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        self.assertTrue(self.encuesta.esta_resuelta)

    def test_una_recibida_todavia_no_esta_resuelta(self):
        self.assertFalse(self.encuesta.esta_resuelta)

    def test_resolver_conserva_la_fecha_de_envio(self):
        """Validar no cambia cuándo el encuestador terminó su trabajo."""
        enviada = self.encuesta.cerrada_en

        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.cerrada_en, enviada)

    def test_devuelve_la_propia_encuesta(self):
        self.assertIs(
            self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor),
            self.encuesta,
        )


# ==========================================================================
# HU-14 — 63. LOS FORMULARIOS
# ==========================================================================


class ValidarFormTest(BaseResolucionTest):
    def test_sin_comentario_es_valido(self):
        formulario = ValidarEncuestaForm({"comentario": ""})

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_recorta_el_comentario(self):
        formulario = ValidarEncuestaForm({"comentario": "  ok  "})
        formulario.is_valid()

        self.assertEqual(formulario.cleaned_data["comentario"], "ok")


class AnularFormTest(BaseResolucionTest):
    def test_un_formulario_completo_es_valido(self):
        formulario = AnularEncuestaForm(self.datos_anulacion())

        self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_la_causa_es_obligatoria(self):
        formulario = AnularEncuestaForm(self.datos_anulacion(causa=""))

        self.assertFalse(formulario.is_valid())
        self.assertIn("causa", formulario.errors)

    def test_una_causa_inventada_se_rechaza(self):
        formulario = AnularEncuestaForm(self.datos_anulacion(causa="Porque sí"))

        self.assertFalse(formulario.is_valid())

    def test_el_motivo_es_obligatorio(self):
        formulario = AnularEncuestaForm(self.datos_anulacion(motivo=""))

        self.assertFalse(formulario.is_valid())

    def test_un_motivo_de_una_letra_no_sirve(self):
        """Lo va a leer la persona cuyo trabajo se descarta."""
        formulario = AnularEncuestaForm(self.datos_anulacion(motivo="x"))

        self.assertFalse(formulario.is_valid())
        self.assertIn("motivo", formulario.errors)

    def test_hay_que_confirmar_la_consecuencia(self):
        datos = self.datos_anulacion()
        del datos["confirmar"]

        formulario = AnularEncuestaForm(datos)

        self.assertFalse(formulario.is_valid())
        self.assertIn("confirmar", formulario.errors)

    def test_el_comentario_junta_la_causa_y_la_explicacion(self):
        """La causa permite contar; el texto explica el caso concreto."""
        formulario = AnularEncuestaForm(self.datos_anulacion())
        formulario.is_valid()

        completo = formulario.comentario_completo()

        self.assertTrue(completo.startswith("Duplicada: "))
        self.assertIn("Pasaje 4", completo)


# ==========================================================================
# HU-14 — 64. LA PANTALLA DE VALIDAR
# ==========================================================================


class ValidarEncuestaVistaTest(BaseResolucionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    def test_muestra_la_confirmacion(self):
        respuesta = self.client.get(self.url_validar)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "¿Validar esta encuesta?")

    def test_muestra_el_resumen_de_lo_que_se_aprueba(self):
        """Última mirada antes de firmar: después el encuestador no la puede tocar."""
        respuesta = self.client.get(self.url_validar)

        self.assertContains(respuesta, "Rosa Elena Millán")
        self.assertContains(respuesta, "Lo que estás aprobando")

    def test_avisa_si_la_ficha_esta_incompleta(self):
        otra = self.recibida(direccion="Calle 2")
        self.con_hogar(otra, declarados=5, personas=1)

        respuesta = self.client.get(
            reverse("fichas:validar_encuesta", kwargs={"pk": otra.pk})
        )

        self.assertContains(respuesta, "datos incompletos")
        self.assertContains(respuesta, "Faltan 4 personas")

    def test_el_post_valida_la_encuesta(self):
        self.client.post(self.url_validar, {"comentario": ""})

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.VALIDADA)

    def test_registra_quien_valido(self):
        self.client.post(self.url_validar, {"comentario": ""})

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.revisada_por, self.supervisor)

    def test_guarda_el_comentario_opcional(self):
        self.client.post(self.url_validar, {"comentario": "Revisada con la familia."})

        self.encuesta.refresh_from_db()
        self.assertIn("Revisada con la familia", self.encuesta.comentario_revision)

    def test_vuelve_a_la_bandeja(self):
        respuesta = self.client.post(self.url_validar, {"comentario": ""})

        self.assertRedirects(respuesta, reverse("fichas:bandeja_revision"))

    def test_el_get_no_valida_nada(self):
        """Con un GET capaz de validar, un <img src> ajeno aprobaría fichas."""
        self.client.get(self.url_validar)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    # -- acceso ------------------------------------------------------------

    def test_sin_el_permiso_de_validar_no_se_entra(self):
        self.rol_supervisor.permisos.remove(
            Permiso.objects.get(codigo="fichas.validar")
        )

        respuesta = self.client.get(self.url_validar)

        self.assertEqual(respuesta.status_code, 302)

    def test_el_encuestador_no_entra(self):
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_validar).status_code, 302)

    def test_no_se_puede_validar_la_propia(self):
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        respuesta = self.client.get(
            reverse("fichas:validar_encuesta", kwargs={"pk": propia.pk})
        )

        self.assertEqual(respuesta.status_code, 302)

    def test_el_post_sobre_la_propia_tampoco(self):
        """Ocultar el botón no es una validación."""
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        self.client.post(
            reverse("fichas:validar_encuesta", kwargs={"pk": propia.pk}),
            {"comentario": ""},
        )

        propia.refresh_from_db()
        self.assertEqual(propia.estado, EstadoEncuesta.COMPLETADA)

    def test_una_ya_validada_no_se_vuelve_a_validar(self):
        """Dos supervisores en la misma bandeja no pueden pisarse."""
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.admin)

        respuesta = self.client.post(self.url_validar, {"comentario": "otra vez"})

        self.assertEqual(respuesta.status_code, 302)
        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.revisada_por, self.admin)

    def test_una_pendiente_no_se_puede_validar(self):
        pendiente = self.crear(direccion="Nunca visitada")

        respuesta = self.client.get(
            reverse("fichas:validar_encuesta", kwargs={"pk": pendiente.pk})
        )

        self.assertEqual(respuesta.status_code, 404)


# ==========================================================================
# HU-14 — 65. LA PANTALLA DE ANULAR
# ==========================================================================


class AnularEncuestaVistaTest(BaseResolucionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    def test_muestra_el_formulario(self):
        respuesta = self.client.get(self.url_anular)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "¿Anular esta encuesta?")

    def test_explica_que_anular_no_es_devolver(self):
        """Es la confusión que hay que evitar antes de pulsar el botón."""
        respuesta = self.client.get(self.url_anular)

        self.assertContains(respuesta, "Anular no es devolver")

    def test_avisa_de_que_los_datos_no_se_borran(self):
        respuesta = self.client.get(self.url_anular)

        self.assertContains(respuesta, "no se borran")

    def test_muestra_lo_que_se_descarta(self):
        respuesta = self.client.get(self.url_anular)

        self.assertContains(respuesta, "Rosa Elena Millán")

    def test_el_post_anula_la_encuesta(self):
        self.client.post(self.url_anular, self.datos_anulacion())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.ANULADA)

    def test_guarda_la_causa_y_el_motivo(self):
        self.client.post(self.url_anular, self.datos_anulacion())

        self.encuesta.refresh_from_db()
        self.assertTrue(self.encuesta.comentario_revision.startswith("Duplicada: "))

    def test_registra_quien_anulo(self):
        self.client.post(self.url_anular, self.datos_anulacion())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.revisada_por, self.supervisor)

    def test_sin_confirmar_no_anula(self):
        datos = self.datos_anulacion()
        del datos["confirmar"]

        self.client.post(self.url_anular, datos)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_sin_motivo_no_anula(self):
        self.client.post(self.url_anular, self.datos_anulacion(motivo=""))

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_no_borra_el_hogar_ni_a_las_personas(self):
        """El registro de que alguien la levantó y otro la descartó es lo auditable."""
        self.client.post(self.url_anular, self.datos_anulacion())

        self.assertTrue(GrupoFamiliar.objects.filter(pk=self.hogar.pk).exists())
        self.assertEqual(self.hogar.integrantes.count(), 2)

    def test_el_encuestador_no_puede_anular(self):
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_anular).status_code, 302)

    def test_no_se_puede_anular_la_propia(self):
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        self.client.post(
            reverse("fichas:anular_encuesta", kwargs={"pk": propia.pk}),
            self.datos_anulacion(),
        )

        propia.refresh_from_db()
        self.assertEqual(propia.estado, EstadoEncuesta.COMPLETADA)


# ==========================================================================
# HU-14 — 66. LO QUE CAMBIA EN LAS PANTALLAS ANTERIORES
# ==========================================================================


class ResolucionEnLasPantallasTest(BaseResolucionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)
        self.url_revisar_esta = self.url_revisar(self.encuesta)

    def test_la_pantalla_de_revision_ofrece_las_dos_acciones(self):
        respuesta = self.client.get(self.url_revisar_esta)

        self.assertTrue(respuesta.context["puede_validar"])
        self.assertContains(respuesta, self.url_validar)
        self.assertContains(respuesta, self.url_anular)

    def test_explica_por_que_no_se_puede_resolver_la_propia(self):
        """Esconder los botones sin decir nada parecería un fallo."""
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        respuesta = self.client.get(self.url_revisar(propia))

        self.assertFalse(respuesta.context["puede_validar"])
        self.assertContains(respuesta, "control cruzado")

    def test_muestra_la_resolucion_ya_tomada(self):
        self.encuesta.resolver(
            EstadoEncuesta.ANULADA,
            usuario=self.supervisor,
            comentario="Duplicada: ya estaba levantada.",
        )

        respuesta = self.client.get(self.url_revisar_esta)

        self.assertContains(respuesta, "Resolución")
        self.assertContains(respuesta, "Duplicada: ya estaba levantada")

    def test_la_bandeja_cuenta_las_anuladas(self):
        self.crear(direccion="Anulada 1", estado=EstadoEncuesta.ANULADA)

        respuesta = self.client.get(self.url_bandeja)

        self.assertEqual(respuesta.context["resumen"]["anuladas"], 1)

    def test_la_bandeja_filtra_las_anuladas(self):
        self.crear(direccion="Anulada 1", estado=EstadoEncuesta.ANULADA)

        respuesta = self.client.get(
            self.url_bandeja, {"estado": EstadoEncuesta.ANULADA}
        )
        direcciones = [e.direccion for e in respuesta.context["encuestas"]]

        self.assertEqual(direcciones, ["Anulada 1"])

    def test_el_grupo_revisadas_incluye_las_anuladas(self):
        self.crear(direccion="Anulada 1", estado=EstadoEncuesta.ANULADA)
        self.crear(direccion="Validada 1", estado=EstadoEncuesta.VALIDADA)

        respuesta = self.client.get(self.url_bandeja, {"estado": "REVISADAS"})

        self.assertEqual(len(respuesta.context["encuestas"]), 2)

    def test_una_validada_sale_de_la_cola(self):
        self.client.post(self.url_validar, {"comentario": ""})

        respuesta = self.client.get(self.url_bandeja)

        self.assertEqual(respuesta.context["resumen"]["recibidas"], 0)

    def test_el_encuestador_ve_el_comentario_en_su_ficha(self):
        """La resolución es un dato que se consulta, no solo un hecho auditable."""
        self.encuesta.resolver(
            EstadoEncuesta.ANULADA,
            usuario=self.supervisor,
            comentario="Duplicada: ya estaba levantada.",
        )
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertEqual(respuesta.status_code, 200)


# ==========================================================================
# HU-14 — 67. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionHU14Test(BaseRevisionTest):
    def test_una_se_valida_y_otra_se_anula(self):
        buena = self.recibida(direccion="Av. Central 100", hace_dias=3)
        self.con_hogar(buena, declarados=2, personas=2)
        mala = self.recibida(direccion="Av. Central 118", hace_dias=1)
        self.con_hogar(mala, declarados=2, personas=2)

        self.client.force_login(self.supervisor)

        # 1. Las dos están en la cola.
        respuesta = self.client.get(self.url_bandeja)
        self.assertEqual(respuesta.context["resumen"]["recibidas"], 2)

        # 2. La primera se valida, con un comentario opcional.
        self.client.post(
            reverse("fichas:validar_encuesta", kwargs={"pk": buena.pk}),
            {"comentario": "Revisada junto al encuestador."},
        )
        buena.refresh_from_db()
        self.assertEqual(buena.estado, EstadoEncuesta.VALIDADA)
        self.assertEqual(buena.revisada_por, self.supervisor)

        # 3. La segunda resulta estar duplicada y se anula.
        self.client.post(
            reverse("fichas:anular_encuesta", kwargs={"pk": mala.pk}),
            {
                "causa": "Duplicada",
                "motivo": "Es la misma vivienda que Av. Central 100.",
                "confirmar": "on",
            },
        )
        mala.refresh_from_db()
        self.assertEqual(mala.estado, EstadoEncuesta.ANULADA)
        self.assertIn("Duplicada", mala.comentario_revision)

        # 4. Los datos de la anulada NO se borraron.
        self.assertTrue(hasattr(mala, "grupo_familiar"))

        # 5. La cola quedó vacía y los contadores lo reflejan.
        respuesta = self.client.get(self.url_bandeja)
        self.assertEqual(respuesta.context["resumen"]["recibidas"], 0)
        self.assertEqual(respuesta.context["resumen"]["validadas"], 1)
        self.assertEqual(respuesta.context["resumen"]["anuladas"], 1)

        # 6. Ninguna de las dos la puede modificar ya el encuestador.
        for encuesta in (buena, mala):
            permitido, _ = encuesta.puede_registrarse()
            self.assertFalse(permitido)

        # 7. Y el supervisor no puede resolver una encuesta suya.
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)
        respuesta = self.client.get(
            reverse("fichas:validar_encuesta", kwargs={"pk": propia.pk})
        )
        self.assertEqual(respuesta.status_code, 302)


# ==========================================================================
# HU-15 — 68. DEVOLVER CON OBSERVACIONES
# ==========================================================================


class BaseDevolucionTest(BaseResolucionTest):
    """Escenario común: una encuesta recibida y la dirección para devolverla."""

    def setUp(self):
        super().setUp()
        self.url_devolver = reverse(
            "fichas:devolver_encuesta", kwargs={"pk": self.encuesta.pk}
        )

    def datos_devolucion(self, **extra):
        base = {
            "aspectos": ["Integrantes del hogar"],
            "observaciones": "Faltan los dos hijos menores que declaró la familia.",
        }
        base.update(extra)
        return base


class EstadoObservadaTest(BaseDevolucionTest):
    """Lo que significa OBSERVADA, ahora que hay quien la ponga."""

    def test_devolver_reabre_la_encuesta(self):
        """Es la diferencia con las otras dos resoluciones: no cierra, reabre."""
        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")

        self.assertEqual(self.encuesta.estado, EstadoEncuesta.OBSERVADA)
        self.assertFalse(self.encuesta.esta_cerrada)
        self.assertTrue(self.encuesta.requiere_trabajo)

    def test_devolver_borra_la_fecha_de_cierre(self):
        """Si siguiera cerrada, seguiría contando como enviada."""
        self.assertIsNotNone(self.encuesta.cerrada_en)

        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")

        self.assertIsNone(self.encuesta.cerrada_en)

    def test_devolver_conserva_la_fecha_de_inicio(self):
        """El trabajo ya hecho no se pierde: la encuesta sigue siendo la misma."""
        iniciada = self.encuesta.iniciada_en

        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")

        self.assertEqual(self.encuesta.iniciada_en, iniciada)

    def test_una_devuelta_vuelve_a_admitir_cambios(self):
        """Sin esto, devolverla sería pedir una corrección imposible."""
        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")

        permitido, _ = self.encuesta.puede_registrarse()

        self.assertTrue(permitido)

    def test_devolver_exige_comentario_en_la_base_de_datos(self):
        """La regla no vive solo en el formulario: la tabla la garantiza."""
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(direccion="Sin motivo 1"),
                    censista=self.marta,
                    estado=EstadoEncuesta.OBSERVADA,
                    iniciada_en=timezone.now(),
                    comentario_revision="",
                )

    def test_la_restriccion_sigue_cubriendo_las_anuladas(self):
        """Renombrarla no debía perder lo que ya comprobaba (HU-14)."""
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Encuesta.objects.create(
                    vivienda=self.crear_vivienda(direccion="Sin motivo 2"),
                    censista=self.marta,
                    estado=EstadoEncuesta.ANULADA,
                    iniciada_en=timezone.now(),
                    cerrada_en=timezone.now(),
                    comentario_revision="",
                )

    def test_una_observada_no_esta_en_revision(self):
        """Está en manos del encuestador otra vez, no del supervisor."""
        self.devolver(self.encuesta)

        self.assertFalse(self.encuesta.esta_en_revision)
        self.assertIsNone(self.encuesta.dias_esperando())


class ContadorDeDevolucionesTest(BaseDevolucionTest):
    """`veces_devuelta`: la señal de que el problema ya no es la ficha."""

    def test_una_encuesta_nueva_no_se_ha_devuelto_nunca(self):
        self.assertEqual(self.encuesta.veces_devuelta, 0)
        self.assertFalse(self.encuesta.fue_devuelta)

    def test_devolver_suma_uno(self):
        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.veces_devuelta, 1)
        self.assertTrue(self.encuesta.fue_devuelta)

    def test_devolver_dos_veces_suma_dos(self):
        """El caso que justifica el campo: la segunda vuelta con el mismo error."""
        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")
        self.encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)
        self.encuesta.devolver(usuario=self.supervisor, observaciones="Siguen faltando.")

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.veces_devuelta, 2)

    def test_validar_no_toca_el_contador(self):
        """Aprobar no es devolver: contar ahí falsearía la señal."""
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.veces_devuelta, 0)

    def test_el_contador_sobrevive_a_la_validacion(self):
        """Que costara dos intentos es información de calidad, no ruido temporal."""
        self.encuesta.devolver(usuario=self.supervisor, observaciones="Faltan datos.")
        self.encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.VALIDADA)
        self.assertTrue(self.encuesta.fue_devuelta)

    def test_avisa_al_llegar_al_umbral(self):
        for veces in range(Encuesta.DEVOLUCIONES_PARA_ALERTAR):
            with self.subTest(veces=veces):
                self.assertFalse(self.encuesta.devuelta_repetidamente)
                self.encuesta.veces_devuelta = veces + 1

        self.assertTrue(self.encuesta.devuelta_repetidamente)

    def test_no_admite_valores_negativos(self):
        """PositiveSmallIntegerField: «devuelta menos una vez» no significa nada."""
        self.encuesta.veces_devuelta = -1

        with self.assertRaises(ValidationError):
            self.encuesta.full_clean()


class ProblemasDetectadosTest(BaseDevolucionTest):
    """Lo que el sistema sugiere escribir, para que el supervisor no parta de cero."""

    def test_una_ficha_completa_no_tiene_problemas_que_sugerir(self):
        """Y no significa que esté bien: significa que no falta nada CONTABLE."""
        self.encuesta.vivienda.latitud = Decimal("-36.8270")
        self.encuesta.vivienda.longitud = Decimal("-73.0503")
        self.encuesta.vivienda.save()

        self.assertEqual(self.encuesta.problemas_detectados(), [])

    def test_detecta_las_personas_que_faltan(self):
        encuesta = self.recibida(direccion="Faltan personas 1")
        self.con_hogar(encuesta, declarados=5, personas=3)

        problemas = " ".join(encuesta.problemas_detectados())

        self.assertIn("Faltan 2 personas", problemas)

    def test_usa_el_singular_cuando_falta_una(self):
        """«Faltan 1 personas» delata que nadie leyó el texto que se envía."""
        encuesta = self.recibida(direccion="Falta una 1")
        self.con_hogar(encuesta, declarados=3, personas=2)

        problemas = " ".join(encuesta.problemas_detectados())

        self.assertIn("Falta 1 persona por registrar", problemas)

    def test_detecta_que_no_hay_hogar(self):
        encuesta = self.recibida(direccion="Sin hogar 1")

        problemas = " ".join(encuesta.problemas_detectados())

        self.assertIn("datos del hogar", problemas)

    def test_sin_hogar_no_habla_de_personas_que_faltan(self):
        """Sin hogar no hay declaración con la que comparar: sobra el detalle."""
        encuesta = self.recibida(direccion="Sin hogar 2")

        problemas = " ".join(encuesta.problemas_detectados())

        self.assertNotIn("Faltan", problemas)

    def test_detecta_la_vivienda_sin_describir(self):
        vivienda = self.crear_vivienda(direccion="Sin describir 1", tipo="")
        encuesta = self.recibida(direccion="Sin describir 1", vivienda=vivienda)
        self.con_hogar(encuesta, declarados=2, personas=2)

        problemas = " ".join(encuesta.problemas_detectados())

        self.assertIn("vivienda no está descrita", problemas)

    def test_detecta_la_ubicacion_que_falta(self):
        problemas = " ".join(self.encuesta.problemas_detectados())

        self.assertIn("ubicación", problemas)

    def test_acumula_varios_problemas(self):
        """Devolverla dos veces por dos cosas que se veían juntas es un viaje de más."""
        vivienda = self.crear_vivienda(direccion="Todo mal 1", tipo="")
        encuesta = self.recibida(direccion="Todo mal 1", vivienda=vivienda)
        self.con_hogar(encuesta, declarados=4, personas=1)

        self.assertEqual(len(encuesta.problemas_detectados()), 3)


# ==========================================================================
# HU-15 — 69. EL FORMULARIO DE LA DEVOLUCIÓN
# ==========================================================================


class DevolverFormTest(BaseDevolucionTest):
    def test_valida_con_un_aspecto_y_observaciones(self):
        formulario = DevolverEncuestaForm(data=self.datos_devolucion())

        self.assertTrue(formulario.is_valid())

    def test_exige_al_menos_un_aspecto(self):
        """Los aspectos dicen DÓNDE mirar: sin ninguno, el texto queda suelto."""
        formulario = DevolverEncuestaForm(data=self.datos_devolucion(aspectos=[]))

        self.assertFalse(formulario.is_valid())
        self.assertIn("aspectos", formulario.errors)

    def test_no_acepta_un_aspecto_inventado(self):
        formulario = DevolverEncuestaForm(
            data=self.datos_devolucion(aspectos=["Lo que sea"])
        )

        self.assertFalse(formulario.is_valid())

    def test_exige_observaciones(self):
        formulario = DevolverEncuestaForm(data=self.datos_devolucion(observaciones=""))

        self.assertFalse(formulario.is_valid())

    def test_rechaza_una_observacion_de_una_palabra(self):
        """«revisar» le cuesta al encuestador un viaje para no saber qué mirar."""
        formulario = DevolverEncuestaForm(
            data=self.datos_devolucion(observaciones="revisar")
        )

        self.assertFalse(formulario.is_valid())
        self.assertIn("observaciones", formulario.errors)

    def test_los_espacios_no_cuentan_como_largo(self):
        formulario = DevolverEncuestaForm(
            data=self.datos_devolucion(observaciones="   ok   ")
        )

        self.assertFalse(formulario.is_valid())

    def test_no_pide_confirmacion_como_anular(self):
        """Devolver es reversible y esperable; anular no. Igualarlas gastaría la barrera."""
        self.assertNotIn("confirmar", DevolverEncuestaForm().fields)

    def test_el_comentario_junta_aspectos_y_texto(self):
        formulario = DevolverEncuestaForm(data=self.datos_devolucion())
        formulario.is_valid()

        comentario = formulario.comentario_completo()

        self.assertTrue(comentario.startswith("Corregir: Integrantes del hogar."))
        self.assertIn("Faltan los dos hijos menores", comentario)

    def test_los_aspectos_se_ordenan_siempre_igual(self):
        """Dos devoluciones por lo mismo tienen que leerse igual."""
        formulario = DevolverEncuestaForm(
            data=self.datos_devolucion(
                aspectos=["Ubicación", "Datos de la vivienda"]
            )
        )
        formulario.is_valid()

        self.assertIn(
            "Corregir: Datos de la vivienda, Ubicación.",
            formulario.comentario_completo(),
        )

    def test_prerrellena_lo_que_el_sistema_detecta(self):
        encuesta = self.recibida(direccion="Prerrelleno 1")
        self.con_hogar(encuesta, declarados=4, personas=1)

        formulario = DevolverEncuestaForm(encuesta=encuesta)

        self.assertIn("Faltan 3 personas", formulario.initial["observaciones"])

    def test_no_prerrellena_si_no_hay_nada_que_decir(self):
        self.encuesta.vivienda.latitud = Decimal("-36.8270")
        self.encuesta.vivienda.longitud = Decimal("-73.0503")
        self.encuesta.vivienda.save()

        formulario = DevolverEncuestaForm(encuesta=self.encuesta)

        self.assertNotIn("observaciones", formulario.initial)

    def test_no_pisa_lo_que_el_supervisor_escribio(self):
        """Si el formulario vuelve con un error, su texto tiene que seguir ahí."""
        formulario = DevolverEncuestaForm(
            data=self.datos_devolucion(aspectos=[], observaciones="Lo mío"),
            encuesta=self.encuesta,
        )
        formulario.is_valid()

        self.assertEqual(formulario.data["observaciones"], "Lo mío")

    def test_funciona_sin_encuesta(self):
        """El prerrelleno es una ayuda, no un requisito para instanciarlo."""
        self.assertTrue(DevolverEncuestaForm(data=self.datos_devolucion()).is_valid())


# ==========================================================================
# HU-15 — 70. LA PANTALLA DE DEVOLVER
# ==========================================================================


class DevolverEncuestaVistaTest(BaseDevolucionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    # -- quién entra -------------------------------------------------------

    def test_el_supervisor_ve_el_formulario(self):
        respuesta = self.client.get(self.url_devolver)

        self.assertEqual(respuesta.status_code, 200)
        self.assertTemplateUsed(respuesta, "fichas/encuesta_devolver.html")

    def test_el_encuestador_no_puede_devolver(self):
        """Exige `fichas.validar`, igual que validar y anular."""
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_devolver)

        self.assertEqual(respuesta.status_code, 302)

    def test_el_encuestador_tampoco_por_POST(self):
        self.client.force_login(self.marta)

        self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_sin_sesion_pide_entrar(self):
        self.client.logout()

        respuesta = self.client.get(self.url_devolver)

        self.assertEqual(respuesta.status_code, 302)

    def test_no_se_puede_devolver_la_propia(self):
        """Quien levanta no resuelve, ni para bien ni para mal."""
        propia = self.recibida(direccion="Del supervisor", censista=self.supervisor)

        respuesta = self.client.get(
            reverse("fichas:devolver_encuesta", kwargs={"pk": propia.pk})
        )

        self.assertEqual(respuesta.status_code, 302)

    def test_no_se_puede_devolver_una_que_no_espera_revision(self):
        borrador = self.crear(direccion="Borrador 1", estado=EstadoEncuesta.BORRADOR)

        respuesta = self.client.get(
            reverse("fichas:devolver_encuesta", kwargs={"pk": borrador.pk})
        )

        self.assertEqual(respuesta.status_code, 302)

    def test_no_se_puede_devolver_dos_veces_seguidas(self):
        """Dos supervisores en la misma bandeja: el segundo no pisa al primero."""
        self.client.post(self.url_devolver, self.datos_devolucion())

        respuesta = self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertEqual(respuesta.status_code, 302)
        self.assertEqual(self.encuesta.veces_devuelta, 1)

    # -- lo que muestra ----------------------------------------------------

    def test_dice_a_quien_vuelve_la_ficha(self):
        """Escribir para una persona con nombre no es lo mismo que para el sistema."""
        respuesta = self.client.get(self.url_devolver)

        self.assertContains(respuesta, "Marta Soto")

    def test_muestra_los_problemas_detectados(self):
        respuesta = self.client.get(self.url_devolver)

        self.assertContains(respuesta, "Lo que el sistema detecta")
        self.assertContains(respuesta, "ubicación")

    def test_prerrellena_el_texto_en_la_pantalla(self):
        encuesta = self.recibida(direccion="Prerrelleno 2")
        self.con_hogar(encuesta, declarados=5, personas=2)

        respuesta = self.client.get(
            reverse("fichas:devolver_encuesta", kwargs={"pk": encuesta.pk})
        )

        self.assertContains(respuesta, "Faltan 3 personas")

    def test_avisa_cuando_no_detecta_nada(self):
        """El silencio no debe leerse como «no hay nada que objetar»."""
        self.encuesta.vivienda.latitud = Decimal("-36.8270")
        self.encuesta.vivienda.longitud = Decimal("-73.0503")
        self.encuesta.vivienda.save()

        respuesta = self.client.get(self.url_devolver)

        self.assertContains(respuesta, "no detecta datos faltantes")

    def test_avisa_de_la_reincidencia(self):
        self.encuesta.veces_devuelta = Encuesta.DEVOLUCIONES_PARA_ALERTAR
        self.encuesta.save(update_fields=["veces_devuelta"])

        respuesta = self.client.get(self.url_devolver)

        self.assertContains(respuesta, "ya se devolvió")

    def test_no_avisa_de_reincidencia_la_primera_vez(self):
        respuesta = self.client.get(self.url_devolver)

        self.assertNotContains(respuesta, "ya se devolvió")

    # -- lo que hace -------------------------------------------------------

    def test_un_GET_no_devuelve_nada(self):
        """Si un GET devolviera fichas, un <img> ajeno las devolvería solo."""
        self.client.get(self.url_devolver)

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_el_POST_la_deja_observada(self):
        self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.OBSERVADA)

    def test_guarda_quien_la_devolvio_y_cuando(self):
        self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.revisada_por, self.supervisor)
        self.assertIsNotNone(self.encuesta.revisada_en)

    def test_guarda_las_observaciones_con_los_aspectos(self):
        self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertIn("Corregir: Integrantes del hogar.", self.encuesta.comentario_revision)
        self.assertIn("hijos menores", self.encuesta.comentario_revision)

    def test_suma_una_devolucion(self):
        self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.veces_devuelta, 1)

    def test_no_borra_nada_de_lo_registrado(self):
        """Devolver pide corregir, no empezar de cero."""
        self.client.post(self.url_devolver, self.datos_devolucion())

        self.encuesta.refresh_from_db()
        self.assertTrue(hasattr(self.encuesta, "grupo_familiar"))
        self.assertEqual(self.encuesta.grupo_familiar.integrantes.count(), 2)

    def test_un_formulario_incompleto_no_devuelve_nada(self):
        self.client.post(self.url_devolver, self.datos_devolucion(observaciones=""))

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)
        self.assertEqual(self.encuesta.veces_devuelta, 0)

    def test_un_formulario_incompleto_vuelve_con_el_error(self):
        respuesta = self.client.post(
            self.url_devolver, self.datos_devolucion(observaciones="ver")
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Escribe qué hay que corregir")

    def test_redirige_a_la_bandeja(self):
        """El trabajo del supervisor es la cola, no esta ficha."""
        respuesta = self.client.post(self.url_devolver, self.datos_devolucion())

        self.assertRedirects(respuesta, self.url_bandeja)

    def test_avisa_a_quien_volvio_la_ficha(self):
        respuesta = self.client.post(
            self.url_devolver, self.datos_devolucion(), follow=True
        )

        self.assertContains(respuesta, "devuelta a Marta Soto")

    def test_una_encuesta_inexistente_responde_404(self):
        respuesta = self.client.get(
            reverse("fichas:devolver_encuesta", kwargs={"pk": 9999})
        )

        self.assertEqual(respuesta.status_code, 404)


# ==========================================================================
# HU-15 — 71. LO QUE LEE EL ENCUESTADOR
# ==========================================================================


class ObservacionesParaElEncuestadorTest(BaseDevolucionTest):
    """La mitad de la historia: sin esto, la devolución no pide nada."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)
        self.client.post(self.url_devolver, self.datos_devolucion())
        self.client.force_login(self.marta)

    def test_el_encuestador_lee_las_observaciones_en_su_ficha(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Faltan los dos hijos menores")

    def test_lee_tambien_qué_aspectos_hay_que_corregir(self):
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Corregir: Integrantes del hogar.")

    def test_sabe_quien_la_devolvio(self):
        """Para poder preguntarle si no entiende qué falta."""
        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "Luis Pérez")

    def test_la_ficha_vuelve_a_su_lista_de_trabajo(self):
        respuesta = self.client.get(self.url_lista)
        direcciones = [e.direccion for e in respuesta.context["encuestas"]]

        self.assertIn("Av. Central 100", direcciones)

    def test_aparece_primera_por_urgencia(self):
        """Lo devuelto va delante: hay alguien esperando para cerrarlo."""
        self.crear(direccion="Pendiente 1", estado=EstadoEncuesta.PENDIENTE)

        respuesta = self.client.get(self.url_lista)
        direcciones = [e.direccion for e in respuesta.context["encuestas"]]

        self.assertEqual(direcciones[0], "Av. Central 100")

    def test_la_puede_editar_otra_vez(self):
        respuesta = self.client.get(
            reverse("fichas:registrar_hogar", kwargs={"pk": self.encuesta.pk})
        )

        self.assertEqual(respuesta.status_code, 200)

    def test_la_puede_volver_a_enviar(self):
        """El hogar ya está completo desde el escenario base: nada bloquea el envío."""
        self.client.post(
            reverse("fichas:completar_encuesta", kwargs={"pk": self.encuesta.pk})
        )

        self.encuesta.refresh_from_db()
        self.assertEqual(self.encuesta.estado, EstadoEncuesta.COMPLETADA)

    def test_otro_encuestador_no_ve_esas_observaciones(self):
        """Devolver no cambia de quién es la ficha."""
        self.client.force_login(self.juan)

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertEqual(respuesta.status_code, 404)

    def test_avisa_cuando_ya_van_dos_devoluciones(self):
        self.encuesta.cambiar_estado(EstadoEncuesta.COMPLETADA)
        self.client.force_login(self.supervisor)
        self.client.post(self.url_devolver, self.datos_devolucion())
        self.client.force_login(self.marta)

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertContains(respuesta, "2.ª vez")

    def test_el_comentario_de_una_validada_tambien_se_lee(self):
        """Aprobada con un comentario, el encuestador tiene derecho a leerlo."""
        otra = self.recibida(direccion="Validada con nota")
        otra.resolver(
            EstadoEncuesta.VALIDADA,
            usuario=self.supervisor,
            comentario="Revisada junto a la familia.",
        )

        respuesta = self.client.get(self.url_detalle(otra))

        self.assertContains(respuesta, "Revisada junto a la familia")

    def test_el_motivo_de_una_anulada_tambien_se_lee(self):
        """Es la ficha cuyo trabajo se descartó: el motivo es lo mínimo."""
        otra = self.recibida(direccion="Anulada con motivo")
        otra.resolver(
            EstadoEncuesta.ANULADA,
            usuario=self.supervisor,
            comentario="Duplicada: ya estaba levantada en el Pasaje 4.",
        )

        respuesta = self.client.get(self.url_detalle(otra))

        self.assertContains(respuesta, "Duplicada: ya estaba levantada")

    def test_el_texto_del_supervisor_no_se_interpreta_como_HTML(self):
        """El comentario es texto de una persona, no una plantilla."""
        self.encuesta.comentario_revision = "Revisar <script>alert(1)</script>"
        self.encuesta.save(update_fields=["comentario_revision"])

        respuesta = self.client.get(self.url_detalle(self.encuesta))

        self.assertNotContains(respuesta, "<script>alert(1)</script>")


# ==========================================================================
# HU-15 — 72. LO QUE CAMBIA EN LA REVISIÓN
# ==========================================================================


class DevolucionEnLaRevisionTest(BaseDevolucionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    def test_la_pantalla_de_revision_ofrece_las_tres_salidas(self):
        respuesta = self.client.get(self.url_revisar(self.encuesta))

        self.assertContains(respuesta, self.url_validar)
        self.assertContains(respuesta, self.url_devolver)
        self.assertContains(respuesta, self.url_anular)

    def test_no_ofrece_devolver_una_ya_resuelta(self):
        self.devolver(self.encuesta)

        respuesta = self.client.get(self.url_revisar(self.encuesta))

        self.assertNotContains(respuesta, self.url_devolver)

    def test_muestra_el_historial_de_devoluciones(self):
        self.devolver(self.encuesta)

        respuesta = self.client.get(self.url_revisar(self.encuesta))

        self.assertContains(respuesta, "Devoluciones")

    def test_no_habla_de_devoluciones_si_no_hubo(self):
        self.encuesta.resolver(EstadoEncuesta.VALIDADA, usuario=self.supervisor)

        respuesta = self.client.get(self.url_revisar(self.encuesta))

        self.assertNotContains(respuesta, "Devoluciones")

    def test_avisa_de_la_reincidencia_donde_se_decide(self):
        self.encuesta.veces_devuelta = Encuesta.DEVOLUCIONES_PARA_ALERTAR
        self.encuesta.save(update_fields=["veces_devuelta"])

        respuesta = self.client.get(self.url_revisar(self.encuesta))

        self.assertContains(respuesta, "hablar con")

    def test_una_devuelta_sale_de_la_cola(self):
        """Está en manos del encuestador: contarla como recibida sería mentir."""
        self.client.post(self.url_devolver, self.datos_devolucion())

        respuesta = self.client.get(self.url_bandeja)

        self.assertEqual(respuesta.context["resumen"]["recibidas"], 0)
        self.assertEqual(respuesta.context["resumen"]["observadas"], 1)

    def test_la_bandeja_filtra_las_devueltas(self):
        self.client.post(self.url_devolver, self.datos_devolucion())

        respuesta = self.client.get(
            self.url_bandeja, {"estado": EstadoEncuesta.OBSERVADA}
        )
        direcciones = [e.direccion for e in respuesta.context["encuestas"]]

        self.assertEqual(direcciones, ["Av. Central 100"])

    def test_el_panel_del_supervisor_cuenta_las_devueltas(self):
        self.client.post(self.url_devolver, self.datos_devolucion())

        respuesta = self.client.get(reverse("dashboards:supervisor"))

        self.assertEqual(respuesta.context["resumen_revision"]["observadas"], 1)


# ==========================================================================
# HU-15 — 73. RECORRIDO COMPLETO
# ==========================================================================


class IntegracionHU15Test(BaseRevisionTest):
    def test_se_devuelve_se_corrige_y_se_valida(self):
        """El circuito completo de una corrección, de punta a punta."""
        encuesta = self.recibida(direccion="Av. Central 100", hace_dias=2)
        self.con_hogar(encuesta, declarados=3, personas=1)

        # 1. El supervisor la encuentra en la cola.
        self.client.force_login(self.supervisor)
        respuesta = self.client.get(self.url_bandeja)
        self.assertEqual(respuesta.context["resumen"]["recibidas"], 1)

        # 2. La pantalla de devolver llega con las observaciones ya redactadas.
        url_devolver = reverse("fichas:devolver_encuesta", kwargs={"pk": encuesta.pk})
        respuesta = self.client.get(url_devolver)
        self.assertContains(respuesta, "Faltan 2 personas")

        # 3. La devuelve, agregando lo que solo él vio.
        self.client.post(
            url_devolver,
            {
                "aspectos": ["Integrantes del hogar", "Ubicación"],
                "observaciones": (
                    "Faltan los dos hijos menores y no se capturó la ubicación."
                ),
            },
        )
        encuesta.refresh_from_db()
        self.assertEqual(encuesta.estado, EstadoEncuesta.OBSERVADA)
        self.assertEqual(encuesta.veces_devuelta, 1)
        self.assertIsNone(encuesta.cerrada_en)

        # 4. Sale de la cola del supervisor.
        respuesta = self.client.get(self.url_bandeja)
        self.assertEqual(respuesta.context["resumen"]["recibidas"], 0)

        # 5. El encuestador la encuentra arriba de su lista y LEE qué corregir.
        self.client.force_login(self.marta)
        respuesta = self.client.get(self.url_lista)
        self.assertEqual(respuesta.context["encuestas"][0].direccion, "Av. Central 100")

        respuesta = self.client.get(self.url_detalle(encuesta))
        self.assertContains(respuesta, "Faltan los dos hijos menores")
        self.assertContains(respuesta, "Luis Pérez")

        # 6. Corrige lo que le pidieron: agrega las dos personas que faltaban.
        for numero in (2, 3):
            self.client.post(
                reverse(
                    "fichas:integrante_nuevo", kwargs={"encuesta_pk": encuesta.pk}
                ),
                {
                    "parentesco": Parentesco.HIJO,
                    "nombres": f"Hijo {numero}",
                    "apellidos": "Millán",
                    "sexo": Sexo.MASCULINO,
                    "fecha_nacimiento": (
                        timezone.localdate() - timedelta(days=10 * 365)
                    ).isoformat(),
                    "nivel_educacional": NivelEducacional.BASICA_INCOMPLETA,
                    "situacion_ocupacional": SituacionOcupacional.ESTUDIA,
                    "pueblo_originario": PuebloOriginario.NINGUNO,
                },
            )
        self.assertEqual(encuesta.grupo_familiar.integrantes.count(), 3)

        # 7. La vuelve a enviar.
        self.client.post(
            reverse("fichas:completar_encuesta", kwargs={"pk": encuesta.pk})
        )
        encuesta.refresh_from_db()
        self.assertEqual(encuesta.estado, EstadoEncuesta.COMPLETADA)

        # 8. Reaparece en la cola del supervisor, que ahora sí la valida.
        self.client.force_login(self.supervisor)
        respuesta = self.client.get(self.url_bandeja)
        self.assertEqual(respuesta.context["resumen"]["recibidas"], 1)

        self.client.post(
            reverse("fichas:validar_encuesta", kwargs={"pk": encuesta.pk}),
            {"comentario": "Corregida correctamente."},
        )
        encuesta.refresh_from_db()
        self.assertEqual(encuesta.estado, EstadoEncuesta.VALIDADA)

        # 9. Y queda registro de que costó dos intentos.
        self.assertEqual(encuesta.veces_devuelta, 1)
        self.assertTrue(encuesta.fue_devuelta)


# ==========================================================================
# HU-16 — 74. ALERTAS DE CALIDAD DE DATOS
# ==========================================================================


class DatosIncompletosTest(BaseRevisionTest):
    """`tiene_datos_incompletos` no mira el estado: eso lo decide quien cuenta."""

    def test_todo_completo_no_es_una_alerta(self):
        encuesta = self.recibida()
        self.con_hogar(encuesta, declarados=2, personas=2)
        encuesta.vivienda.latitud = -36.8270
        encuesta.vivienda.longitud = -73.0498
        encuesta.vivienda.save()

        self.assertFalse(encuesta.tiene_datos_incompletos)

    def test_sin_hogar_es_una_alerta(self):
        encuesta = self.recibida()

        self.assertTrue(encuesta.tiene_datos_incompletos)

    def test_menos_personas_que_declaradas_es_una_alerta(self):
        encuesta = self.recibida()
        self.con_hogar(encuesta, declarados=4, personas=2)

        self.assertTrue(encuesta.tiene_datos_incompletos)

    def test_vivienda_sin_describir_es_una_alerta(self):
        vivienda = Vivienda.objects.create(zona=self.zona1, direccion="Sin describir")
        encuesta = self.recibida(vivienda=vivienda, direccion="Sin describir")
        self.con_hogar(encuesta, declarados=1, personas=1)

        self.assertTrue(encuesta.tiene_datos_incompletos)

    def test_sin_ubicacion_es_una_alerta(self):
        encuesta = self.recibida()
        self.con_hogar(encuesta, declarados=1, personas=1)

        self.assertTrue(encuesta.tiene_datos_incompletos)

    def test_no_depende_del_estado(self):
        """Una vivienda sin describir es igual de real en cualquier estado.

        Que solo las COMPLETADA cuenten como alerta es una decisión de
        `resumen_alertas_calidad()`, no de esta propiedad.
        """
        encuesta = self.crear(estado=EstadoEncuesta.BORRADOR)

        self.assertTrue(encuesta.tiene_datos_incompletos)


class ResumenAlertasCalidadTest(BaseRevisionTest):
    def test_sin_encuestas_todo_en_cero(self):
        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(
            resumen,
            {"datos_incompletos": 0, "espera_prolongada": 0, "visitas_vencidas": 0},
        )

    def test_cuenta_datos_incompletos_solo_entre_las_completadas(self):
        self.recibida(direccion="Enviada incompleta")  # COMPLETADA, sin hogar
        self.crear(direccion="Borrador incompleto", estado=EstadoEncuesta.BORRADOR)

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["datos_incompletos"], 1)

    def test_una_completada_sin_problemas_no_cuenta(self):
        encuesta = self.recibida()
        self.con_hogar(encuesta, declarados=1, personas=1)
        encuesta.vivienda.latitud = -36.8270
        encuesta.vivienda.longitud = -73.0498
        encuesta.vivienda.save()

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["datos_incompletos"], 0)

    def test_cuenta_espera_prolongada_con_el_mismo_umbral(self):
        self.recibida(direccion="Recién llegada", hace_dias=2)
        self.recibida(direccion="Esperando demasiado", hace_dias=10)

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["espera_prolongada"], 1)

    def test_una_visita_vencida_cuenta(self):
        self.crear(
            direccion="Visita atrasada",
            estado=EstadoEncuesta.BORRADOR,
            proxima_visita=timezone.localdate() - timedelta(days=1),
        )

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["visitas_vencidas"], 1)

    def test_una_visita_futura_no_cuenta(self):
        self.crear(
            direccion="Visita a tiempo",
            estado=EstadoEncuesta.BORRADOR,
            proxima_visita=timezone.localdate() + timedelta(days=3),
        )

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["visitas_vencidas"], 0)

    def test_sin_proxima_visita_no_cuenta(self):
        self.crear(direccion="Sin fecha", estado=EstadoEncuesta.BORRADOR)

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["visitas_vencidas"], 0)

    def test_una_encuesta_cerrada_con_visita_vencida_no_cuenta(self):
        """`proxima_visita` sobrevive al cierre, pero ya no es trabajo pendiente."""
        encuesta = self.recibida(direccion="Ya resuelta")
        self.con_hogar(encuesta, declarados=1, personas=1)
        encuesta.proxima_visita = timezone.localdate() - timedelta(days=5)
        encuesta.save(update_fields=["proxima_visita"])
        encuesta.resolver(
            EstadoEncuesta.VALIDADA, usuario=self.supervisor, comentario="Correcta."
        )

        resumen = Encuesta.resumen_alertas_calidad()

        self.assertEqual(resumen["visitas_vencidas"], 0)

    def test_el_queryset_acota_el_universo(self):
        """Pasar un queryset debe respetarlo y no mirar el resto del sistema."""
        self.recibida(direccion="Fuera del universo")
        acotado = Encuesta.objects.none()

        resumen = Encuesta.resumen_alertas_calidad(acotado)

        self.assertEqual(resumen["datos_incompletos"], 0)


class AlertasEnElPanelTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.url_panel = reverse("dashboards:supervisor")

    def test_el_panel_muestra_las_fichas_incompletas(self):
        self.recibida(direccion="Incompleta")
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["alertas_calidad"]["datos_incompletos"], 1)

    def test_el_panel_muestra_las_visitas_vencidas(self):
        self.crear(
            direccion="Atrasada",
            estado=EstadoEncuesta.BORRADOR,
            proxima_visita=timezone.localdate() - timedelta(days=2),
        )
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(respuesta.context["alertas_calidad"]["visitas_vencidas"], 1)
        self.assertContains(respuesta, "Visita vencida")

    def test_sin_alertas_el_panel_lo_dice(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertContains(respuesta, "Sin alertas")

    def test_el_umbral_de_espera_viene_del_modelo_y_no_esta_fijo_en_la_plantilla(self):
        # El texto del umbral solo se pinta cuando hay alguna alerta que mostrar;
        # sin esto el panel toma la rama «Sin alertas» y la prueba no vería nada.
        self.recibida()
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_panel)

        self.assertEqual(
            respuesta.context["dias_espera_prolongada"], Encuesta.DIAS_ESPERA_PROLONGADA
        )
        self.assertContains(
            respuesta,
            f"hace más de {Encuesta.DIAS_ESPERA_PROLONGADA} días",
        )


# ==========================================================================
# HU-18 — 75. FILTRAR POR FECHA
# ==========================================================================
#
# El resto del filtro —estado, sector, censista, y de regalo operativo y texto
# libre— ya existía desde la HU-13 (FiltrosRevisionTest, arriba). Lo único que
# faltaba era acotar por fecha, así que esta sección solo prueba el rango
# nuevo: FiltroRevisionTest cubre que sí conviva con los filtros de siempre.


class FiltroPorFechaTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)
        self.antigua = self.recibida(direccion="Antigua", hace_dias=10)
        self.reciente = self.recibida(direccion="Reciente", hace_dias=1)

    def direcciones(self, **filtros):
        respuesta = self.client.get(
            self.url_bandeja, {"estado": "TODAS", **filtros}
        )
        return [e.direccion for e in respuesta.context["encuestas"]]

    def test_desde_excluye_lo_anterior(self):
        desde = timezone.localdate() - timedelta(days=5)

        self.assertEqual(self.direcciones(fecha_desde=desde), ["Reciente"])

    def test_hasta_excluye_lo_posterior(self):
        hasta = timezone.localdate() - timedelta(days=5)

        self.assertEqual(self.direcciones(fecha_hasta=hasta), ["Antigua"])

    def test_el_rango_combina_ambos_extremos(self):
        entremedio = self.recibida(direccion="Entremedio", hace_dias=5)

        direcciones = self.direcciones(
            fecha_desde=timezone.localdate() - timedelta(days=7),
            fecha_hasta=timezone.localdate() - timedelta(days=3),
        )

        self.assertEqual(direcciones, [entremedio.direccion])

    def test_convive_con_el_filtro_de_censista(self):
        de_juan = self.recibida(
            direccion="De Juan", hace_dias=1, censista=self.juan
        )

        direcciones = self.direcciones(
            fecha_desde=timezone.localdate() - timedelta(days=5),
            censista=self.juan.pk,
        )

        self.assertEqual(direcciones, [de_juan.direccion])

    def test_un_rango_invertido_no_rompe_la_pantalla(self):
        respuesta = self.client.get(
            self.url_bandeja,
            {
                "estado": "TODAS",
                "fecha_desde": timezone.localdate(),
                "fecha_hasta": timezone.localdate() - timedelta(days=5),
            },
        )

        self.assertEqual(respuesta.status_code, 200)

    def test_un_rango_invertido_avisa_en_el_campo_y_no_filtra_nada(self):
        """Sin filtros válidos, la bandeja cae al valor por defecto: lo recibido."""
        respuesta = self.client.get(
            self.url_bandeja,
            {
                "estado": "TODAS",
                "fecha_desde": timezone.localdate(),
                "fecha_hasta": timezone.localdate() - timedelta(days=5),
            },
        )

        # Con el formulario inválido, la vista ni siquiera mira el «estado» del
        # GET: cae al valor por defecto sin filtro alguno (ambas COMPLETADA).
        self.assertContains(respuesta, "no puede ser anterior")
        direcciones = {e.direccion for e in respuesta.context["encuestas"]}
        self.assertEqual(direcciones, {"Antigua", "Reciente"})

    def test_una_encuesta_sin_cerrar_no_entra_en_ningun_rango(self):
        """Sin `cerrada_en` no hay con qué comparar: filtrar por fecha la deja fuera."""
        self.crear(direccion="A medias", estado=EstadoEncuesta.BORRADOR)

        direcciones = self.direcciones(
            fecha_desde=timezone.localdate() - timedelta(days=30)
        )

        self.assertNotIn("A medias", direcciones)


# ==========================================================================
# HU-19 — 76. REPORTES EN PDF Y EXCEL
# ==========================================================================


class ResumenParaReporteTest(BaseRevisionTest):
    def test_total_cuenta_el_queryset_recibido(self):
        self.recibida(direccion="A")
        self.crear(direccion="B", estado=EstadoEncuesta.VALIDADA)

        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.all())

        self.assertEqual(resumen["total"], 2)

    def test_incluye_la_fecha_de_generacion(self):
        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.none())

        self.assertAlmostEqual(
            resumen["generado_en"], timezone.localtime(timezone.now()),
            delta=timedelta(seconds=5),
        )

    def test_reutiliza_las_alertas_de_calidad_de_la_hu16(self):
        """No inventa un segundo criterio de "ficha con problemas"."""
        encuesta = self.recibida()

        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.all())

        self.assertEqual(
            resumen["alertas"],
            Encuesta.resumen_alertas_calidad(Encuesta.objects.all()),
        )
        self.assertTrue(encuesta.tiene_datos_incompletos)
        self.assertEqual(resumen["alertas"]["datos_incompletos"], 1)

    def test_por_estado_usa_la_etiqueta_legible_y_calcula_el_porcentaje(self):
        self.recibida(direccion="A")
        self.recibida(direccion="B")
        self.crear(direccion="C", estado=EstadoEncuesta.VALIDADA)

        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.all())

        self.assertEqual(
            resumen["por_estado"],
            [
                {"etiqueta": "Completada", "total": 2, "porcentaje": 66.7},
                {"etiqueta": "Validada", "total": 1, "porcentaje": 33.3},
            ],
        )

    def test_por_sector_agrupa_por_nombre_e_incluye_la_comuna(self):
        self.recibida()  # cae en self.boldos («Los Boldos», comuna Concepción)

        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.all())

        self.assertEqual(
            resumen["por_sector"],
            [
                {
                    "sector": "Los Boldos",
                    "comuna": "Concepción",
                    "total": 1,
                    "porcentaje": 100.0,
                }
            ],
        )

    def test_por_censista_junta_nombre_y_apellido_y_desglosa_la_resolucion(self):
        self.recibida(censista=self.juan)
        self.crear(
            direccion="Devuelta", censista=self.juan, estado=EstadoEncuesta.OBSERVADA
        )
        self.crear(
            direccion="Aprobada", censista=self.juan, estado=EstadoEncuesta.VALIDADA
        )

        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.all())

        self.assertEqual(
            resumen["por_censista"],
            [{"censista": "Juan Vera", "total": 3, "validadas": 1, "observadas": 1}],
        )

    def test_respeta_el_queryset_que_recibe(self):
        """No mira `Encuesta.objects.all()` por su cuenta: cuenta lo que le pasan."""
        self.recibida(direccion="Dentro", censista=self.marta)
        self.recibida(direccion="Fuera", censista=self.juan)

        resumen = Encuesta.resumen_para_reporte(
            Encuesta.objects.filter(censista=self.marta)
        )

        self.assertEqual(resumen["total"], 1)

    def test_un_queryset_vacio_no_rompe_nada(self):
        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.none())

        self.assertEqual(resumen["total"], 0)
        self.assertEqual(resumen["por_estado"], [])
        self.assertEqual(resumen["por_sector"], [])
        self.assertEqual(resumen["por_censista"], [])
        self.assertEqual(
            resumen["alertas"],
            {"datos_incompletos": 0, "espera_prolongada": 0, "visitas_vencidas": 0},
        )


class ConstruirReporteExcelTest(BaseRevisionTest):
    def leer(self, resumen):
        libro = construir_reporte_excel(resumen)
        buffer = BytesIO()
        libro.save(buffer)
        buffer.seek(0)
        return load_workbook(buffer)

    def test_incluye_el_resumen_general(self):
        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.none())
        resumen["total"] = 5

        celdas = [fila[0].value for fila in self.leer(resumen).active.iter_rows()]

        self.assertIn("Resumen general", celdas)
        self.assertIn("Total de encuestas", celdas)
        self.assertIn("Visitas vencidas", celdas)

    def test_incluye_los_tres_bloques_con_sus_columnas_y_datos(self):
        self.recibida(censista=self.marta)

        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.all())
        filas = list(self.leer(resumen).active.iter_rows(values_only=True))
        celdas = [fila[0] for fila in filas]

        self.assertIn("Por estado", celdas)
        self.assertIn("Por sector", celdas)
        self.assertIn("Por censista", celdas)
        self.assertIn("Marta Soto", celdas)
        # Las columnas nuevas de esta ronda: comuna en «por sector», y el
        # desglose de validadas/observadas en «por censista».
        self.assertIn(("Sector", "Comuna", "Total", "%"), filas)
        self.assertIn(("Censista", "Total", "Validadas", "Observadas"), filas)

    def test_un_bloque_sin_filas_lo_dice_en_vez_de_dejar_solo_el_encabezado(self):
        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.none())

        celdas = [fila[0].value for fila in self.leer(resumen).active.iter_rows()]

        self.assertIn("Sin datos para este filtro.", celdas)


class ConstruirReportePdfTest(BaseRevisionTest):
    def test_produce_un_pdf_valido(self):
        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.none())
        buffer = BytesIO()

        construir_reporte_pdf(buffer, resumen)

        self.assertTrue(buffer.getvalue().startswith(b"%PDF"))

    def test_un_bloque_vacio_no_lanza_una_excepcion(self):
        """Un filtro sin resultados en algún bloque no debe romper el armado."""
        resumen = Encuesta.resumen_para_reporte(Encuesta.objects.none())
        buffer = BytesIO()

        construir_reporte_pdf(buffer, resumen)  # no debe lanzar

        self.assertTrue(buffer.getvalue().startswith(b"%PDF"))


class ExportarReportesTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.url_excel = reverse("fichas:reporte_excel")
        self.url_pdf = reverse("fichas:reporte_pdf")

    def test_el_supervisor_descarga_el_excel(self):
        self.recibida()
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_excel)

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", respuesta["Content-Disposition"])

    def test_el_supervisor_descarga_el_pdf(self):
        self.recibida()
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_pdf)

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta["Content-Type"], "application/pdf")
        self.assertIn("attachment", respuesta["Content-Disposition"])

    def test_el_administrador_tambien_descarga(self):
        self.client.force_login(self.admin)

        self.assertEqual(self.client.get(self.url_excel).status_code, 200)
        self.assertEqual(self.client.get(self.url_pdf).status_code, 200)

    def test_el_censista_no_puede_exportar(self):
        self.client.force_login(self.marta)

        self.assertEqual(self.client.get(self.url_excel).status_code, 302)
        self.assertEqual(self.client.get(self.url_pdf).status_code, 302)

    def test_un_visitante_anonimo_va_al_login(self):
        respuesta = self.client.get(self.url_excel)

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(reverse("usuarios:login"), respuesta.url)

    def test_sin_ver_todas_el_supervisor_igual_puede_exportar(self):
        """`reportes.exportar` es el permiso correcto, no `fichas.ver_todas`."""
        self.rol_supervisor.permisos.remove(
            Permiso.objects.get(codigo="fichas.ver_todas")
        )
        self.client.force_login(self.supervisor)

        self.assertEqual(self.client.get(self.url_excel).status_code, 200)

    def test_sin_reportes_exportar_el_supervisor_no_puede(self):
        self.rol_supervisor.permisos.remove(
            Permiso.objects.get(codigo="reportes.exportar")
        )
        self.client.force_login(self.supervisor)

        self.assertEqual(self.client.get(self.url_excel).status_code, 302)

    def test_respeta_los_filtros_de_la_bandeja(self):
        """Lo exportado coincide con lo que la bandeja mostraría con el mismo filtro."""
        self.recibida(direccion="De Marta", censista=self.marta)
        self.recibida(direccion="De Juan", censista=self.juan)
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(
            self.url_excel, {"estado": "TODAS", "censista": self.juan.pk}
        )

        celdas = [
            fila[0].value
            for fila in load_workbook(BytesIO(respuesta.content)).active.iter_rows()
        ]

        self.assertIn("Juan Vera", celdas)
        self.assertNotIn("Marta Soto", celdas)

    def test_el_pdf_tambien_respeta_el_filtro_por_defecto(self):
        """No repite lo anterior: solo confirma que el PDF sigue el mismo camino."""
        self.recibida(direccion="Enviada")
        self.crear(direccion="Sin enviar")  # PENDIENTE: no debería contarse
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(self.url_pdf)

        self.assertEqual(respuesta.status_code, 200)
        self.assertTrue(respuesta.content.startswith(b"%PDF"))

    def test_un_rango_de_fecha_invertido_no_rompe_la_exportacion(self):
        self.client.force_login(self.supervisor)

        respuesta = self.client.get(
            self.url_excel,
            {
                "fecha_desde": timezone.localdate(),
                "fecha_hasta": timezone.localdate() - timedelta(days=5),
            },
        )

        self.assertEqual(respuesta.status_code, 200)


class BotonesDeExportarTest(BaseRevisionTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.supervisor)

    def test_el_supervisor_ve_los_botones(self):
        respuesta = self.client.get(self.url_bandeja)

        self.assertContains(respuesta, "Exportar Excel")
        self.assertContains(respuesta, "Exportar PDF")

    def test_sin_reportes_exportar_no_se_ven_los_botones(self):
        self.rol_supervisor.permisos.remove(
            Permiso.objects.get(codigo="reportes.exportar")
        )

        respuesta = self.client.get(self.url_bandeja)

        self.assertNotContains(respuesta, "Exportar Excel")

    def test_los_enlaces_llevan_los_filtros_activos(self):
        """Exportar es «descargar lo que estoy viendo», no una pantalla aparte."""
        respuesta = self.client.get(self.url_bandeja, {"censista": self.juan.pk})

        self.assertContains(respuesta, f"reporte.xlsx?censista={self.juan.pk}")
        self.assertContains(respuesta, f"reporte.pdf?censista={self.juan.pk}")
